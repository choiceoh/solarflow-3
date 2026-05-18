# -*- coding: utf-8 -*-
"""면장/수입신고필증 + 납부고지서 PDF → 정형 데이터 추출 → 시트 22 추가.

추출 양식:
1. 수입신고필증 (면장) — 신고번호/신고일/입항일/BL/수입자/CIF/환율/관세/부가세/총세액 등
2. 납부고지서 — BL/수입신고번호/관세/부가세/합계금액

cost_details.customs_fee (현재 0/100) 백필의 1순위 소스.
"""
from __future__ import annotations

import os
import re
import sys
import json
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import pdfplumber

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
TARGET = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')

PDF_DIRS = [
    (os.path.join(ROOT, '2024년 모듈발주', '수입면장'), '2024'),
    (os.path.join(ROOT, '2024년 모듈발주', '수입면장', '수입면장'), '2024'),
    (os.path.join(ROOT, '2025년 모듈 발주', '수입신고필증'), '2025'),
    (os.path.join(ROOT, '2026년 모듈 발주', '수입면장'), '2026'),
]

# ---------- 스타일 ----------
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='1F4E78')
NOTE_FONT = Font(name='맑은 고딕', italic=True, size=9, color='666666')
BODY_FONT = Font(name='맑은 고딕', size=10)
BORDER_THIN = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))


# ---------- 파싱 헬퍼 ----------
def num(s):
    if s is None:
        return None
    s = s.replace(',', '').strip()
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def search1(pat, text, group=1, flags=0):
    m = re.search(pat, text, flags)
    return m.group(group).strip() if m else None


def is_payment_notice(text):
    return '납부영수증서' in text or '납부고지서' in text or '납부서[수납기관용]' in text


def is_declaration(text):
    return '수입신고필증' in text or '신고필증' in text


def parse_declaration(text):
    """수입신고필증 (면장) → dict.

    핵심 필드만. 일부 PDF 는 layout 차이로 NULL 가능 — 베스트-에포트.
    """
    d = {}
    d['declaration_no'] = search1(r'(\d{4,5}-\d{2}-\d{6,7}M)', text)
    # 신고일: 라인 '2 신고일' 다음의 첫 yyyy/mm/dd
    d['decl_date'] = search1(r'(\d{4}/\d{2}/\d{2})', text)  # 보통 첫 등장이 신고일

    dates = re.findall(r'\d{4}/\d{2}/\d{2}', text)
    d['arrival_date'] = dates[1] if len(dates) >= 2 else None  # 입항일
    d['release_date'] = dates[-1] if dates else None  # 수리일자 (마지막)

    # B/L 번호
    d['bl_number'] = search1(r'B/L\(AWB\)번호.*?\n([A-Z0-9]{6,30})', text, flags=re.DOTALL) \
        or search1(r'(JWSH\d+|SNK[oO]03[A-Z0-9]+|HDMUSHAA\d+|SHKWA\d+|EASE[KD]\w+|EASHO\w+|SELHTZ\w+|SELYIT\d+|SHACYV\w+|SHACYR\w+|NPSELHT\d+|LS\d+|RSPN\d+|JAHF\d+|MCKRJH\w+|TMSHKPTP\d+|DFS\d+|SHADFC\w+|ESZX\d+|HDMU\w+)', text)
    # 수입자 (회사명) — '11 수 입 자' 라인
    d['importer'] = search1(r'수\s*입\s*자\s+([가-힣()㈜0-9A-Za-z\s\.]+?)\(', text)
    if not d['importer']:
        d['importer'] = search1(r'(탑솔라\(주\)|디원|화신이엔지|\(주\)디원)', text)
    # 사업자번호
    d['biz_no'] = search1(r'(\d{3}-\d{2}-\d{5})', text)
    # 운송주선인 (포워더)
    d['forwarder'] = search1(r'운송주선인\s+(.+?)(?:\d+종류|18종류|\n)', text)
    # 무역거래처 (제조사)
    d['trade_partner'] = search1(r'무역거래처\s+(.+?)(?:\d{2}MASTER|\n)', text)
    # MASTER B/L
    d['master_bl'] = search1(r'MASTER B/L번호\s+([A-Z0-9]+)', text)
    # 적출국
    d['origin_country'] = search1(r'적출국\s+([A-Z]{2})', text)
    # 도착항
    d['arrival_port'] = search1(r'국내도착항\s+([A-Z]{4,6})', text)
    # 모델/규격
    d['model'] = search1(r'(JKM\d+\w*-\w+|LR7-\w+|TSM-\w+|RSM\d+\w*-\w+|JAM\w+|JAHF\d+|JCM\d+|JC\d+\w*)', text)
    # Wp
    d['wp'] = search1(r'\((\d{3,4})W\)', text) or search1(r'(\d{3,4})W\b', text)
    # PCS
    d['pcs'] = num(search1(r'\(([\d,]+)\s*PCS?\)', text))
    if d['pcs'] is None:
        d['pcs'] = num(search1(r'([\d,]+)\s*PC\b', text))
    # 단가 USD
    d['unit_price_usd'] = num(search1(r'단가\(USD\).*?(0\.\d+)', text, flags=re.DOTALL))
    # 금액 USD (37금액)
    # 과세가격(CIF) USD/KRW
    d['cif_usd'] = num(search1(r'\$\s*([\d,]+(?:\.\d+)?)', text))  # 첫 $ 가 과세가격
    d['cif_krw'] = num(search1(r'￦\s*([\d,]+)', text))
    # 환율
    d['exchange_rate'] = num(search1(r'환\s*율\s+([\d,]+\.\d+)', text))
    # 관세 (61세종/62세액 다음 '관 세' 라인)
    customs_match = re.search(r'관\s*세\s+([\d,]+)\s*\n.*?개별소비세', text, re.DOTALL)
    d['customs_duty'] = num(customs_match.group(1)) if customs_match else None
    # 부가가치세
    d['vat'] = num(search1(r'부가가치세\s+([\d,]+)', text))
    # 총세액합계
    d['total_tax'] = num(search1(r'총세액합계\s+([\d,]+)', text))
    # 발행번호
    d['issue_no'] = search1(r'발\s*행\s*번\s*호\s*[:：]\s*([\d]+)', text)
    # L/C No
    d['lc_no'] = search1(r'L/C[\s\.]*NO[\s\.：:]*([A-Z0-9]+)', text, flags=re.IGNORECASE)
    return d


def parse_payment_notice(text):
    """납부고지서 → dict."""
    d = {}
    # B/L No (납부고지서 시리얼)
    d['notice_bl'] = search1(r'B/L\s*No\.?\s*[:：]\s*(\w+)', text)
    # 수입신고번호
    d['declaration_no'] = search1(r'수입신고번호\s+([\d]{4,5}-\d{2}-\d{6,7}M)', text)
    # 수입자 (상 호)
    d['importer'] = search1(r'상\s*호\s+(.+?)(?:\n|주\s*소)', text)
    # 사업번호
    d['biz_no'] = search1(r'사업번호\s*[:：]\s*(\d{3}-\d{2}-\d{5})', text)
    # 발행일자
    d['issue_date'] = search1(r'발행일자\s+(\d{4}년\d{2}월\d{2}일)', text)
    # 납부기한
    d['due_date'] = search1(r'납부기한\s+(\d{4}년\d{2}월\d{2}일)', text)
    # 관세
    d['customs_duty'] = num(search1(r'관\s*세\s+([\d,]+)', text))
    # 부가가치세
    d['vat'] = num(search1(r'부가가치세\s+([\d,]+)', text))
    # 합계 금액 (납기내 금액)
    d['amount_total'] = num(search1(r'납기내\s*\n.*?([\d,]{4,})', text, flags=re.DOTALL))
    # 수입징수관서
    d['customs_office'] = search1(r'수입징수관서\s+(\S+세관)', text)
    return d


# ---------- PDF 일괄 처리 ----------
def extract_one(args):
    full, year = args
    name = os.path.basename(full)
    try:
        with pdfplumber.open(full) as pdf:
            text = ''
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    text += t + '\n'
        if not text.strip():
            return {'year': year, 'file': name, 'kind': 'EMPTY', 'error': 'no text (scan?)'}
        kind = 'PAYMENT' if is_payment_notice(text) else ('DECLARATION' if is_declaration(text) else 'OTHER')
        if kind == 'DECLARATION':
            data = parse_declaration(text)
        elif kind == 'PAYMENT':
            data = parse_payment_notice(text)
        else:
            data = {'note': text[:200]}
        data['year'] = year
        data['file'] = name
        data['kind'] = kind
        return data
    except Exception as e:
        return {'year': year, 'file': name, 'kind': 'ERROR', 'error': str(e)[:200]}


def collect_pdfs():
    pdfs = []
    seen = set()
    for folder, year in PDF_DIRS:
        if not os.path.isdir(folder):
            continue
        for f in sorted(os.listdir(folder)):
            full = os.path.join(folder, f)
            if os.path.isfile(full) and f.lower().endswith('.pdf') and full not in seen:
                pdfs.append((full, year))
                seen.add(full)
    return pdfs


print('PDF 수집 ...')
pdfs = collect_pdfs()
print(f'총 {len(pdfs)} PDF 처리 시작')

results = []
for i, args in enumerate(pdfs):
    if i % 20 == 0:
        print(f'  {i}/{len(pdfs)} ...')
    results.append(extract_one(args))

print(f'추출 완료: {len(results)}건')

# 종류별 통계
kinds = {}
for r in results:
    k = r.get('kind', '?')
    kinds[k] = kinds.get(k, 0) + 1
print(f'종류 분포: {kinds}')

# 면장 추출 성공 카운트
decl_results = [r for r in results if r.get('kind') == 'DECLARATION']
pay_results = [r for r in results if r.get('kind') == 'PAYMENT']
print(f'면장 (DECLARATION): {len(decl_results)}, 납부고지서 (PAYMENT): {len(pay_results)}')

# 관세/부가세 추출 성공률
decl_with_customs = sum(1 for r in decl_results if r.get('customs_duty') is not None)
decl_with_vat = sum(1 for r in decl_results if r.get('vat') is not None)
print(f'면장 - 관세 추출: {decl_with_customs}/{len(decl_results)}, 부가세 추출: {decl_with_vat}/{len(decl_results)}')

# JSON 백업
backup = os.path.join(os.path.dirname(__file__), '..', 'pdf_extract_backup.json')
with open(backup, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)
print(f'JSON 백업: {backup}')

# ---------- xlsx 시트 22 추가 ----------
print('\nxlsx 갱신 ...')
wb = load_workbook(TARGET)

# 시트 22 (면장 정본)
S22 = '22. 면장 정본 데이터 (PDF 추출)'
if S22 in wb.sheetnames:
    del wb[S22]
ws = wb.create_sheet(S22)

ws['A1'] = '면장 정본 데이터 — 수입신고필증 + 납부고지서 PDF 추출'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:V1')
ws['A2'] = (f'추출 결과: 총 {len(results)} PDF / 면장 {len(decl_results)} / 납부고지서 {len(pay_results)} / '
            f'관세 추출 {decl_with_customs}/{len(decl_results)} = {100*decl_with_customs/max(1,len(decl_results)):.0f}% '
            f'(cost_details.customs_fee 백필 1순위 소스)')
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:V2')

# === 섹션 1. 면장 정본 ===
ws.cell(row=4, column=1, value='1. 수입신고필증 (면장) 정본 — 관세/부가세/CIF 직접 추출').font = SECTION_FONT
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=22)

decl_cols = [
    ('연도', 'year', 6),
    ('PDF 파일', 'file', 38),
    ('신고번호', 'declaration_no', 18),
    ('신고일', 'decl_date', 12),
    ('입항일', 'arrival_date', 12),
    ('수리일자', 'release_date', 12),
    ('B/L 번호', 'bl_number', 22),
    ('MASTER B/L', 'master_bl', 18),
    ('수입자', 'importer', 14),
    ('사업자번호', 'biz_no', 14),
    ('포워더', 'forwarder', 24),
    ('무역거래처', 'trade_partner', 28),
    ('적출국', 'origin_country', 8),
    ('도착항', 'arrival_port', 8),
    ('모델', 'model', 22),
    ('Wp', 'wp', 6),
    ('PCS', 'pcs', 10),
    ('단가($)', 'unit_price_usd', 10),
    ('CIF($)', 'cif_usd', 14),
    ('CIF(₩)', 'cif_krw', 16),
    ('환율', 'exchange_rate', 10),
    ('관세(₩)', 'customs_duty', 14),
    ('부가세(₩)', 'vat', 14),
    ('총세액(₩)', 'total_tax', 14),
    ('L/C No', 'lc_no', 22),
]
hdr_row = 5
for j, (h, _, _) in enumerate(decl_cols):
    c = ws.cell(row=hdr_row, column=j + 1, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER_THIN
cur = hdr_row + 1
for r in decl_results:
    for j, (_, key, _) in enumerate(decl_cols):
        val = r.get(key)
        c = ws.cell(row=cur, column=j + 1, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER_THIN
        # 숫자 포맷
        if key in ('cif_krw', 'customs_duty', 'vat', 'total_tax', 'pcs') and isinstance(val, (int, float)):
            c.number_format = '#,##0'
        elif key in ('cif_usd', 'unit_price_usd', 'exchange_rate') and isinstance(val, (int, float)):
            c.number_format = '#,##0.0000'
    cur += 1

# === 섹션 2. 납부고지서 정본 ===
cur += 1
ws.cell(row=cur, column=1, value='2. 납부고지서 — 관세/부가세 합계 (cost_details.customs_fee 직접 매칭 가능)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=22)
cur += 1

pay_cols = [
    ('연도', 'year', 6),
    ('PDF 파일', 'file', 50),
    ('신고번호', 'declaration_no', 20),
    ('수입자', 'importer', 14),
    ('사업번호', 'biz_no', 14),
    ('발행일자', 'issue_date', 14),
    ('납부기한', 'due_date', 14),
    ('관세(₩)', 'customs_duty', 14),
    ('부가세(₩)', 'vat', 14),
    ('합계금액(₩)', 'amount_total', 14),
    ('수입징수관서', 'customs_office', 18),
    ('Notice B/L', 'notice_bl', 16),
]
for j, (h, _, _) in enumerate(pay_cols):
    c = ws.cell(row=cur, column=j + 1, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER_THIN
cur += 1
for r in pay_results:
    for j, (_, key, _) in enumerate(pay_cols):
        val = r.get(key)
        c = ws.cell(row=cur, column=j + 1, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER_THIN
        if key in ('customs_duty', 'vat', 'amount_total') and isinstance(val, (int, float)):
            c.number_format = '#,##0'
    cur += 1

# === 섹션 3. 추출 실패/기타 ===
errors = [r for r in results if r.get('kind') in ('OTHER', 'EMPTY', 'ERROR')]
if errors:
    cur += 1
    ws.cell(row=cur, column=1, value='3. 추출 실패 / 기타 양식').font = SECTION_FONT
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=22)
    cur += 1
    err_hdr = ['연도', 'PDF 파일', '종류', '비고']
    for j, h in enumerate(err_hdr):
        c = ws.cell(row=cur, column=j + 1, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cur += 1
    for r in errors:
        vals = [r.get('year'), r.get('file'), r.get('kind'), r.get('error') or r.get('note', '')[:100]]
        for j, v in enumerate(vals):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
        cur += 1

# 컬럼 너비
for j, (_, _, w) in enumerate(decl_cols):
    ws.column_dimensions[get_column_letter(j + 1)].width = w
ws.freeze_panes = 'A6'

# ---------- 보조: 카탈로그 P entry 갱신 (백필 상태 강조) ----------
ws_cat = wb['2. 자료 카탈로그']
for r in range(1, ws_cat.max_row + 1):
    if ws_cat.cell(row=r, column=1).value == 'P':
        ws_cat.cell(row=r, column=6, value=f'🔥 PDF 추출 완료 — 면장 {len(decl_results)} / 납부고지서 {len(pay_results)} (시트 22) — customs_fee 백필 1순위')
        break

# ---------- 보조: 대시보드 갱신 ----------
ws_d = wb['0. 대시보드']
for r in range(1, ws_d.max_row + 1):
    v = ws_d.cell(row=r, column=1).value
    if v == 'P':
        ws_d.cell(row=r, column=4,
                  value=f'🔥 시트 21 메타 + 시트 22 정본 데이터 (관세/부가세/CIF/환율/모델/PCS — '
                        f'{len(decl_results)}건 면장 + {len(pay_results)}건 납부고지서)')

# 백필 후보 M132 (customs_fee) 의 소스란을 강화
for r in range(1, ws_d.max_row + 1):
    v = ws_d.cell(row=r, column=1).value
    if v == 'M132 (제안)':
        ws_d.cell(row=r, column=4, value='시트 22 면장 정본 (PDF 추출 — 관세 / 부가세 직접 보유) + D 회계전표')
        ws_d.cell(row=r, column=5, value='🔥🔥 PDF 정본 추출 완료 — 즉시 백필 가능')
        break

# ---------- README 갱신 (시트 22 안내 추가) ----------
ws_r = wb['1. README']
# 시트 21 안내 다음 행에 시트 22 추가
inserted = False
for r in range(1, ws_r.max_row + 1):
    v = ws_r.cell(row=r, column=1).value
    if v and '21. (NEW) P. 면장·기타 PDF 인벤토리' in str(v):
        ws_r.insert_rows(r + 1)
        ws_r.cell(row=r + 1, column=1, value=' 22. (NEW) 면장 정본 데이터 — PDF 텍스트 추출 결과 (관세/부가세/CIF/환율/모델 정형 데이터)').font = BODY_FONT
        inserted = True
        break

# 시트 순서 재정렬
def sort_key(name):
    try:
        n = int(name.split('.')[0])
        return (0, n)
    except ValueError:
        return (1, name)

sorted_names = sorted(wb.sheetnames, key=sort_key)
wb._sheets = [wb[n] for n in sorted_names]

wb.save(TARGET)
print(f'\n저장 완료: {TARGET}')

wb2 = load_workbook(TARGET, read_only=True)
print(f'최종 시트 ({len(wb2.sheetnames)}개):')
for i, n in enumerate(wb2.sheetnames):
    print(f'  {i:2d}. {n}')
sz = os.path.getsize(TARGET)
print(f'파일 크기: {sz:,}B ({sz/1024:.1f}KB)')
wb2.close()
