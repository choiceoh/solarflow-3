# -*- coding: utf-8 -*-
"""추가 흡수: D 회계 raw / C BL 출고 raw / 운송료 청구서 합계 / 소형 보조 xlsx 통합.

시트 23, 24, 25, 26 추가. 카탈로그/대시보드/README 갱신.
"""
from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
TARGET = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')

HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='1F4E78')
NOTE_FONT = Font(name='맑은 고딕', italic=True, size=9, color='666666')
BODY_FONT = Font(name='맑은 고딕', size=10)
BORDER_THIN = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))


def safe_set(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == 'MergedCell':
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row <= row <= mr.max_row) and (mr.min_col <= col <= mr.max_col):
                ws.unmerge_cells(str(mr))
                ws.cell(row=mr.min_row, column=mr.min_col, value=value)
                return
    cell.value = value


def write_header_row(ws, row, headers):
    for j, v in enumerate(headers):
        c = ws.cell(row=row, column=j + 1, value=v)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER_THIN


def append_rows(ws, start_row, rows, num_cols=None, date_cols=(), int_cols=(), float_cols=()):
    cur = start_row
    for row in rows:
        if num_cols:
            row = list(row)[:num_cols] + [None] * max(0, num_cols - len(row))
        for j, v in enumerate(row):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if j in date_cols and isinstance(v, datetime):
                c.number_format = 'yyyy-mm-dd'
            elif j in int_cols and isinstance(v, (int, float)):
                c.number_format = '#,##0'
            elif j in float_cols and isinstance(v, (int, float)):
                c.number_format = '#,##0.0000'
        cur += 1
    return cur


# ---------- 1. wb 로드 ----------
print(f'Loading {TARGET}')
wb = load_workbook(TARGET)
print(f'기존 시트: {len(wb.sheetnames)}')

# ============================================================
# 시트 23: D 회계 전표 raw (304행, 6 거래처×년도 시트 통합)
# ============================================================
S23 = '23. D-회계 전표 raw (부대비용·운송료)'
if S23 in wb.sheetnames:
    del wb[S23]
ws = wb.create_sheet(S23)
ws['A1'] = 'D. 부대비용·운송료 회계 전표 raw — 6 시트 (선진/블루오션/스마일 × 25/26) 통합 (304행)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:N1')
ws['A2'] = '원본: Dropbox/2025년 운송료/2025년, 2026년 모듈 부대비용, 운송료 내역.xlsx (오늘 갱신)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:N2')

hdr_d = ['거래처 시트', '거래처코드', '거래처명', '회계단위코드', '회계단위명', '계정과목명',
         '승인일', '승인번호', '적요', '차변(₩)', '대변(₩)', '잔액', '전표번호', 'BL 추출']
write_header_row(ws, 4, hdr_d)

p_d = os.path.join(ROOT, '2025년 운송료', '2025년, 2026년 모듈 부대비용, 운송료 내역.xlsx')
src_wb = load_workbook(p_d, read_only=True, data_only=True)

bl_re = re.compile(r'(SHACYV\w+|SHACYR\w+|SNK[oO]03[A-Z0-9]+|JWSH\d+|HDMUSHAA\d+|SHKWA\d+|EASE[KD]\w+|EASHO\w+|SELHTZ\w+|SELYIT\d+|NPSELHT\d+|LS\d+|RSPN\d+|JAHF\d+|MCKRJH\w+|TMSHKPTP\d+|DFS\d+|SHADFC\w+|ESZX\d+|HDMU\w+)')

cur = 5
total_credit_25 = 0
total_credit_26 = 0
for sheet_name in src_wb.sheetnames:
    sws = src_wb[sheet_name]
    rows = list(sws.iter_rows(values_only=True))
    # 헤더 row 1 (개념: r0=타이틀, r1=헤더, r2=서브헤더, r3+=data)
    for i, r in enumerate(rows):
        if i < 3:
            continue
        if all(v is None for v in r[:13]):
            continue
        # 적요에서 BL 추출
        memo = r[7] if len(r) > 7 and r[7] else ''
        bl = ''
        if memo:
            m = bl_re.search(str(memo))
            bl = m.group(1) if m else ''
        # 컬럼: [0]거래처코드 [1]거래처명 [2]회계단위코드 [3]회계단위명 [4]계정과목명 [5]승인일 [6]승인번호 [7]적요 [8]차변 [9]대변 [10]잔액 [11]기준잔액 [12]전표번호
        out = [sheet_name, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[11], r[12], bl]
        for j, v in enumerate(out):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if j == 6 and isinstance(v, datetime):
                c.number_format = 'yyyy-mm-dd'
            elif j in (9, 10, 11) and isinstance(v, (int, float)):
                c.number_format = '#,##0'
        # 합계 누적
        if isinstance(r[9], (int, float)):
            if '25년' in sheet_name or '(25년)' in sheet_name:
                total_credit_25 += r[9]
            elif '26년' in sheet_name or '(26년)' in sheet_name:
                total_credit_26 += r[9]
        cur += 1
src_wb.close()

# 합계행
sum_row = cur
ws.cell(row=sum_row, column=1, value='25년 합계').font = HEADER_FONT
ws.cell(row=sum_row, column=1).fill = HEADER_FILL
ws.cell(row=sum_row, column=10, value=total_credit_25).number_format = '#,##0'
ws.cell(row=sum_row, column=10).font = HEADER_FONT
ws.cell(row=sum_row, column=10).fill = HEADER_FILL
sum_row += 1
ws.cell(row=sum_row, column=1, value='26년 합계').font = HEADER_FONT
ws.cell(row=sum_row, column=1).fill = HEADER_FILL
ws.cell(row=sum_row, column=10, value=total_credit_26).number_format = '#,##0'
ws.cell(row=sum_row, column=10).font = HEADER_FONT
ws.cell(row=sum_row, column=10).fill = HEADER_FILL

# 너비
widths = [22, 12, 22, 12, 14, 14, 12, 10, 50, 14, 14, 14, 22, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'
print(f'시트 [{S23}] 작성 완료 — {cur - 5}행, 25년 합 ₩{total_credit_25:,.0f}, 26년 합 ₩{total_credit_26:,.0f}')


# ============================================================
# 시트 24: C BL별 출고현황 raw (2947행, 6 제조사 시트 통합)
# ============================================================
S24 = '24. C-BL별 출고현황 raw (제조사 통합)'
if S24 in wb.sheetnames:
    del wb[S24]
ws = wb.create_sheet(S24)
ws['A1'] = 'C. BL별 출고현황리스트 raw — 6 제조사 시트 (진코/JA/트리나/라이젠/론지/진코(2)) 통합'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:Q1')
ws['A2'] = '원본: Dropbox/(수입) BL별 출고현황리스트.xlsx — 운영자 수기 정본 (M111~M115 백필 1차 소스, outbound_bl_items 정본)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:Q2')
ws['A3'] = '컬럼 구조 시트마다 다름 — 제조사 + 시트별 raw 첫 16컬럼 통합. 빈 행 스킵.'
ws['A3'].font = NOTE_FONT
ws.merge_cells('A3:Q3')

hdr_c = ['제조사', '발주처', 'B/L', '항구', '포워더', 'ETD', 'ETA',
         '모델명', '모듈수량', '용량(kW)', '출고일', '출고지', '지역', 'WP', '출고수량', '비고']
write_header_row(ws, 5, hdr_c)

p_c = os.path.join(ROOT, '(수입) BL별 출고현황리스트.xlsx')
src_wb = load_workbook(p_c, read_only=True, data_only=True)
cur = 6
for sheet_name in src_wb.sheetnames:
    sws = src_wb[sheet_name]
    # 각 시트 헤더 위치 다름 — 첫 비-None 행으로 추정. 보수적으로 row 1-3 스킵.
    # raw row 그대로 가져와서 첫 15 컬럼만
    for i, r in enumerate(sws.iter_rows(values_only=True)):
        if i < 2:  # 헤더 건너뜀 (보통 r0/r1 = 헤더)
            continue
        # 다 비어있으면 skip
        relevant = r[:15] if len(r) > 0 else []
        if all(v is None for v in relevant):
            continue
        out = [sheet_name] + list(r[:15])
        # row pad
        out = (out + [None] * 16)[:16]
        for j, v in enumerate(out):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if j in (5, 6, 10) and isinstance(v, datetime):
                c.number_format = 'yyyy-mm-dd'
            elif j in (8, 9, 14) and isinstance(v, (int, float)):
                c.number_format = '#,##0'
        cur += 1
src_wb.close()

widths = [12, 14, 22, 12, 16, 12, 12, 26, 12, 12, 12, 22, 12, 8, 12, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A6'
print(f'시트 [{S24}] 작성 완료 — {cur - 6}행')


# ============================================================
# 시트 25: 운송료 청구서 합계 추출
# ============================================================
S25 = '25. 운송료 청구서 합계 (xlsx 추출)'
if S25 in wb.sheetnames:
    del wb[S25]
ws = wb.create_sheet(S25)
ws['A1'] = '운송료 청구서 xlsx — BL × 거래처 × 청구금액 자동 추출 (incidental_expenses 백필 정본)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:I1')
ws['A2'] = '원본: 25년 운송료 폴더 (~17 청구서) + 26년 운송료 청구자료/{월별}/* — 청구서 시트의 합계 라인 추출'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:I2')

hdr_inv = ['연도', '폴더', '파일명', 'BL 추정', '거래처 추정', '공급가(₩)', '부가세(₩)', '합계(₩)', '추출 노트']
write_header_row(ws, 4, hdr_inv)

def detect_vendor(name):
    n = name.lower()
    if '블루' in name or '블루오션' in name or 'bluo' in n:
        return '블루오션에어'
    if '선진' in name:
        return '선진로지스틱스'
    if '스마일' in name:
        return '스마일로지스'
    return ''


def extract_invoice_total(p):
    """운송료 청구서 xlsx 에서 합계 라인 추출 — '합계' 키워드 다음 셀에 3개 숫자."""
    try:
        wb_inv = load_workbook(p, read_only=True, data_only=True)
        for sn in wb_inv.sheetnames:
            ws_inv = wb_inv[sn]
            for r in ws_inv.iter_rows(values_only=True):
                # 합계 라인 패턴: ['합계', supply, vat, total, ...] 혹은 마지막 숫자 3개
                row_vals = list(r)
                for i, v in enumerate(row_vals):
                    if v == '합계' or v == '합 계':
                        # 그 다음 3개 숫자 추출
                        nums = [x for x in row_vals[i+1:i+8] if isinstance(x, (int, float))]
                        if len(nums) >= 3:
                            wb_inv.close()
                            return nums[0], nums[1], nums[2], f'{sn} 합계 라인'
                        elif len(nums) == 1:
                            wb_inv.close()
                            return None, None, nums[0], f'{sn} 합계 단일'
        wb_inv.close()
        return None, None, None, '합계 라인 미발견'
    except Exception as e:
        return None, None, None, f'ERROR: {str(e)[:60]}'


bl_re_inv = re.compile(r'(SHACYV\w+|SHACYR\w+|SNK[oO]03[A-Z0-9]+|JWSH\d+|HDMUSHAA\d+|SHKWA\d+|EASE[KD]\w+|EASHO\w+|SELHTZ\w+|SELYIT\d+|NPSELHT\d+|LS\d+|RSPN\d+|JAHF\d+|MCKRJH\w+|TMSHKPTP\d+|DFS\d+|SHADFC\w+|ESZX\d+)')

# 25년 운송료 폴더
folder_25 = os.path.join(ROOT, '2025년 운송료')
folder_26 = os.path.join(ROOT, '2026년 모듈 발주', '운송료 청구자료')

cur = 5
total_supply = 0
total_vat = 0
total_amt = 0
processed = 0

# 25년 폴더 직속 xlsx
if os.path.isdir(folder_25):
    for f in sorted(os.listdir(folder_25)):
        full = os.path.join(folder_25, f)
        if not os.path.isfile(full) or not f.lower().endswith('.xlsx') or f.startswith('~$'):
            continue
        if '부대비용' in f:  # D 회계 전표는 별도
            continue
        bl = ''
        m = bl_re_inv.search(f)
        if m: bl = m.group(1)
        vendor = detect_vendor(f)
        supply, vat_v, total_v, note = extract_invoice_total(full)
        out = ['2025', '(루트)', f, bl, vendor, supply, vat_v, total_v, note]
        for j, v in enumerate(out):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if j in (5, 6, 7) and isinstance(v, (int, float)):
                c.number_format = '#,##0'
        if isinstance(supply, (int, float)): total_supply += supply
        if isinstance(vat_v, (int, float)): total_vat += vat_v
        if isinstance(total_v, (int, float)): total_amt += total_v
        cur += 1
        processed += 1

# 26년 운송료 청구자료/{월별}/* xlsx
if os.path.isdir(folder_26):
    for sub in sorted(os.listdir(folder_26)):
        sub_full = os.path.join(folder_26, sub)
        if not os.path.isdir(sub_full):
            continue
        for f in sorted(os.listdir(sub_full)):
            full = os.path.join(sub_full, f)
            if not os.path.isfile(full) or not f.lower().endswith('.xlsx') or f.startswith('~$'):
                continue
            bl = ''
            m = bl_re_inv.search(f)
            if m: bl = m.group(1)
            vendor = detect_vendor(sub) or detect_vendor(f)
            supply, vat_v, total_v, note = extract_invoice_total(full)
            out = ['2026', sub, f, bl, vendor, supply, vat_v, total_v, note]
            for j, v in enumerate(out):
                c = ws.cell(row=cur, column=j + 1, value=v)
                c.font = BODY_FONT
                c.alignment = Alignment(vertical='center', wrap_text=True)
                c.border = BORDER_THIN
                if j in (5, 6, 7) and isinstance(v, (int, float)):
                    c.number_format = '#,##0'
            if isinstance(supply, (int, float)): total_supply += supply
            if isinstance(vat_v, (int, float)): total_vat += vat_v
            if isinstance(total_v, (int, float)): total_amt += total_v
            cur += 1
            processed += 1

# 합계 라인
sum_row = cur + 1
ws.cell(row=sum_row, column=1, value='합계').font = HEADER_FONT
ws.cell(row=sum_row, column=1).fill = HEADER_FILL
ws.cell(row=sum_row, column=6, value=total_supply).number_format = '#,##0'
ws.cell(row=sum_row, column=6).font = HEADER_FONT
ws.cell(row=sum_row, column=6).fill = HEADER_FILL
ws.cell(row=sum_row, column=7, value=total_vat).number_format = '#,##0'
ws.cell(row=sum_row, column=7).font = HEADER_FONT
ws.cell(row=sum_row, column=7).fill = HEADER_FILL
ws.cell(row=sum_row, column=8, value=total_amt).number_format = '#,##0'
ws.cell(row=sum_row, column=8).font = HEADER_FONT
ws.cell(row=sum_row, column=8).fill = HEADER_FILL

widths = [8, 26, 50, 22, 18, 14, 14, 14, 28]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'
print(f'시트 [{S25}] 작성 완료 — {processed}건, 합계 ₩{total_amt:,.0f}')


# ============================================================
# 시트 26: 소형 보조 xlsx 9개 통합
# ============================================================
S26 = '26. 보조 xlsx 통합 (PO·정산·재고·면장리스트)'
if S26 in wb.sheetnames:
    del wb[S26]
ws = wb.create_sheet(S26)
ws['A1'] = '보조 xlsx 9개 통합 — PO정리/론지정산/JA솔라/KNK/항만재고/면장리스트/포워더 견적'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:M1')
ws['A2'] = '각 시트별 raw 흡수 (헤더 + 데이터). 큰 시트는 첫 30행만.'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:M2')

aux_files = [
    ('2025년 모듈 발주/JA솔라 모듈/JA솔라.xlsx', 'JA솔라 25년 발주'),
    ('2025년 모듈 발주/JA솔라 모듈/KNK/개설자료.xlsx', 'KNK 개설자료 25'),
    ('2025년 모듈 발주/JA솔라 모듈/KNK/발주 계산.xlsx', 'KNK 발주 계산 25'),
    ('2025년 모듈 발주/론지솔라/발주서 내용 정리.xlsx', '론지 발주서 정리 25'),
    ('2025년 모듈 발주/징코모듈/PO정리.xlsx', '징코 PO 정리 25'),
    ('2024년 모듈발주/론지모듈/론지정산.xlsx', '론지 정산 24 (T/T 0.1/0.9)'),
    ('2024년 모듈발주/포워더 견적비교(제주).xlsx', '포워더 견적비교 (제주向)'),
    ('2024년 모듈발주/2024년 재고/항만재고현황.xlsx', '항만재고현황 24'),
    ('2024년 모듈발주/수입면장/수입면장/수입신고필증 리스트.xlsx', '24년 면장 리스트'),
]

cur = 4
for rel, label in aux_files:
    p = os.path.join(ROOT, rel)
    if not os.path.isfile(p):
        continue
    # 섹션 타이틀
    ws.cell(row=cur, column=1, value=f'■ {label}').font = SECTION_FONT
    ws.cell(row=cur, column=2, value=f'경로: {rel}').font = NOTE_FONT
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=13)
    cur += 1
    try:
        src_wb = load_workbook(p, read_only=True, data_only=True)
        for sn in src_wb.sheetnames:
            sws = src_wb[sn]
            ws.cell(row=cur, column=1, value=f'  └ 시트: {sn} ({sws.max_row}r×{sws.max_column}c)').font = NOTE_FONT
            cur += 1
            for i, r in enumerate(sws.iter_rows(values_only=True)):
                if i >= 35:  # 시트당 max 35행
                    ws.cell(row=cur, column=1, value=f'  ... ({sws.max_row - 35} more)').font = NOTE_FONT
                    cur += 1
                    break
                if all(v is None for v in r):
                    continue
                # 첫 13컬럼만
                for j, v in enumerate(list(r)[:13]):
                    c = ws.cell(row=cur, column=j + 1, value=v)
                    c.font = BODY_FONT
                    c.alignment = Alignment(vertical='center', wrap_text=True)
                    if isinstance(v, datetime):
                        c.number_format = 'yyyy-mm-dd'
                    elif isinstance(v, (int, float)) and abs(v) > 1000:
                        c.number_format = '#,##0'
                cur += 1
            cur += 1
        src_wb.close()
    except Exception as e:
        ws.cell(row=cur, column=1, value=f'  [ERROR] {e}').font = NOTE_FONT
        cur += 1
    cur += 1

# 너비
widths = [20, 18, 18, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
print(f'시트 [{S26}] 작성 완료')


# ============================================================
# 카탈로그 / 대시보드 / README 갱신
# ============================================================
ws_cat = wb['2. 자료 카탈로그']
# D, C entry 의 백필 상태 갱신
for r in range(1, ws_cat.max_row + 1):
    id_ = ws_cat.cell(row=r, column=1).value
    if id_ == 'C':
        safe_set(ws_cat, r, 6, '✅ M111~M115 백필 완료 + 시트 24 raw 보유 (2,947행)')
    elif id_ == 'D':
        safe_set(ws_cat, r, 6, '🔥 M130/M131 부분 백필 + 시트 23 raw 보유 (304행, BL 추출 컬럼 포함)')
    elif id_ == 'J':
        safe_set(ws_cat, r, 6, '🔥 M131 9건 + 시트 25 합계 추출 (xlsx 청구서)')
    elif id_ == 'J1':
        safe_set(ws_cat, r, 6, '🔥 M131 9건 + 시트 25 합계 자동 추출')

# 대시보드 — 자료 흡수 진행률에 D, C 갱신 + 행 추가
ws_d = wb['0. 대시보드']
for r in range(1, ws_d.max_row + 1):
    id_ = ws_d.cell(row=r, column=1).value
    if id_ == 'C':
        safe_set(ws_d, r, 3, '시트 8 (매트릭스) + 시트 24 (raw 2947행)')
    elif id_ == 'D':
        safe_set(ws_d, r, 3, '시트 9 (요약) + 시트 23 (raw 304행)')
    elif id_ == 'J':
        safe_set(ws_d, r, 3, '시트 19 (인벤토리) + 시트 25 (합계 자동 추출)')

# README 갱신 — 23~26 안내 추가
ws_r = wb['1. README']
for r in range(1, ws_r.max_row + 1):
    v = ws_r.cell(row=r, column=1).value
    if v and '22. (NEW) 면장 정본' in str(v):
        # 22 다음에 23~26 4줄 추가
        ws_r.insert_rows(r + 1, amount=4)
        new_lines = [
            ' 23. (NEW) D-회계 전표 raw — 부대비용·운송료 6시트 304행 통합 (M134 백필 직접 소스)',
            ' 24. (NEW) C-BL별 출고현황 raw — 6 제조사 시트 2,947행 통합 (운영자 수기 정본)',
            ' 25. (NEW) 운송료 청구서 합계 추출 — 25/26년 BL별 청구서 xlsx 자동 합계 (incidental_expenses)',
            ' 26. (NEW) 보조 xlsx 통합 — PO 정리·론지 정산·JA/KNK·재고·면장리스트·포워더 견적 9개',
        ]
        for i, line in enumerate(new_lines):
            ws_r.cell(row=r + 1 + i, column=1, value=line).font = BODY_FONT
        break

# 시트 순서 정렬
def sort_key(name):
    try:
        n = int(name.split('.')[0])
        return (0, n)
    except ValueError:
        return (1, name)
sorted_names = sorted(wb.sheetnames, key=sort_key)
wb._sheets = [wb[n] for n in sorted_names]

wb.save(TARGET)
print(f'\n저장 완료: {TARGET}')

wb2 = load_workbook(TARGET, read_only=True)
print(f'\n최종 시트 ({len(wb2.sheetnames)}개):')
for i, n in enumerate(wb2.sheetnames):
    print(f'  {i:2d}. {n}')
sz = os.path.getsize(TARGET)
print(f'\n파일 크기: {sz:,}B ({sz/1024:.1f}KB)')
wb2.close()
