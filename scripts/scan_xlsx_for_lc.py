#!/usr/bin/env python
"""Dropbox 의 모든 xlsx 에서 LC no 검색 — 4 누락 LC + 전체 신규 LC.

xlsx 안의 셀 텍스트 + 시트명 + 파일명 종합 검색.
같은 행에서 amount/bank/date 등 보조 정보 수집.
"""

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DROPBOX_ROOT = Path(r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료')
OUT_DIR = Path(__file__).parent / 'output'

LC_RE = re.compile(r'\b(M[A-Z0-9]{2,8}\d{2}NU\d{5,6})\b')

# DB LCs
DB_LCS = set()
with (OUT_DIR / 'lc_records_all.psv').open(encoding='utf-8') as f:
    for line in f:
        parts = line.strip().split('|')
        if parts and parts[0] and parts[0] != '없음':
            DB_LCS.add(parts[0])

# M144 추가분
DB_LCS |= {'M0215402NU00071','M0215402NU00089','M0215403NU00300','M0215405NU00228',
           'M0215405NU00331','M0215407NU00370','M0215407NU00395','M0215410NU00281','M0215509NU00317'}

# 4 target
TARGET = {'M04NG2512NU00018', 'M04NG2512NU00025', 'M04PH2512NU00032', 'M42M62602NU00018'}


def main():
    # 루트 xlsx + 전체 xlsx
    import glob
    all_xlsx = []
    # 루트 우선
    for f in sorted(DROPBOX_ROOT.glob('*.xlsx')):
        all_xlsx.append(f)
    # 그 다음 전체
    for f in DROPBOX_ROOT.rglob('*.xlsx'):
        if f not in all_xlsx and not f.name.startswith('~'):
            all_xlsx.append(f)

    print(f'xlsx 대상: {len(all_xlsx)}')

    target_hits = defaultdict(list)  # lc -> [(file, sheet, row_text)]
    all_lcs_hits = defaultdict(list)
    rows = []

    for i, path in enumerate(all_xlsx):
        rel = path.relative_to(DROPBOX_ROOT).as_posix()
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            continue

        for sn in wb.sheetnames:
            try:
                ws = wb[sn]
                # 셀 단위 row scan
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    row_text = ' | '.join(str(c) for c in row if c is not None)
                    if not row_text:
                        continue
                    lcs = LC_RE.findall(row_text)
                    if not lcs:
                        continue
                    for lc in lcs:
                        all_lcs_hits[lc].append((rel, sn, row_idx, row_text[:200]))
                        if lc in TARGET:
                            target_hits[lc].append((rel, sn, row_idx, row_text[:300]))
            except Exception:
                continue
        wb.close()

        if (i + 1) % 30 == 0:
            print(f'  진행: {i+1}/{len(all_xlsx)}')

    # 결과 저장
    new_lcs = set()
    for lc in all_lcs_hits.keys():
        if lc not in DB_LCS:
            new_lcs.add(lc)

    print(f'\n=== xlsx LC scan 결과 ===')
    print(f'스캔 xlsx:           {len(all_xlsx)}')
    print(f'LC no 발견 고유:     {len(all_lcs_hits)}')
    print(f'  DB 미등록:         {len(new_lcs)}')
    print(f'  Target 4 LC 매칭:  {len(target_hits)}')

    print(f'\n--- Target 4 LC 검색 결과 ---')
    for lc in sorted(TARGET):
        hits = target_hits.get(lc, [])
        print(f'\n  {lc}: {len(hits)} 행')
        for h in hits[:5]:
            print(f'    [{h[0][-60:]}::{h[1]}] row {h[2]}')
            print(f'      {h[3][:200]}')

    # CSV 저장
    with (OUT_DIR / 'lc_xlsx_target_hits.csv').open('w', encoding='utf-8-sig', newline='') as f:
        wr = csv.writer(f)
        wr.writerow(['lc_number', 'file', 'sheet', 'row', 'row_text'])
        for lc in sorted(TARGET):
            for h in target_hits.get(lc, []):
                wr.writerow([lc, h[0], h[1], h[2], h[3]])

    with (OUT_DIR / 'lc_xlsx_all_new.csv').open('w', encoding='utf-8-sig', newline='') as f:
        wr = csv.writer(f)
        wr.writerow(['lc_number', 'file_count', 'sample_file', 'sample_sheet', 'sample_row'])
        for lc in sorted(new_lcs):
            hits = all_lcs_hits[lc]
            wr.writerow([lc, len(hits), hits[0][0], hits[0][1], hits[0][3]])

    if new_lcs - TARGET:
        print(f'\n--- 추가 신규 LC (xlsx 에서만 발견) ---')
        for lc in sorted(new_lcs - TARGET)[:20]:
            hits = all_lcs_hits[lc]
            print(f'  {lc} ({len(hits)} 행): {hits[0][0][-60:]}')


if __name__ == '__main__':
    main()
