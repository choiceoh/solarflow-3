# -*- coding: utf-8 -*-
"""잔여 PDF 추출 + 시트 22 (면장 정본) 데이터 augment.

추가 대상:
- 25년 운송료 폴더 PDF 5개 (청구서/INV)
- 블루 재고 확인서 PDF 5개 (재고)
- 26년 운송료 청구자료/* 24개 (월별 청구서)
- 26년 수입면장/*.zip 안 PDF (납부고지서+필증, 4 zip)
"""
from __future__ import annotations

import os
import re
import sys
import json
import tempfile
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import pdfplumber

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
TARGET = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')
JSON_OLD = os.path.join(os.path.dirname(__file__), '..', 'pdf_extract_backup.json')
JSON_NEW = os.path.join(os.path.dirname(__file__), '..', 'pdf_extract_round2.json')

HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='1F4E78')
NOTE_FONT = Font(name='맑은 고딕', italic=True, size=9, color='666666')
BODY_FONT = Font(name='맑은 고딕', size=10)
BORDER_THIN = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))


# ---------- 파싱 헬퍼 (extract_pdf_data.py 재사용) ----------
def num(s):
    if s is None:
        return None
    s = s.replace(',', '').strip()
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def search1(pat, text, group=1, flags=0):
    m = re.search(pat, text, flags)
    return m.group(group).strip() if m else None


def is_payment_notice(text):
    return '납부영수증서' in text or '납부고지서' in text or '납부서[수납기관용]' in text or '납부영수' in text


def is_declaration(text):
    return '수입신고필증' in text or '신고필증' in text


def is_invoice(text):
    return ('청구서' in text or 'INVOICE' in text.upper() or '작업료' in text) and '수입신고필증' not in text


def is_inventory_cert(text):
    return '재고확인' in text or '재고 확인' in text or '타처보관' in text or '재고자산' in text


def parse_declaration(text):
    d = {}
    d['declaration_no'] = search1(r'(\d{4,5}-\d{2}-\d{6,7}M)', text)
    dates = re.findall(r'\d{4}/\d{2}/\d{2}', text)
    d['decl_date'] = dates[0] if dates else None
    d['arrival_date'] = dates[1] if len(dates) >= 2 else None
    d['release_date'] = dates[-1] if dates else None
    d['bl_number'] = search1(r'(JWSH\d+|SNK[oO]03[A-Z0-9]+|HDMUSHAA\d+|SHKWA\d+|EASE[KD]\w+|EASHO\w+|SELHTZ\w+|SELYIT\d+|SHACYV\w+|SHACYR\w+|NPSELHT\d+|LS\d+|RSPN\d+|JAHF\d+|MCKRJH\w+|TMSHKPTP\d+|DFS\d+|SHADFC\w+|ESZX\d+|HDMU\w+|HGHDCS\w+|SHADDP\w+|SHADGN\w+|SNKO02N\w+|PCSLJBL\w+|EASED\w+|EASEK\w+|KD\d+)', text)
    d['importer'] = search1(r'수\s*입\s*자\s+([가-힣()㈜0-9A-Za-z\s\.]+?)\(', text) \
        or search1(r'(탑솔라\(주\)|디원|화신이엔지|\(주\)디원)', text)
    d['biz_no'] = search1(r'(\d{3}-\d{2}-\d{5})', text)
    d['forwarder'] = search1(r'운송주선인\s+(.+?)(?:\d+종류|18종류|\n)', text)
    d['trade_partner'] = search1(r'무역거래처\s+(.+?)(?:\d{2}MASTER|\n)', text)
    d['master_bl'] = search1(r'MASTER B/L번호\s+([A-Z0-9]+)', text)
    d['origin_country'] = search1(r'적출국\s+([A-Z]{2})', text)
    d['arrival_port'] = search1(r'국내도착항\s+([A-Z]{4,6})', text)
    d['model'] = search1(r'(JKM\d+\w*-\w+|LR7-\w+|TSM-\w+|RSM\d+\w*-\w+|JAM\w+|JAHF\d+|JCM\d+|JC\d+\w*)', text)
    d['wp'] = search1(r'\((\d{3,4})W\)', text) or search1(r'(\d{3,4})W\b', text)
    d['pcs'] = num(search1(r'\(([\d,]+)\s*PCS?\)', text)) or num(search1(r'([\d,]+)\s*PC\b', text))
    d['unit_price_usd'] = num(search1(r'단가\(USD\).*?(0\.\d+)', text, flags=re.DOTALL))
    d['cif_usd'] = num(search1(r'\$\s*([\d,]+(?:\.\d+)?)', text))
    d['cif_krw'] = num(search1(r'￦\s*([\d,]+)', text))
    d['exchange_rate'] = num(search1(r'환\s*율\s+([\d,]+\.\d+)', text))
    d['vat'] = num(search1(r'부가가치세\s+([\d,]+)', text))
    d['total_tax'] = num(search1(r'총세액합계\s+([\d,]+)', text))
    d['lc_no'] = search1(r'L/C[\s\.]*NO[\s\.：:]*([A-Z0-9]+)', text, flags=re.IGNORECASE)
    if d['vat'] and d['total_tax'] == d['vat']:
        d['customs_duty'] = 0
    else:
        d['customs_duty'] = num(search1(r'관\s*세\s+([\d,]+)\s', text))
    return d


def parse_payment_notice(text):
    d = {}
    d['notice_bl'] = search1(r'B/L\s*No\.?\s*[:：]\s*(\w+)', text)
    d['declaration_no'] = search1(r'수입신고번호\s+([\d]{4,5}-\d{2}-\d{6,7}M)', text)
    d['importer'] = search1(r'상\s*호\s+(.+?)(?:\n|주\s*소)', text)
    d['biz_no'] = search1(r'사업번호\s*[:：]\s*(\d{3}-\d{2}-\d{5})', text)
    d['issue_date'] = search1(r'발행일자\s+(\d{4}년\d{2}월\d{2}일)', text)
    d['due_date'] = search1(r'납부기한\s+(\d{4}년\d{2}월\d{2}일)', text)
    d['customs_duty'] = num(search1(r'관\s*세\s+([\d,]+)', text))
    d['vat'] = num(search1(r'부가가치세\s+([\d,]+)', text))
    d['amount_total'] = num(search1(r'납기내\s*\n.*?([\d,]{4,})', text, flags=re.DOTALL))
    d['customs_office'] = search1(r'수입징수관서\s+(\S+세관)', text)
    return d


def parse_invoice(text):
    d = {}
    # 청구금액 — 보통 합계 옆 가장 큰 숫자
    nums = re.findall(r'([\d,]{4,})', text)
    big_nums = sorted(set(int(n.replace(',', '')) for n in nums if int(n.replace(',', '')) > 10000), reverse=True)[:5]
    d['top_amounts'] = big_nums
    # BL 추정
    d['bl_estimated'] = search1(r'(JWSH\d+|SNK[oO]03[A-Z0-9]+|HDMUSHAA\d+|SHKWA\d+|EASE[KD]\w+|SHACYV\w+|HGHDCS\w+|SHADDP\w+|SHADGN\w+|SNKO02N\w+|PCSLJBL\w+|EASED\w+|KD\d+)', text)
    # 거래처
    d['vendor'] = search1(r'(\(주\)?블루오션\w*|\(주\)?선진로지스틱\w*|스마일로지\w*|블루오션에어|선진로지스틱스)', text)
    return d


def parse_inventory_cert(text):
    d = {}
    # 재고 (단위: kg, ea, mw 추정)
    d['date'] = search1(r'(\d{4}[\.년]\s*\d{1,2}[\.월]\s*\d{1,2}일?)', text)
    d['company'] = search1(r'(탑솔라\(주\)|디원|\(주\)디원|화신이엔지)', text)
    # 합계 수량 추정
    nums = re.findall(r'([\d,]{4,})', text)
    big_nums = sorted(set(int(n.replace(',', '')) for n in nums if int(n.replace(',', '')) > 10000), reverse=True)[:5]
    d['top_amounts'] = big_nums
    d['note'] = '재고확인서/타처보관 — bl_shipments 정합 검증 소스'
    return d


def extract_one(args):
    full, year, source_label = args
    name = os.path.basename(full)
    try:
        with pdfplumber.open(full) as pdf:
            text = ''
            for p in pdf.pages:
                t = p.extract_text()
                if t: text += t + '\n'
        if not text.strip():
            return {'year': year, 'file': name, 'source': source_label, 'kind': 'EMPTY', 'error': 'no text (scan?)'}
        if is_payment_notice(text):
            kind = 'PAYMENT'; data = parse_payment_notice(text)
        elif is_declaration(text):
            kind = 'DECLARATION'; data = parse_declaration(text)
        elif is_inventory_cert(text):
            kind = 'INVENTORY_CERT'; data = parse_inventory_cert(text)
        elif is_invoice(text):
            kind = 'INVOICE'; data = parse_invoice(text)
        else:
            kind = 'OTHER'; data = {'note': text[:300]}
        data['year'] = year
        data['file'] = name
        data['source'] = source_label
        data['kind'] = kind
        return data
    except Exception as e:
        return {'year': year, 'file': name, 'source': source_label, 'kind': 'ERROR', 'error': str(e)[:200]}


# ---------- 잔여 PDF 수집 ----------
pdfs = []

# 25년 운송료 폴더 PDF
for f in os.listdir(os.path.join(ROOT, '2025년 운송료')):
    if f.lower().endswith('.pdf'):
        pdfs.append((os.path.join(ROOT, '2025년 운송료', f), '2025', '운송료_폴더'))

# 블루 재고 확인서
inv_dir = os.path.join(ROOT, '2025년 모듈 발주', '2025. 블루 재고 확인서')
for f in os.listdir(inv_dir):
    if f.lower().endswith('.pdf'):
        pdfs.append((os.path.join(inv_dir, f), '2025', '재고확인서'))

# 26년 운송료 청구자료/*
trans_26_root = os.path.join(ROOT, '2026년 모듈 발주', '운송료 청구자료')
for sub in os.listdir(trans_26_root):
    sub_full = os.path.join(trans_26_root, sub)
    if not os.path.isdir(sub_full): continue
    for f in os.listdir(sub_full):
        if f.lower().endswith('.pdf'):
            pdfs.append((os.path.join(sub_full, f), '2026', f'운송료_{sub}'))

# 26년 수입면장 폴더 안 zip 4개 → 임시 풀어서 PDF 추출
print(f'기본 PDF: {len(pdfs)}건')
zip_dir = os.path.join(ROOT, '2026년 모듈 발주', '수입면장')
zip_pdfs_extracted = []
with tempfile.TemporaryDirectory() as tmpdir:
    zip_count = 0
    for f in os.listdir(zip_dir):
        if f.lower().endswith('.zip'):
            zip_count += 1
            zfp = os.path.join(zip_dir, f)
            try:
                with zipfile.ZipFile(zfp) as zf:
                    for n in zf.namelist():
                        if n.lower().endswith('.pdf'):
                            extracted = zf.extract(n, tmpdir)
                            zip_pdfs_extracted.append((extracted, '2026', f'수입면장_zip:{f[:50]}'))
            except Exception as e:
                print(f'  zip 풀기 실패 {f}: {e}')
    print(f'26년 수입면장 zip {zip_count}개 → PDF {len(zip_pdfs_extracted)}건 추출')
    pdfs.extend(zip_pdfs_extracted)

    print(f'\n총 {len(pdfs)} PDF 처리')
    results = []
    for i, args in enumerate(pdfs):
        if i % 15 == 0:
            print(f'  {i}/{len(pdfs)} ...')
        results.append(extract_one(args))

print(f'추출 완료: {len(results)}건')

kinds = {}
for r in results:
    k = r.get('kind', '?')
    kinds[k] = kinds.get(k, 0) + 1
print(f'종류 분포: {kinds}')

# JSON 백업
with open(JSON_NEW, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)
print(f'JSON 백업: {JSON_NEW}')

# 기존 데이터 + 신규 데이터 병합
with open(JSON_OLD, 'r', encoding='utf-8') as f:
    old_results = json.load(f)

# normalize old (관세 0 normalize)
for r in old_results:
    if r.get('kind') == 'DECLARATION':
        if r.get('customs_duty') is None and r.get('vat') and r.get('total_tax') == r.get('vat'):
            r['customs_duty'] = 0
        if 'source' not in r:
            r['source'] = '면장_폴더'

all_results = old_results + results

decl_all = [r for r in all_results if r.get('kind') == 'DECLARATION']
pay_all = [r for r in all_results if r.get('kind') == 'PAYMENT']
inv_all = [r for r in all_results if r.get('kind') == 'INVOICE']
inv_cert_all = [r for r in all_results if r.get('kind') == 'INVENTORY_CERT']
err_all = [r for r in all_results if r.get('kind') in ('OTHER', 'EMPTY', 'ERROR')]

print(f'\n전체 합산: 면장 {len(decl_all)} / 납부고지서 {len(pay_all)} / 청구서 {len(inv_all)} / 재고확인 {len(inv_cert_all)} / 기타 {len(err_all)}')

total_cif_krw = sum(r.get('cif_krw') or 0 for r in decl_all if isinstance(r.get('cif_krw'), (int, float)))
total_vat_decl = sum(r.get('vat') or 0 for r in decl_all if isinstance(r.get('vat'), (int, float)))
total_customs_decl = sum(r.get('customs_duty') or 0 for r in decl_all if isinstance(r.get('customs_duty'), (int, float)))
total_pay_amt = sum(r.get('amount_total') or 0 for r in pay_all if isinstance(r.get('amount_total'), (int, float)))

# ---------- 시트 22 재빌드 (확장) ----------
print('\nxlsx 시트 22 재빌드 ...')
wb = load_workbook(TARGET)

S22 = '22. 면장 정본 데이터 (PDF 추출)'
if S22 in wb.sheetnames:
    del wb[S22]
ws = wb.create_sheet(S22)

ws['A1'] = '면장 정본 데이터 — 수입신고필증 + 납부고지서 + 청구서 + 재고확인서 PDF 텍스트 추출'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:Y1')
ws['A2'] = (f'전체 PDF 처리: 면장 {len(decl_all)} / 납부고지서 {len(pay_all)} / 청구서 {len(inv_all)} / '
            f'재고확인 {len(inv_cert_all)} / 기타 {len(err_all)}')
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:Y2')
ws['A3'] = (f'합계: CIF ₩{total_cif_krw:,.0f} ({total_cif_krw/1e8:.0f}억)  /  관세 ₩{total_customs_decl:,.0f}  /  '
            f'부가세 ₩{total_vat_decl:,.0f} ({total_vat_decl/1e8:.0f}억)  /  납부고지서 합 ₩{total_pay_amt:,.0f}')
ws['A3'].font = NOTE_FONT
ws.merge_cells('A3:Y3')

# 섹션 1. 면장
ws.cell(row=5, column=1, value='1. 수입신고필증 (면장)').font = SECTION_FONT
ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=25)

decl_cols = [
    ('연도', 'year', 6), ('출처', 'source', 16), ('PDF 파일', 'file', 38),
    ('신고번호', 'declaration_no', 18), ('신고일', 'decl_date', 12), ('입항일', 'arrival_date', 12),
    ('수리일자', 'release_date', 12), ('B/L 번호', 'bl_number', 22), ('MASTER B/L', 'master_bl', 18),
    ('수입자', 'importer', 14), ('사업자번호', 'biz_no', 14), ('포워더', 'forwarder', 26),
    ('무역거래처', 'trade_partner', 28), ('적출국', 'origin_country', 8), ('도착항', 'arrival_port', 8),
    ('모델', 'model', 22), ('Wp', 'wp', 6), ('PCS', 'pcs', 10), ('단가($)', 'unit_price_usd', 10),
    ('CIF($)', 'cif_usd', 14), ('CIF(₩)', 'cif_krw', 16), ('환율', 'exchange_rate', 10),
    ('관세(₩)', 'customs_duty', 12), ('부가세(₩)', 'vat', 14), ('총세액(₩)', 'total_tax', 14),
    ('L/C No', 'lc_no', 22),
]
hdr_row = 6
for j, (h, _, _) in enumerate(decl_cols):
    c = ws.cell(row=hdr_row, column=j + 1, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER_THIN
cur = hdr_row + 1
for r in decl_all:
    for j, (_, key, _) in enumerate(decl_cols):
        val = r.get(key)
        c = ws.cell(row=cur, column=j + 1, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER_THIN
        if key in ('cif_krw', 'customs_duty', 'vat', 'total_tax', 'pcs') and isinstance(val, (int, float)):
            c.number_format = '#,##0'
        elif key in ('cif_usd', 'unit_price_usd', 'exchange_rate') and isinstance(val, (int, float)):
            c.number_format = '#,##0.0000'
    cur += 1

# 합계
sum_row = cur
ws.cell(row=sum_row, column=1, value='합계').font = HEADER_FONT
ws.cell(row=sum_row, column=1).fill = HEADER_FILL
for col_idx, key in [(21, 'cif_krw'), (23, 'customs_duty'), (24, 'vat'), (25, 'total_tax')]:
    s = sum(r.get(key) or 0 for r in decl_all if isinstance(r.get(key), (int, float)))
    c = ws.cell(row=sum_row, column=col_idx, value=s)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.number_format = '#,##0'
cur = sum_row + 2

# 섹션 2. 납부고지서
ws.cell(row=cur, column=1, value='2. 납부고지서').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=25)
cur += 1
pay_cols = [('연도', 'year', 6), ('출처', 'source', 16), ('PDF 파일', 'file', 50),
            ('신고번호', 'declaration_no', 20), ('수입자', 'importer', 14), ('사업번호', 'biz_no', 14),
            ('발행일자', 'issue_date', 14), ('납부기한', 'due_date', 14),
            ('관세(₩)', 'customs_duty', 14), ('부가세(₩)', 'vat', 14), ('합계금액(₩)', 'amount_total', 14),
            ('수입징수관서', 'customs_office', 18), ('Notice B/L', 'notice_bl', 16)]
for j, (h, _, _) in enumerate(pay_cols):
    c = ws.cell(row=cur, column=j + 1, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = BORDER_THIN
cur += 1
for r in pay_all:
    for j, (_, key, _) in enumerate(pay_cols):
        val = r.get(key)
        c = ws.cell(row=cur, column=j + 1, value=val); c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True); c.border = BORDER_THIN
        if key in ('customs_duty', 'vat', 'amount_total') and isinstance(val, (int, float)):
            c.number_format = '#,##0'
    cur += 1
sum_row = cur
ws.cell(row=sum_row, column=1, value='합계').font = HEADER_FONT
ws.cell(row=sum_row, column=1).fill = HEADER_FILL
for col_idx, key in [(9, 'customs_duty'), (10, 'vat'), (11, 'amount_total')]:
    s = sum(r.get(key) or 0 for r in pay_all if isinstance(r.get(key), (int, float)))
    c = ws.cell(row=sum_row, column=col_idx, value=s)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.number_format = '#,##0'
cur = sum_row + 2

# 섹션 3. 청구서 (운송료)
if inv_all:
    ws.cell(row=cur, column=1, value='3. 청구서 (운송료/CFS) — PDF 추출').font = SECTION_FONT
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=25)
    cur += 1
    inv_cols = [('연도', 'year'), ('출처', 'source'), ('PDF 파일', 'file'),
                ('BL 추정', 'bl_estimated'), ('거래처', 'vendor'),
                ('상위 금액 (₩, 최대 5)', 'top_amounts')]
    for j, (h, _) in enumerate(inv_cols):
        c = ws.cell(row=cur, column=j + 1, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = BORDER_THIN
    cur += 1
    for r in inv_all:
        for j, (_, key) in enumerate(inv_cols):
            val = r.get(key)
            if key == 'top_amounts' and isinstance(val, list):
                val = ' / '.join(f'{n:,.0f}' for n in val)
            c = ws.cell(row=cur, column=j + 1, value=val); c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True); c.border = BORDER_THIN
        cur += 1
    cur += 1

# 섹션 4. 재고확인서
if inv_cert_all:
    ws.cell(row=cur, column=1, value='4. 거래처 재고확인서').font = SECTION_FONT
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=25)
    cur += 1
    cert_cols = [('연도', 'year'), ('출처', 'source'), ('PDF 파일', 'file'),
                 ('일자', 'date'), ('회사', 'company'), ('상위 수량/금액', 'top_amounts'), ('비고', 'note')]
    for j, (h, _) in enumerate(cert_cols):
        c = ws.cell(row=cur, column=j + 1, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = BORDER_THIN
    cur += 1
    for r in inv_cert_all:
        for j, (_, key) in enumerate(cert_cols):
            val = r.get(key)
            if key == 'top_amounts' and isinstance(val, list):
                val = ' / '.join(f'{n:,.0f}' for n in val)
            c = ws.cell(row=cur, column=j + 1, value=val); c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True); c.border = BORDER_THIN
        cur += 1
    cur += 1

# 섹션 5. 기타
if err_all:
    ws.cell(row=cur, column=1, value='5. 추출 실패 / 기타 양식').font = SECTION_FONT
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=25)
    cur += 1
    for j, h in enumerate(['연도', '출처', 'PDF 파일', '종류', '비고']):
        c = ws.cell(row=cur, column=j + 1, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
    cur += 1
    for r in err_all:
        vals = [r.get('year'), r.get('source'), r.get('file'), r.get('kind'),
                r.get('error') or (r.get('note', '') or '')[:120]]
        for j, v in enumerate(vals):
            c = ws.cell(row=cur, column=j + 1, value=v); c.font = BODY_FONT
        cur += 1

for j, (_, _, w) in enumerate(decl_cols):
    ws.column_dimensions[get_column_letter(j + 1)].width = w
ws.freeze_panes = 'A7'

# 카탈로그 갱신
ws_cat = wb['2. 자료 카탈로그']
for r in range(1, ws_cat.max_row + 1):
    id_ = ws_cat.cell(row=r, column=1).value
    if id_ == 'P':
        cell = ws_cat.cell(row=r, column=6)
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = (f'🔥 PDF 정본 추출 완료 — 면장 {len(decl_all)} / 납부고지서 {len(pay_all)} / '
                         f'청구서 {len(inv_all)} / 재고확인 {len(inv_cert_all)} (시트 22)')
        break

# 대시보드 갱신
ws_d = wb['0. 대시보드']
for r in range(1, ws_d.max_row + 1):
    id_ = ws_d.cell(row=r, column=1).value
    if id_ == 'P':
        cell = ws_d.cell(row=r, column=4)
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = (f'🔥 시트 21 메타 + 시트 22 정본 — 면장 {len(decl_all)} + 납부고지서 {len(pay_all)} + '
                          f'청구서 {len(inv_all)} + 재고확인 {len(inv_cert_all)}건')

wb.save(TARGET)
print(f'\n저장 완료: {TARGET}')
sz = os.path.getsize(TARGET)
print(f'파일 크기: {sz:,}B ({sz/1024:.1f}KB)')
