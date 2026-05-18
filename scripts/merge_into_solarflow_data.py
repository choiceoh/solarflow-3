# -*- coding: utf-8 -*-
"""산출물(통합정리.xlsx) 을 solarflow 자료.xlsx 에 머지 + 통합본 별도 파일 생성.

생성물:
1. solarflow 자료.xlsx — 기존 ERP 8시트 + 통합정리 30시트 (값 + 기본 스타일 복사)
2. 솔라플로우_자료_통합본.xlsx — 통합정리.xlsx 의 30시트 + ERP 8시트 (분석이 앞, ERP 가 뒤)
"""
from __future__ import annotations

import os
import sys
import shutil
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
ANALYSIS = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')
ERP = os.path.join(ROOT, 'solarflow 자료.xlsx')
COMBINED = os.path.join(ROOT, '솔라플로우_자료_통합본.xlsx')

# 원본 ERP 시트 (재실행 시 분석 시트가 머지된 상태여도 정확히 이 8개만 picking)
ERP_ORIGINAL_SHEETS = ['DB-3', '디원화신fifo', '탑솔라Fifo_복사본', '수불', '재고', '출고', '매출', '입고']


def copy_sheet_cells(src_ws, dst_ws):
    """워크북 간 시트 복사 — 셀 값 + Font/Fill/Alignment/Border + column widths + merged ranges."""
    # 컬럼 너비
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width
    # 행 높이
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_num].height = dim.height
    # 셀 값 + 스타일
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and cell.fill.fgColor.rgb in (None, '00000000'):
                continue
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.number_format = cell.number_format
    # 머지 영역
    for mr in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mr))
    # freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes


def copy_sheet_simple(src_ws, dst_ws, header_row=1):
    """간단한 ERP 시트 복사 — 값 only (read_only 모드 호환). 헤더만 굵게."""
    HEADER_FILL = PatternFill('solid', start_color='D9E1F2')
    HEADER_FONT = Font(name='맑은 고딕', bold=True, size=10)
    BODY_FONT = Font(name='맑은 고딕', size=10)
    for i, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
        for j, v in enumerate(row, start=1):
            c = dst_ws.cell(row=i, column=j, value=v)
            if i == header_row:
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
            else:
                c.font = BODY_FONT
    dst_ws.freeze_panes = f'A{header_row + 1}'


# ============================================================
# 1. 통합본 파일 생성: 통합정리 사본 → ERP 8시트 추가
# ============================================================
print(f'[1/2] 통합본 파일 생성: {os.path.basename(COMBINED)}')
shutil.copy2(ANALYSIS, COMBINED)
print(f'  통합정리 사본 복사 완료')

dst_wb = load_workbook(COMBINED)
# 이전 실행의 ERP_* 잔재 제거 (재실행 안전성)
for sn in list(dst_wb.sheetnames):
    if sn.startswith('ERP_'):
        del dst_wb[sn]

src_erp = load_workbook(ERP, read_only=True, data_only=True)

# 원본 ERP 8 시트만 picking (분석 시트가 ERP 파일에 머지된 상태여도 안전)
erp_sheets = [sn for sn in ERP_ORIGINAL_SHEETS if sn in src_erp.sheetnames]
print(f'  ERP 원본 시트 {len(erp_sheets)}개 추가 중 (전체 {len(src_erp.sheetnames)} 중 필터링)...')
for sn in erp_sheets:
    new_name = f'ERP_{sn}'[:31]
    if new_name in dst_wb.sheetnames:
        del dst_wb[new_name]
    dst_ws = dst_wb.create_sheet(new_name)
    src_ws = src_erp[sn]
    copy_sheet_simple(src_ws, dst_ws, header_row=1)
    print(f'    [{new_name}] {dst_ws.max_row}r × {dst_ws.max_column}c')
src_erp.close()

dst_wb.save(COMBINED)
sz = os.path.getsize(COMBINED)
print(f'  저장 완료: {COMBINED}  ({sz:,}B / {sz/1024/1024:.1f}MB)')

# ============================================================
# 2. solarflow 자료.xlsx 에 분석 30 시트 머지 (값+스타일 복사)
# ============================================================
print(f'\n[2/2] solarflow 자료.xlsx 에 분석 30 시트 머지')
print(f'  주의: ERP export 로 덮어쓰면 추가 시트 사라짐 → 운영자 안내 필요')

# 직접 수정은 위험 — 먼저 dry-run 으로 사이즈 확인
src_an = load_workbook(ANALYSIS)
dst_wb2 = load_workbook(ERP)

an_sheets = src_an.sheetnames
print(f'  분석 시트 {len(an_sheets)}개 → 머지 중 ...')
for sn in an_sheets:
    if sn in dst_wb2.sheetnames:
        del dst_wb2[sn]
    dst_ws = dst_wb2.create_sheet(sn)
    src_ws = src_an[sn]
    copy_sheet_cells(src_ws, dst_ws)
    print(f'    [{sn}] {dst_ws.max_row}r × {dst_ws.max_column}c')
src_an.close()

dst_wb2.save(ERP)
sz2 = os.path.getsize(ERP)
print(f'  저장 완료: {ERP}  ({sz2:,}B / {sz2/1024/1024:.1f}MB)')

# 결과 확인
print('\n=== 최종 확인 ===')
for path, label in [(COMBINED, '통합본'), (ERP, 'solarflow 자료 (머지)')]:
    wb = load_workbook(path, read_only=True)
    print(f'\n[{label}] {os.path.basename(path)} ({os.path.getsize(path):,}B / {os.path.getsize(path)/1024/1024:.1f}MB) — {len(wb.sheetnames)} 시트')
    for i, n in enumerate(wb.sheetnames):
        print(f'  {i:2d}. {n}')
    wb.close()
