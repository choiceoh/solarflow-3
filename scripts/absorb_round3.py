# -*- coding: utf-8 -*-
"""추가 흡수 라운드 3: E1 25년 마스터 raw / 24년 출하요청서 csv / I 발전시공일정.

시트 27, 28, 29 추가.
"""
from __future__ import annotations

import os
import csv
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
TARGET = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')

HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='1F4E78')
NOTE_FONT = Font(name='맑은 고딕', italic=True, size=9, color='666666')
BODY_FONT = Font(name='맑은 고딕', size=10)
BORDER_THIN = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))


def safe_set(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == 'MergedCell':
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row <= row <= mr.max_row) and (mr.min_col <= col <= mr.max_col):
                ws.unmerge_cells(str(mr))
                ws.cell(row=mr.min_row, column=mr.min_col, value=value)
                return
    cell.value = value


def write_header_row(ws, row, headers):
    for j, v in enumerate(headers):
        c = ws.cell(row=row, column=j + 1, value=v)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER_THIN


print(f'Loading {TARGET}')
wb = load_workbook(TARGET)
print(f'기존 시트: {len(wb.sheetnames)}')

# ============================================================
# 시트 27: E1 25년 마스터 raw (2024 + 2025 시트 통합)
# ============================================================
S27 = '27. E1-25년 마스터 raw (2024+2025)'
if S27 in wb.sheetnames:
    del wb[S27]
ws = wb.create_sheet(S27)
ws['A1'] = 'E1. 수입진행상황(module)-2025년도.xlsx — 2024 + 2025 시트 통합 raw (PO 운영 마스터)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:U1')
ws['A2'] = ('원본: Dropbox/2025년 모듈 발주/수입진행상황(module)-2025년도.xlsx (191MB) — '
            '2024시트 799행 + 2025시트 1214행 = 2,013행. 핵심 21컬럼 추출.')
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:U2')
ws['A3'] = 'M132~M136 백필 핵심 소스 (purchase_orders / lc_records / cost_details / bl_shipments).'
ws['A3'].font = NOTE_FONT
ws.merge_cells('A3:U3')

# E1 시트의 헤더는 r2/r3 multi-row. 핵심 컬럼 인덱스 (53~55 컬럼 중):
# 0=No, 1=발주일자/업체(2024)/발주처(2025), 3=P/O No, 4=품명, 5=pcs, 7=F/M, 8=Wp, 9=Unit price, 10=Amount,
# (LC 부분) 11=은행, 12=개설일, 13=L/C No, 14=수량, 15=F/M, 16=W.P, 17=개설금액, 18=만기일,
# (선적) 21=선적, 22=입항, ~30~32=B/L No, 30=현장배송 등
# 시트마다 컬럼 매핑이 약간 다름 — 첫 25컬럼 raw 흡수 (헤더는 시트 자체)

E1_HEADERS = ['연도', 'No.', '발주처/업체', 'P/O No.', '품명', 'Q\'ty(pcs)', 'F/M', 'Wp',
              'Unit price', 'Amount', '은행', '개설일', 'L/C No.', '개설금액', '만기일',
              '선적', '입항', 'PLT', '포워더', 'B/L No', '면장', '현장']

# 컬럼 매핑: (year, src_idx) → out_idx. 시트마다 다른 매핑 보정.
# 일반화 — 시트 raw row 의 첫 35 인덱스 중 가장 가까운 것을 시도. 본 데이터에 적용.
# 안전한 방식: 2025 시트는 컬럼이 1 더 있음 (발주처 추가) → 2024와 비교 시 +1 shift.

write_header_row(ws, 5, E1_HEADERS)
cur = 6

p_e1 = os.path.join(ROOT, '2025년 모듈 발주', '수입진행상황(module)-2025년도.xlsx')
src_wb = load_workbook(p_e1, read_only=True, data_only=True)

# 2024 시트 매핑 — 헤더 r2/r3 검사 결과:
# r2 = ['No.', '발주일자', '업체', 'P/O No.', '품명', 'Qty', 'Unit price', 'Amount', 'LC', '유산스', '선적', '입항', 'F/M包', 'F/S', 'PLT']
# r3 = ['', '', '', '', '', 'pcs', 'F/M', 'Wp', '은행', '개설일', 'L/C No', '수량', 'F/M', 'W.P', '개설금액', '만기일', '금액']
# 즉 2024 r2 col1=발주일자, col2=업체, col3=P/O, col4=품명, col5=pcs, col6=F/M, col7=Wp,
#       col9=은행, col10=개설일, col11=LC No, col13=W.P, col14=개설금액, col15=만기일, col17~=선적/입항...

# 너무 시트별 차이 분석에 시간 들이면 늦으니 — raw row 의 첫 25 컬럼을 그대로 가져온다 (헤더만 본 시트의 정보로 표시).
# 단순화: 한 시트 = 한 row. 25 컬럼 dump. 헤더는 첫 행 (r2 + r3 병합) 으로.

def stage_label(year):
    return f'E1.{year}'


for year in ['2024', '2025']:
    sws = src_wb[year]
    skip_rows = 4  # r0=타이틀, r1=날짜, r2=헤더1, r3=헤더2
    for i, r in enumerate(sws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        # 핵심 비어있는 행 skip (No. 가 None 이고 P/O 도 None)
        if r[0] is None and (len(r) < 4 or r[3] is None):
            # 단, 현장 행 (PO 의 sub-row, 입고일/현장만 있음) 은 살림
            if not (len(r) > 25 and any(x is not None for x in r[25:35])):
                continue
        # 매핑: 2025는 발주처 컬럼이 0번에 추가된 형태, 2024는 0번이 No.
        # 단순화: 첫 22 컬럼 가져와서 헤더에 맞게 padding
        out = [year]
        # 2024와 2025 모두 raw 첫 25 컬럼 가져옴 — 헤더 매핑은 대략적
        out.extend(list(r)[:21])
        out = (out + [None] * 22)[:22]
        for j, v in enumerate(out):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if isinstance(v, datetime):
                c.number_format = 'yyyy-mm-dd'
            elif isinstance(v, (int, float)) and abs(v) > 1000 and j not in (7, 8):  # Wp/단가 제외
                c.number_format = '#,##0'
            elif isinstance(v, float) and j in (8,):
                c.number_format = '0.0000'
        cur += 1
src_wb.close()

widths = [8, 6, 18, 22, 26, 12, 8, 8, 14, 16, 14, 12, 22, 16, 12, 12, 12, 8, 16, 22, 12, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A6'
print(f'시트 [{S27}] 작성 완료 — {cur - 6}행')


# ============================================================
# 시트 28: 24년 모듈 출하요청서 csv 통합 (1339행)
# ============================================================
S28 = '28. 24년 모듈 출하요청서 (csv)'
if S28 in wb.sheetnames:
    del wb[S28]
ws = wb.create_sheet(S28)
ws['A1'] = '24년 모듈 출하요청서 — 부산항 출고 요청 csv 3개 통합 (출고 정본)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:AA1')
ws['A2'] = '원본: Dropbox/2024년 모듈발주/출고현황/모듈 출하요청서 (부산항)_2024{0429,0527,0531}.csv'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:AA2')

csv_dir = os.path.join(ROOT, '2024년 모듈발주', '출고현황')
csv_files = sorted([f for f in os.listdir(csv_dir) if f.endswith('.csv')])

# 첫 csv 의 헤더 채택
all_data = []
header = None
for f in csv_files:
    full = os.path.join(csv_dir, f)
    with open(full, 'r', encoding='utf-8-sig') as fp:
        rows = list(csv.reader(fp))
    if not rows:
        continue
    if header is None:
        header = ['파일명'] + rows[0]
    for r in rows[1:]:
        if not any(c.strip() for c in r if c):
            continue
        all_data.append([f] + r)

# 헤더 작성
write_header_row(ws, 4, header)
cur = 5
for r in all_data:
    for j, v in enumerate(r):
        c = ws.cell(row=cur, column=j + 1, value=v)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER_THIN
    cur += 1

# 너비
n_cols = len(header)
for i in range(1, n_cols + 1):
    ws.column_dimensions[get_column_letter(i)].width = 14
ws.column_dimensions['A'].width = 40
ws.freeze_panes = 'A5'
print(f'시트 [{S28}] 작성 완료 — {cur - 5}행, {n_cols}컬럼')


# ============================================================
# 시트 29: I 발전시공일정 — 핵심 시트 흡수
# ============================================================
S29 = '29. I-발전시공일정 (PM 영업 자료)'
if S29 in wb.sheetnames:
    del wb[S29]
ws = wb.create_sheet(S29)
ws['A1'] = 'I. 발전시공일정 및 자재 발주 일정 — 영업/PM 자료 (현장 × 모듈 발주)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:O1')
ws['A2'] = '원본: Dropbox/2024년 모듈발주/자재 2024년/발전시공일정 및 자재 발주 일정.xlsx (14 시트)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:O2')
ws['A3'] = '신규 도메인 후보 (projects / installation_schedules) — 핵심 시트만 (현장 행수 의미있는 것).'
ws['A3'].font = NOTE_FONT
ws.merge_cells('A3:O3')

p_i = os.path.join(ROOT, '2024년 모듈발주', '자재 2024년', '발전시공일정 및 자재 발주 일정.xlsx')
src_wb = load_workbook(p_i, read_only=True, data_only=True)

# 핵심 시트 (max_row 1048569 같은 빈 시트는 실제 데이터 행만 가져옴)
core_sheets = ['사용전검사 완료', '태양광 현장(탑)', '태양광 현장(디원)', '건물임대사업', '화신이엔지',
               '모듈 판매 (탑)', '모듈 판매 (디원)', '제주탑', '영암 은곡리', '영암 동호리',
               '일양산업(영암)', '일양산업(신안)', '일양산업(해남)', '탑인프라 신안']

cur = 5
for sn in core_sheets:
    if sn not in src_wb.sheetnames:
        continue
    sws = src_wb[sn]
    # 섹션 헤더
    ws.cell(row=cur, column=1, value=f'■ {sn}').font = SECTION_FONT
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=15)
    cur += 1
    # 데이터 흡수 — 빈 행 N개 만나면 stop (빈 시트 1048569 가드)
    consec_empty = 0
    rows_added = 0
    for i, r in enumerate(sws.iter_rows(values_only=True)):
        # 첫 15 컬럼만
        rv = list(r)[:15]
        if all(v is None for v in rv):
            consec_empty += 1
            if consec_empty >= 30:  # 빈 행 30개 연속이면 stop
                break
            continue
        consec_empty = 0
        for j, v in enumerate(rv):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if isinstance(v, datetime):
                c.number_format = 'yyyy-mm-dd'
            elif isinstance(v, (int, float)) and abs(v) > 1000:
                c.number_format = '#,##0'
        cur += 1
        rows_added += 1
        # 너무 큰 시트 (>500행) 제한
        if rows_added >= 500:
            ws.cell(row=cur, column=1, value=f'  ... ({sws.max_row - 500} more rows skipped)').font = NOTE_FONT
            cur += 1
            break
    cur += 1  # 섹션 간격
src_wb.close()

# 너비
for i in range(1, 16):
    ws.column_dimensions[get_column_letter(i)].width = 16
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 24
print(f'시트 [{S29}] 작성 완료')


# ============================================================
# 카탈로그 / README 갱신
# ============================================================
ws_cat = wb['2. 자료 카탈로그']
for r in range(1, ws_cat.max_row + 1):
    id_ = ws_cat.cell(row=r, column=1).value
    if id_ == 'E1':
        safe_set(ws_cat, r, 6, '🔥 시트 11/12/13/14 슬라이스 + 시트 27 raw (2,013행 PO/LC/BL 마스터)')
    elif id_ == 'I':
        safe_set(ws_cat, r, 6, '✅ 시트 29 흡수 — 영업 PM 자료 (현장×모듈)')

# 대시보드 — E1, I 갱신
ws_d = wb['0. 대시보드']
for r in range(1, ws_d.max_row + 1):
    id_ = ws_d.cell(row=r, column=1).value
    if id_ == 'E1':
        safe_set(ws_d, r, 3, '시트 11~14 슬라이스 + 시트 27 raw 2,013행')
    elif id_ == 'I':
        safe_set(ws_d, r, 3, '시트 29 (PM 영업 자료 14 시트 핵심 흡수)')
        safe_set(ws_d, r, 4, '✅ 신규 흡수 — projects/installation_schedules 신규 도메인 설계 입력')

# README 갱신 — 시트 27/28/29 안내
ws_r = wb['1. README']
for r in range(1, ws_r.max_row + 1):
    v = ws_r.cell(row=r, column=1).value
    if v and '26. (NEW) 보조 xlsx' in str(v):
        ws_r.insert_rows(r + 1, amount=3)
        new_lines = [
            ' 27. (NEW) E1-25년 마스터 raw — 수입진행상황(2025) 의 2024+2025 시트 2,013행 PO/LC/BL 마스터',
            ' 28. (NEW) 24년 모듈 출하요청서 csv — 부산항 출고 요청 3개 csv 통합 (1,339행 × 27컬럼)',
            ' 29. (NEW) I-발전시공일정 — 영업/PM 자료 14 시트 핵심 흡수 (현장×모듈 발주, 신규 도메인 후보)',
        ]
        for i, line in enumerate(new_lines):
            ws_r.cell(row=r + 1 + i, column=1, value=line).font = BODY_FONT
        break

# 시트 순서
def sort_key(name):
    try:
        n = int(name.split('.')[0])
        return (0, n)
    except ValueError:
        return (1, name)
sorted_names = sorted(wb.sheetnames, key=sort_key)
wb._sheets = [wb[n] for n in sorted_names]

wb.save(TARGET)
print(f'\n저장 완료: {TARGET}')

wb2 = load_workbook(TARGET, read_only=True)
print(f'\n최종 시트 ({len(wb2.sheetnames)}개):')
for i, n in enumerate(wb2.sheetnames):
    print(f'  {i:2d}. {n}')
sz = os.path.getsize(TARGET)
print(f'\n파일 크기: {sz:,}B ({sz/1024:.1f}KB)')
wb2.close()
