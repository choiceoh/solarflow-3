#!/usr/bin/env python
"""수입진행상황(module)-2025년도.xlsx + 2026년도.xlsx 의 LC 마스터 자료 파싱.

xlsx 의 헤더 구조 자동 인식 → 컬럼 매핑 → LC 단위 row 추출.
필드: 회사/제조사/모델/수량/단가/amount/은행/개설일/ETD/ETA/BL no/도착지
"""

import csv
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DROPBOX_ROOT = Path(r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료')
OUT_DIR = Path(__file__).parent / 'output'

MASTER_FILES = [
    DROPBOX_ROOT / '2025년 모듈 발주' / '수입진행상황(module)-2025년도.xlsx',
    DROPBOX_ROOT / '2026년 모듈 발주' / '수입진행상황(module)-2026년도.xlsx',
]

LC_RE = re.compile(r'\b(M[A-Z0-9]{2,8}\d{2}NU\d{5,6})\b')


def to_iso(v):
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    s = str(v) if v else ''
    m = re.search(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', s)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return s


def main():
    rows = []
    for src in MASTER_FILES:
        if not src.exists():
            # glob fallback (한글 콤마 path 우회)
            from glob import glob
            cand = glob(str(src.parent) + '/수입진행상황*.xlsx')
            if cand:
                src = Path(cand[0])
            else:
                print(f'SKIP: {src}')
                continue
        print(f'\n=== {src.name} ===')
        wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            # 헤더 행 찾기: LC no 가 처음 등장하는 행의 위/위에 위치
            all_rows = list(ws.iter_rows(values_only=True))
            # 보통 첫 1~3행 안에 헤더
            header_idx = None
            for i, r in enumerate(all_rows[:5]):
                rt = ' '.join(str(c) for c in r if c is not None)
                if 'LC' in rt or '신용장' in rt or '발행은행' in rt or 'L/C' in rt:
                    header_idx = i
                    break
            if header_idx is None:
                # default 첫 행
                header_idx = 0
            header = [str(c) if c is not None else '' for c in all_rows[header_idx]]
            print(f'  sheet={sn}, header_row={header_idx+1}, cols={len(header)}')
            # 데이터 row 추출 (LC no 가 있는 행만)
            for ri, r in enumerate(all_rows[header_idx + 1:], header_idx + 2):
                rt = ' | '.join(str(c) for c in r if c is not None)
                m = LC_RE.search(rt)
                if not m:
                    continue
                lc = m.group(1)
                # 셀 단위로 dict
                row_dict = {}
                for ci, c in enumerate(r):
                    if ci < len(header):
                        col = header[ci].replace('\n', ' ').strip() or f'col{ci}'
                        row_dict[col] = c
                row_dict['__lc_number'] = lc
                row_dict['__source_file'] = src.name
                row_dict['__sheet'] = sn
                row_dict['__row'] = ri
                rows.append(row_dict)
        wb.close()
    print(f'\n총 LC row: {len(rows)}')

    if not rows:
        return

    # 헤더 합집합
    all_cols = set()
    for r in rows:
        all_cols |= set(r.keys())
    print(f'고유 컬럼: {len(all_cols)}')
    # 컬럼 빈도
    from collections import Counter
    col_freq = Counter()
    for r in rows:
        for k, v in r.items():
            if v is not None and v != '':
                col_freq[k] += 1
    print('\n상위 컬럼:')
    for k, n in col_freq.most_common(30):
        print(f'  {k[:40]:<42} {n}')

    # CSV 저장
    fieldnames = ['__lc_number', '__source_file', '__sheet', '__row'] + sorted(
        c for c in all_cols if not c.startswith('__')
    )
    with (OUT_DIR / 'lc_master_xlsx_rows.csv').open('w', encoding='utf-8-sig', newline='') as f:
        wr = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        wr.writeheader()
        for r in rows:
            # date 객체는 ISO 로 변환
            r2 = {}
            for k, v in r.items():
                if isinstance(v, (datetime, date)):
                    r2[k] = to_iso(v)
                else:
                    r2[k] = str(v) if v is not None else ''
            wr.writerow(r2)
    print(f'\n저장: {OUT_DIR / "lc_master_xlsx_rows.csv"}')


if __name__ == '__main__':
    main()
