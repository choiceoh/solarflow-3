# -*- coding: utf-8 -*-
"""솔라플로우_통합정리자료.xlsx 에 PDF·기타 파일 인벤토리 시트 21 추가.

추가/갱신:
1. 신규 시트 21 — 면장·납부고지서·기타 운영 자료 (PDF/HWP/CSV/XLS) 인벤토리
2. 시트 2 (자료 카탈로그) — P entry (면장 PDF), Q entry (운영 보조 자료) 추가
3. 시트 1 (README) — 시트 안내 행 추가
4. 시트 0 (대시보드) — 자료 흡수 진행률 P/Q row 추가
"""
from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
TARGET = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')

HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='1F4E78')
NOTE_FONT = Font(name='맑은 고딕', italic=True, size=9, color='666666')
BODY_FONT = Font(name='맑은 고딕', size=10)
BORDER_THIN = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))

BL_RE = re.compile(
    r'(SHACYV\w+|SHACYR\w+|SNK[oO]03[A-Z0-9]+|JWSH\d+|HDMUSHAA\d+|SHKWA\d+|EASED\d+|SELYIT\d+|'
    r'SHADFC\w+|ESZX\d+|NPSELHT\d+|LS\d+|RSPN\d+|JAHF\d+|MCKRJH\w+|KD\d+|TMSHKPTP\d+|DFS\d+|'
    r'EASEK\w+|EASHO\w+|SELHTZ\w+|SHA\w*\d+|2000\d+|8000\d+|JWSH\d+|HDMU\w+)'
)
DECL_RE = re.compile(r'(\d{11})')  # 면장번호 11자리
LICENSE_RE = re.compile(r'(\d{4,5})')


def detect_bl(name):
    m = BL_RE.search(name)
    return m.group(1) if m else ''


def detect_decl(name):
    """면장번호 추정 — 보통 11자리 숫자, 또는 '수입면장 - <ID>' 패턴."""
    if '수입면장' in name or '수입필증' in name or '납부고지서' in name:
        m = DECL_RE.search(name)
        if m:
            return m.group(1)
    return ''


def detect_company_from_name(name):
    if '디원' in name: return '디원'
    if '화신' in name: return '화신이엔지'
    if '탑솔라' in name or '탑' in name and '솔라' in name: return '탑솔라(주)'
    if '바로' in name: return '바로(주)'
    return ''


def file_meta(full):
    sz = os.path.getsize(full)
    mt = datetime.fromtimestamp(os.path.getmtime(full)).strftime('%Y-%m-%d')
    return sz, mt


def scan_dir_flat(folder, year_label):
    """주어진 폴더의 모든 파일 (재귀 1단계) → 메타 list."""
    rows = []
    if not os.path.isdir(folder):
        return rows
    for entry in sorted(os.listdir(folder)):
        full = os.path.join(folder, entry)
        if os.path.isdir(full):
            for sub in sorted(os.listdir(full)):
                sf = os.path.join(full, sub)
                if os.path.isdir(sf):
                    continue
                sz, mt = file_meta(sf)
                ext = os.path.splitext(sub)[1].lower()
                rows.append((year_label, f'{entry}/{sub}', ext, sz, mt,
                             detect_bl(sub) or detect_bl(entry),
                             detect_decl(sub),
                             detect_company_from_name(sub) or detect_company_from_name(entry)))
        else:
            sz, mt = file_meta(full)
            ext = os.path.splitext(entry)[1].lower()
            rows.append((year_label, entry, ext, sz, mt,
                         detect_bl(entry), detect_decl(entry),
                         detect_company_from_name(entry)))
    return rows


# ---------- 1. wb 로드 ----------
print(f'Loading {TARGET} ...')
wb = load_workbook(TARGET)
print(f'기존 시트: {len(wb.sheetnames)}개')

# ---------- 2. 신규 시트 21 추가 ----------
S21 = '21. P-면장·기타 PDF 인벤토리'
if S21 in wb.sheetnames:
    del wb[S21]
ws = wb.create_sheet(S21)

ws['A1'] = 'P. 면장 PDF + 운영 보조 자료 (HWP/CSV/XLS) 인벤토리'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:H1')
ws['A2'] = '원본: Dropbox/{24,25,26}년 모듈(발주)/수입(면장|신고필증) + 24년 공정진행현황·모듈 입찰·출고현황·재고 + 25년 운송료'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:H2')

hdr = ['연도', '파일명/경로', '확장자', '크기(B)', '갱신일', '추정 BL/PI', '면장번호 추정', '추정 회사']
hdr_row = 4

def write_header(row):
    for j, v in enumerate(hdr):
        c = ws.cell(row=row, column=j + 1, value=v)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER_THIN

def write_data_rows(start_row, rows):
    cur = start_row
    for row in rows:
        for j, v in enumerate(row):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER_THIN
            if j == 3:
                c.number_format = '#,##0'
        cur += 1
    return cur


# 섹션 1: 24년 수입면장 (두 폴더 합산)
ws.cell(row=4, column=1, value='1. 2024년 수입면장 PDF (Dropbox/2024년 모듈발주/수입면장/)').font = SECTION_FONT
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)
write_header(5)
rows_24 = scan_dir_flat(os.path.join(ROOT, '2024년 모듈발주', '수입면장'), '2024')
cur = write_data_rows(6, rows_24)

# 섹션 2: 25년 수입신고필증
cur += 1
ws.cell(row=cur, column=1, value='2. 2025년 수입신고필증 PDF (Dropbox/2025년 모듈 발주/수입신고필증/)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
cur += 1
write_header(cur)
cur += 1
rows_25 = scan_dir_flat(os.path.join(ROOT, '2025년 모듈 발주', '수입신고필증'), '2025')
cur = write_data_rows(cur, rows_25)

# 섹션 3: 26년 수입면장 (PDF + 납부고지서 + 기타)
cur += 1
ws.cell(row=cur, column=1, value='3. 2026년 수입면장 PDF + 납부고지서 (Dropbox/2026년 모듈 발주/수입면장/)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
cur += 1
write_header(cur)
cur += 1
rows_26 = scan_dir_flat(os.path.join(ROOT, '2026년 모듈 발주', '수입면장'), '2026')
cur = write_data_rows(cur, rows_26)

# 섹션 4: 24년 운영 보조 자료 (공정진행/모듈 입찰/출고현황/재고)
cur += 1
ws.cell(row=cur, column=1, value='4. 2024년 운영 보조 자료 (공정진행현황·모듈 입찰·출고현황·재고)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
cur += 1
write_header(cur)
cur += 1
aux_dirs = [
    ('2024년 모듈발주/공정진행현황', '24-공정진행'),
    ('2024년 모듈발주/모듈 입찰', '24-입찰'),
    ('2024년 모듈발주/출고현황', '24-출고요청'),
    ('2024년 모듈발주/2024년 재고', '24-재고'),
]
aux_rows = []
for sub, lbl in aux_dirs:
    folder = os.path.join(ROOT, sub)
    aux_rows.extend(scan_dir_flat(folder, lbl))
cur = write_data_rows(cur, aux_rows)

# 섹션 5: 25년 운송료 폴더 (블루오션/선진/씨앤아이 등 하위)
cur += 1
ws.cell(row=cur, column=1, value='5. 2025년 운송료 (블루오션/선진/씨앤아이/운송료 폴더 하위 PDF·xlsx)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
cur += 1
write_header(cur)
cur += 1
folder_25t = os.path.join(ROOT, '2025년 모듈 발주', '운송료')
trans_rows = scan_dir_flat(folder_25t, '2025-운송료')
cur = write_data_rows(cur, trans_rows)

# 너비
widths = [12, 60, 8, 12, 12, 22, 18, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A6'

print(f'시트 [{S21}] 작성 완료 — 24년 {len(rows_24)} + 25년 {len(rows_25)} + 26년 {len(rows_26)} + 24보조 {len(aux_rows)} + 25운송료 {len(trans_rows)}')

# ---------- 3. 카탈로그 (시트 2) 에 P, Q entry 추가 ----------
ws_cat = wb['2. 자료 카탈로그']
new_entries = [
    ('P', '수입면장 / 수입신고필증 PDF (3개년 142+ 건)', '면장·납부고지서·OBL 첨부 PDF (24/25/26년 폴더)', '면장 단위', 'import_declarations.declaration_number ↔ bl_shipments.bl_number', '🟡 PDF — 메타만 인덱싱 (텍스트 추출 미수행)', '24/25/26년 폴더/수입면장 (또는 수입신고필증)/'),
    ('Q', '운영 보조 자료 (공정진행/모듈 입찰/출고요청/재고/운송료 PDF)', 'HWP/CSV/XLS/PDF 혼재 (24년 공정진행현황·모듈 입찰·출고현황·재고 + 25년 운송료 하위)', '수시', '(영업·운영 reference, DB 매핑 없음)', '🟡 인벤토리만 — 일부 신규 도메인 후보', '24년/{공정진행현황,모듈 입찰,출고현황,2024년 재고}, 25년 모듈 발주/운송료/'),
]
start = ws_cat.max_row + 1
for i, row in enumerate(new_entries):
    for j, v in enumerate(row):
        c = ws_cat.cell(row=start + i, column=1 + j, value=v)
        c.border = BORDER_THIN
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.font = BODY_FONT
print(f'카탈로그에 P, Q entry 추가 (총 {ws_cat.max_row - 2}개 자료)')

# ---------- 4. README (시트 1) 시트 안내 갱신 ----------
ws_r = wb['1. README']
# r32 (시트 20) 다음에 r33 (시트 21) 추가하기 위해, 마지막 안내 라인 다음에 삽입
# 기존 r33 = 빈줄, r34 = 범례 시작
# 시트 21 안내 라인 추가 — r33 위치에 insert
ws_r.insert_rows(33)
c = ws_r.cell(row=33, column=1, value=' 21. (NEW) P. 면장·기타 PDF 인벤토리 — 24/25/26년 수입면장 142+ PDF + 운영 보조 자료')
c.font = BODY_FONT
print(f'README 갱신 — 시트 21 안내 추가')

# ---------- 5. 대시보드 (시트 0) "자료 흡수 진행률" 표에 P/Q row 추가 ----------
ws_d = wb['0. 대시보드']
# 마지막 데이터 행 (intake 마지막 = 'N' 행) 다음을 찾는다
# 더 안전하게는: 모든 행 스캔 → '4. 추천 다음 액션' 섹션 발견 → 그 직전 row 에 두 줄 추가
target_section = None
for r_idx in range(1, ws_d.max_row + 1):
    v = ws_d.cell(row=r_idx, column=1).value
    if v == '4. 추천 다음 액션':
        target_section = r_idx
        break
if target_section:
    # 섹션 4 시작 직전에 빈 줄이 1개 있을 것. 그 위에 두 줄 추가.
    ws_d.insert_rows(target_section, amount=2)
    new_rows = [
        ('P', '수입면장 / 수입신고필증 PDF (24/25/26년 142+ 건)', '시트 21', '✅ 신규 흡수 — PDF 메타 인덱스 (BL/면장번호/회사 추정 컬럼 포함)'),
        ('Q', '운영 보조 자료 (공정진행/입찰/출고요청/재고/운송료 PDF)', '시트 21', '✅ 신규 흡수 — HWP/CSV/XLS/PDF 혼재 운영 reference'),
    ]
    for i, row in enumerate(new_rows):
        for j, v in enumerate(row):
            c = ws_d.cell(row=target_section + i, column=1 + j, value=v)
            c.font = BODY_FONT
            c.border = BORDER_THIN
            c.alignment = Alignment(vertical='center', wrap_text=True)
    print(f'대시보드 — 자료 흡수 진행률에 P, Q 추가 (행 {target_section})')

# ---------- 6. 시트 순서 보장 ----------
def sort_key(name):
    try:
        n = int(name.split('.')[0])
        return (0, n)
    except ValueError:
        return (1, name)

sorted_names = sorted(wb.sheetnames, key=sort_key)
wb._sheets = [wb[n] for n in sorted_names]

# ---------- 7. 저장 ----------
wb.save(TARGET)
print(f'\n저장 완료: {TARGET}')

wb2 = load_workbook(TARGET, read_only=True)
print(f'\n최종 시트 ({len(wb2.sheetnames)}개):')
for i, n in enumerate(wb2.sheetnames):
    print(f'  {i:2d}. {n}')
sz = os.path.getsize(TARGET)
print(f'\n파일 크기: {sz:,}B ({sz/1024:.1f}KB)')
wb2.close()
