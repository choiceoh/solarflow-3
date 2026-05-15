#!/usr/bin/env python
"""빈 13 BL 의 BL no. 가 운송료/부대비용 xlsx 에 존재하는지 확인."""

import os
import sys
from collections import defaultdict

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DROPBOX_COST_DIR = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료\2025년 운송료'
COST_XLSX = os.path.join(
    r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료',
    '2025년, 2026년 모듈 부대비용, 운송료 내역.xlsx',
)

EMPTY_BLS = [
    'SNKO03K250200546', 'SNKO03K250200547', 'SNKO03K250201370', 'SHACZA82185',
    'PCSLJBL001250720', 'JWSH25090055', 'SHKWA25019767', 'SHKWA25019768',
    'SNKO03K251001475', 'ESZX2502368', 'HGHDC502961', 'SHADCF57885', 'SNKO03K251101919',
]


def scan_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_blobs = {}
    for sn in wb.sheetnames:
        parts = []
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    parts.append(str(v))
        sheet_blobs[sn] = ' | '.join(parts)
    wb.close()
    return sheet_blobs


def main():
    # 메인 xlsx 는 글로브로 찾기 (한글 콤마 경로 검증 우회)
    import glob
    parent = os.path.dirname(COST_XLSX)
    candidates = glob.glob(os.path.join(parent, '*부대비용*운송료*.xlsx'))
    targets = list(candidates)
    for f in sorted(os.listdir(DROPBOX_COST_DIR)):
        if f.lower().endswith('.xlsx'):
            targets.append(os.path.join(DROPBOX_COST_DIR, f))
    print(f'대상 xlsx: {len(targets)}')
    for t in targets:
        print(f'  - {os.path.basename(t)}')

    hits = defaultdict(list)
    for path in targets:
        if not os.path.exists(path):
            print(f'  SKIP (not found): {os.path.basename(path)}')
            continue
        try:
            blobs = scan_workbook(path)
        except Exception as e:
            print(f'  FAIL {os.path.basename(path)}: {e}')
            continue
        for sn, blob in blobs.items():
            for bl in EMPTY_BLS:
                if bl in blob:
                    hits[bl].append(f'{os.path.basename(path)}::{sn}')
                else:
                    tail = bl[-8:]
                    if len(tail) >= 6 and tail in blob:
                        hits[f'{bl} (tail8={tail})'].append(f'{os.path.basename(path)}::{sn}')

    print('\n=== 13 빈 BL xlsx 내 검색 결과 ===')
    for bl in EMPTY_BLS:
        if bl in hits:
            print(f'  ✓ {bl}: {hits[bl]}')
        else:
            tail_key = f'{bl} (tail8={bl[-8:]})'
            if tail_key in hits:
                print(f'  ? {bl} (tail 매치): {hits[tail_key][:3]}')
            else:
                print(f'  ✗ {bl}: 없음')


if __name__ == '__main__':
    main()
