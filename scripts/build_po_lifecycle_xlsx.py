"""PO 라이프사이클 통합 엑셀 빌더.

입력: /tmp/sf_lifecycle/*.json (gx10 운영 DB 에서 추출한 6개 시트 원본).
출력: C:/Users/user/Downloads/PO_라이프사이클_YYYYMMDD.xlsx

시트 구성:
  1. 요약          — PO 1행 = 라이프사이클 단계별 요약 (62행)
  2. PO 헤더       — purchase_orders 전체 (62행)
  3. PO 라인       — po_line_items (111행)
  4. LC 마스터     — lc_records (63행, PO 매핑)
  5. BL 마스터     — bl_shipments (178행, PO/LC 매핑)
  6. 면장          — import_declarations + cost_details (171행)
  7. BL별 출고집계 — outbound_bl_items 그룹 (90행)
  8. 전체 사슬     — PO → LC → BL → 면장 → 출고 펼친 long format
"""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path('C:/Users/user/AppData/Local/Temp/sf_lifecycle')
OUT = Path('C:/Users/user/Downloads') / f'PO_라이프사이클_{datetime.now():%Y%m%d}.xlsx'


def load(name: str):
    raw = (SRC / f'{name}.json').read_text(encoding='utf-8').strip()
    data = json.loads(raw) if raw and raw != '\\N' else []
    return data or []


def to_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def fmt_date_range(values):
    dts = sorted([d for d in (to_date(v) for v in values) if d])
    if not dts:
        return ''
    if dts[0] == dts[-1]:
        return dts[0].isoformat()
    return f'{dts[0].isoformat()} ~ {dts[-1].isoformat()}'


def num(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


# ─── styles ───────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='맑은 고딕', size=10)
SUB_FILL = PatternFill('solid', fgColor='F3F4F6')
BORDER = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)
# 진행 단계 6 색 (1=발주만 ~ 6=매출확정)
STAGE_FILL = {
    1: PatternFill('solid', fgColor='FEE2E2'),  # red-100
    2: PatternFill('solid', fgColor='FED7AA'),  # orange-100
    3: PatternFill('solid', fgColor='FEF3C7'),  # amber-100
    4: PatternFill('solid', fgColor='DBEAFE'),  # blue-100
    5: PatternFill('solid', fgColor='BFDBFE'),  # blue-200
    6: PatternFill('solid', fgColor='BBF7D0'),  # green-100
}
STAGE_LABEL = {
    1: '1. PO 만',
    2: '2. + LC',
    3: '3. + BL',
    4: '4. + 면장',
    5: '5. + 출고',
    6: '6. + 매출',
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_w=60, min_w=8):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = min_w
        for cell in col[:200]:  # 처음 200행만 계산
            v = cell.value
            if v is None:
                continue
            s = str(v)
            # 한글 가로폭 보정
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            max_len = max(max_len, w)
        ws.column_dimensions[letter].width = min(max_w, max_len + 2)


def write_rows(ws, headers, rows, *, num_cols=None, date_cols=None):
    ws.append(headers)
    style_header(ws)
    num_cols = set(num_cols or [])
    date_cols = set(date_cols or [])
    for r in rows:
        ws.append(r)
    # 포맷
    for col_idx in num_cols:
        letter = get_column_letter(col_idx)
        for cell in ws[letter][1:]:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
    for col_idx in date_cols:
        letter = get_column_letter(col_idx)
        for cell in ws[letter][1:]:
            cell.number_format = 'yyyy-mm-dd'
    # 바디 폰트
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
    ws.freeze_panes = 'A2'


# ─── 데이터 로드 ─────────────────────────────────────────────
pos = load('po_list')
po_lines = load('po_lines')
lcs = load('lc_list')
bls = load('bl_list')
decls = load('decl_list')
obi_agg = load('obi_agg')

print(f'PO {len(pos)} / PO라인 {len(po_lines)} / LC {len(lcs)} / BL {len(bls)} / 면장 {len(decls)} / 출고집계 {len(obi_agg)}')

# 인덱스
po_by_id = {p['po_id']: p for p in pos}
lc_by_id = {l['lc_id']: l for l in lcs}
bl_by_id = {b['bl_id']: b for b in bls}

lines_by_po = defaultdict(list)
for pl in po_lines:
    lines_by_po[pl['po_id']].append(pl)

lcs_by_po = defaultdict(list)
for lc in lcs:
    lcs_by_po[lc['po_id']].append(lc)

# BL 은 po_id 직접 또는 lc_id 경유 양쪽으로 PO 에 연결될 수 있음.
bls_by_po_direct = defaultdict(list)
bls_by_lc = defaultdict(list)
for bl in bls:
    if bl.get('po_id'):
        bls_by_po_direct[bl['po_id']].append(bl)
    if bl.get('lc_id'):
        bls_by_lc[bl['lc_id']].append(bl)

decls_by_bl = defaultdict(list)
for d in decls:
    if d.get('bl_id'):
        decls_by_bl[d['bl_id']].append(d)

obi_by_bl = {o['bl_id']: o for o in obi_agg}


def bls_for_po(po_id):
    """직접 + LC 경유 BL 모음 (중복 제거)."""
    seen = set()
    out = []
    for bl in bls_by_po_direct.get(po_id, []):
        if bl['bl_id'] not in seen:
            seen.add(bl['bl_id'])
            out.append(bl)
    for lc in lcs_by_po.get(po_id, []):
        for bl in bls_by_lc.get(lc['lc_id'], []):
            if bl['bl_id'] not in seen:
                seen.add(bl['bl_id'])
                out.append(bl)
    return out


# ─── workbook ────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ─── Sheet 1: 요약 ──────────────────────────────────────────
ws = wb.create_sheet('요약')
headers = [
    'PO 번호', '회사', '제조사', '계약유형', '계약일', 'Incoterms', '결제조건',
    '총수량(pcs)', '총MW', 'PO 상태',
    'LC 건수', 'LC 번호', 'LC 개설금액합(USD)', 'LC 개설일범위', 'LC 만기일범위',
    'LC 미상환건수',
    'BL 건수', 'BL 번호', 'ETD 범위', 'ETA 범위', '입항 범위',
    'BL CIF 합(KRW)', 'BL 상태',
    '면장 건수', '면장 paid 수량합', '면장 free 수량합', 'landed 합(KRW)',
    '출고건수', '출고 수량합', '출고처 수', '매출 출고건수', '시공 출고건수',
    '매출 공급가합(KRW)',
    '진행단계', 'PO Memo',
]
ws.append(headers)
style_header(ws)

for po in pos:
    pid = po['po_id']
    lcs_p = lcs_by_po.get(pid, [])
    bls_p = bls_for_po(pid)
    bl_ids = [b['bl_id'] for b in bls_p]
    decls_p = [d for bid in bl_ids for d in decls_by_bl.get(bid, [])]
    obi_p = [obi_by_bl[bid] for bid in bl_ids if bid in obi_by_bl]

    lc_amount = sum(num(l['amount_usd']) for l in lcs_p)
    lc_unsettled = sum(1 for l in lcs_p if not l.get('repaid'))
    cif_total = sum(num(b['cif_amount_krw']) for b in bls_p)
    paid_qty = sum(num(d['paid_qty']) for d in decls_p)
    free_qty = sum(num(d['free_qty']) for d in decls_p)
    landed = sum(num(d.get('landed_total_krw')) for d in decls_p)
    ob_count = sum(int(num(o['outbound_count'])) for o in obi_p)
    ob_qty = sum(num(o['allocated_qty']) for o in obi_p)
    sale_count = sum(int(num(o['sale_outbounds'])) for o in obi_p)
    constr_count = sum(int(num(o['construction_outbounds'])) for o in obi_p)
    supply_sum = sum(num(o.get('supply_amount_sum')) for o in obi_p)

    # 출고처 (BL별 site_list 들을 합쳐서 distinct count)
    sites = set()
    for o in obi_p:
        if o.get('site_list'):
            for s in o['site_list'].split(' | '):
                if s:
                    sites.add(s)
    site_n = len(sites)

    # 단계 (1~6)
    stage = 1
    if lcs_p: stage = 2
    if bls_p: stage = max(stage, 3)
    if decls_p: stage = max(stage, 4)
    if ob_count > 0: stage = max(stage, 5)
    if supply_sum > 0: stage = max(stage, 6)

    bl_statuses = sorted({(b.get('status') or '?') for b in bls_p})

    row = [
        po['po_number'], po.get('company_name'), po.get('manufacturer'),
        po.get('contract_type'), to_date(po.get('contract_date')),
        po.get('incoterms'), po.get('payment_terms'),
        po.get('total_qty'), po.get('total_mw'), po.get('status'),
        len(lcs_p),
        ' / '.join(l.get('lc_number') or '?' for l in lcs_p) if lcs_p else '',
        lc_amount if lc_amount else None,
        fmt_date_range([l.get('open_date') for l in lcs_p]),
        fmt_date_range([l.get('maturity_date') for l in lcs_p]),
        lc_unsettled if lcs_p else None,
        len(bls_p),
        ' / '.join(b.get('bl_number') or '?' for b in bls_p) if bls_p else '',
        fmt_date_range([b.get('etd') for b in bls_p]),
        fmt_date_range([b.get('eta') for b in bls_p]),
        fmt_date_range([b.get('actual_arrival') for b in bls_p]),
        cif_total if cif_total else None,
        ' / '.join(bl_statuses) if bls_p else '',
        len(decls_p),
        paid_qty if paid_qty else None,
        free_qty if free_qty else None,
        landed if landed else None,
        ob_count if ob_count else None,
        ob_qty if ob_qty else None,
        site_n if site_n else None,
        sale_count if sale_count else None,
        constr_count if constr_count else None,
        supply_sum if supply_sum else None,
        STAGE_LABEL[stage],
        po.get('memo'),
    ]
    ws.append(row)
    # stage 색칠
    last_row = ws.max_row
    stage_cell = ws.cell(row=last_row, column=headers.index('진행단계') + 1)
    stage_cell.fill = STAGE_FILL[stage]
    stage_cell.font = Font(name='맑은 고딕', size=10, bold=True)

# 포맷
num_cols_idx = [8, 9, 11, 13, 16, 17, 22, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33]
for col_idx in num_cols_idx:
    for cell in ws[get_column_letter(col_idx)][1:]:
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
# 계약일 (col 5)
for cell in ws['E'][1:]:
    cell.number_format = 'yyyy-mm-dd'

# 폰트/테두리
for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.font.name != '맑은 고딕':
            cell.font = BODY_FONT
        cell.border = BORDER

ws.freeze_panes = 'C2'
autosize(ws, max_w=40)


# ─── Sheet 2: PO 헤더 ───────────────────────────────────────
ws = wb.create_sheet('PO 헤더')
headers = ['PO 번호', '회사', '제조사', '계약유형', '계약일', 'Incoterms', '결제조건',
           '총수량(pcs)', '총MW', '계약시작', '계약종료', '상태', 'Memo',
           '상위 PO ID', 'PO_ID']
rows = []
for po in pos:
    rows.append([
        po['po_number'], po.get('company_name'), po.get('manufacturer'),
        po.get('contract_type'), to_date(po.get('contract_date')),
        po.get('incoterms'), po.get('payment_terms'),
        po.get('total_qty'), po.get('total_mw'),
        to_date(po.get('contract_period_start')), to_date(po.get('contract_period_end')),
        po.get('status'), po.get('memo'),
        po.get('parent_po_id'), po['po_id'],
    ])
write_rows(ws, headers, rows,
           num_cols=[8, 9], date_cols=[5, 10, 11])
autosize(ws, max_w=40)


# ─── Sheet 3: PO 라인 ───────────────────────────────────────
ws = wb.create_sheet('PO 라인')
headers = ['PO 번호', '품번', '품명', '규격', '수량', '단가 USD', 'Wp단가 USD',
           '총액 USD', '아이템', '결제유형', 'Memo']
rows = []
for pl in po_lines:
    po = po_by_id.get(pl['po_id'])
    rows.append([
        po['po_number'] if po else '?',
        pl.get('product_code'), pl.get('product_name'), pl.get('spec_wp'),
        pl.get('quantity'), pl.get('unit_price_usd'), pl.get('unit_price_usd_wp'),
        pl.get('total_amount_usd'), pl.get('item_type'), pl.get('payment_type'),
        pl.get('memo'),
    ])
# 계약일 기준 정렬은 PO 번호 정렬로 대체
rows.sort(key=lambda r: (r[0], r[1] or ''))
write_rows(ws, headers, rows, num_cols=[5, 6, 7, 8])
autosize(ws, max_w=40)


# ─── Sheet 4: LC 마스터 ─────────────────────────────────────
ws = wb.create_sheet('LC 마스터')
headers = ['LC 번호', '연결 PO', '은행', '회사', '개설일', '금액 USD', '목표수량',
           '목표MW', 'Usance(일)', 'Usance 유형', '만기일', '결제예정일',
           '상환일', '상환여부', '상태', 'Memo']
rows = []
for lc in lcs:
    po = po_by_id.get(lc.get('po_id'))
    rows.append([
        lc.get('lc_number'),
        po['po_number'] if po else '(PO 미연결)',
        lc.get('bank_name'), lc.get('company_name'),
        to_date(lc.get('open_date')), lc.get('amount_usd'),
        lc.get('target_qty'), lc.get('target_mw'),
        lc.get('usance_days'), lc.get('usance_type'),
        to_date(lc.get('maturity_date')), to_date(lc.get('settlement_date')),
        to_date(lc.get('repayment_date')),
        '✓' if lc.get('repaid') else '',
        lc.get('status'), lc.get('memo'),
    ])
write_rows(ws, headers, rows,
           num_cols=[6, 7, 8, 9], date_cols=[5, 11, 12, 13])
autosize(ws, max_w=40)


# ─── Sheet 5: BL 마스터 ─────────────────────────────────────
ws = wb.create_sheet('BL 마스터')
headers = ['BL 번호', '연결 PO', '연결 LC', '회사', '제조사', 'inbound_type',
           '통화', '환율', 'ETD', 'ETA', '실제입항', '항구', '포워더',
           'Invoice', '면장번호', 'CIF KRW', '상태', 'ERP 등록', 'Memo']
rows = []
for bl in bls:
    po = po_by_id.get(bl.get('po_id'))
    lc = lc_by_id.get(bl.get('lc_id'))
    rows.append([
        bl.get('bl_number'),
        po['po_number'] if po else '',
        lc.get('lc_number') if lc else '',
        bl.get('company_name'), bl.get('manufacturer'),
        bl.get('inbound_type'), bl.get('currency'), bl.get('exchange_rate'),
        to_date(bl.get('etd')), to_date(bl.get('eta')), to_date(bl.get('actual_arrival')),
        bl.get('port'), bl.get('forwarder'),
        bl.get('invoice_number'), bl.get('declaration_number'),
        bl.get('cif_amount_krw'), bl.get('status'),
        '✓' if bl.get('erp_registered') else '', bl.get('memo'),
    ])
write_rows(ws, headers, rows,
           num_cols=[8, 16], date_cols=[9, 10, 11])
autosize(ws, max_w=40)


# ─── Sheet 6: 면장 ──────────────────────────────────────────
ws = wb.create_sheet('면장')
headers = ['면장번호', 'BL 번호', 'CIF KRW', '환율', 'USD/Wp 계약단가',
           '계약총액 USD', '관세율', '관세액 KRW', 'VAT KRW',
           'paid 수량', 'free 수량', '합계 수량', '용량 kW',
           'ERP 입고번호', '부대비용 KRW', '통관수수료 KRW', 'landed_total KRW']
rows = []
for d in decls:
    bl = bl_by_id.get(d.get('bl_id'))
    rows.append([
        d.get('declaration_number'),
        bl['bl_number'] if bl else '?',
        d.get('cif_krw'), d.get('exchange_rate'),
        d.get('contract_unit_price_usd_wp'), d.get('contract_total_usd'),
        d.get('customs_rate'), d.get('customs_amount'), d.get('vat_amount'),
        d.get('paid_qty'), d.get('free_qty'), d.get('quantity'),
        d.get('capacity_kw'), d.get('erp_inbound_no'),
        d.get('incidental_cost'), d.get('customs_fee'), d.get('landed_total_krw'),
    ])
write_rows(ws, headers, rows, num_cols=[3, 6, 8, 9, 10, 11, 12, 13, 15, 16, 17])
autosize(ws, max_w=40)


# ─── Sheet 7: BL별 출고 집계 ────────────────────────────────
ws = wb.create_sheet('BL별 출고집계')
headers = ['BL 번호', '연결 PO', '연결 LC', '회사', '제조사',
           '출고건수', '할당 수량합', '첫 출고일', '마지막 출고일',
           '출고처 수', '매출 출고건수', '시공 출고건수', '매출 공급가합 KRW',
           '출고처 목록']
rows = []
for o in obi_agg:
    bl = bl_by_id.get(o['bl_id'])
    if not bl:
        continue
    po = po_by_id.get(bl.get('po_id'))
    lc = lc_by_id.get(bl.get('lc_id'))
    rows.append([
        bl.get('bl_number'),
        po['po_number'] if po else '',
        lc.get('lc_number') if lc else '',
        bl.get('company_name'), bl.get('manufacturer'),
        int(num(o['outbound_count'])),
        num(o['allocated_qty']),
        to_date(o.get('first_outbound_date')),
        to_date(o.get('last_outbound_date')),
        int(num(o.get('distinct_sites'))),
        int(num(o.get('sale_outbounds'))),
        int(num(o.get('construction_outbounds'))),
        num(o.get('supply_amount_sum')),
        o.get('site_list'),
    ])
write_rows(ws, headers, rows,
           num_cols=[6, 7, 10, 11, 12, 13], date_cols=[8, 9])
autosize(ws, max_w=60)


# ─── Sheet 8: 전체 사슬 (Long format) ───────────────────────
ws = wb.create_sheet('전체 사슬')
headers = ['PO 번호', '회사', '제조사', 'PO 계약일', 'PO 총수량', 'PO MW', 'PO 상태',
           'LC 번호', 'LC 은행', 'LC 개설일', 'LC 금액 USD', 'LC 만기일', 'LC 상환',
           'BL 번호', 'BL inbound_type', 'BL ETD', 'BL ETA', 'BL 입항',
           'BL CIF KRW', 'BL 상태',
           '면장번호', '면장 paid', '면장 free', '면장 landed KRW',
           '출고건수', '출고 수량합', '매출 공급가합 KRW', '단계']
ws.append(headers)
style_header(ws)

for po in pos:
    pid = po['po_id']
    po_row_base = [
        po['po_number'], po.get('company_name'), po.get('manufacturer'),
        to_date(po.get('contract_date')), po.get('total_qty'), po.get('total_mw'),
        po.get('status'),
    ]
    lcs_p = lcs_by_po.get(pid, [])
    bls_p = bls_for_po(pid)

    # 매핑: LC 별로 어떤 BL 들이 묶이는지
    bls_under_lc = defaultdict(list)
    bls_no_lc = []
    for bl in bls_p:
        if bl.get('lc_id') and bl['lc_id'] in {l['lc_id'] for l in lcs_p}:
            bls_under_lc[bl['lc_id']].append(bl)
        else:
            bls_no_lc.append(bl)

    def emit(po_row, lc, bl):
        lc_row = ['', '', '', '', '', ''] if lc is None else [
            lc.get('lc_number'), lc.get('bank_name'),
            to_date(lc.get('open_date')), num(lc.get('amount_usd')) or '',
            to_date(lc.get('maturity_date')),
            '✓' if lc.get('repaid') else '',
        ]
        if bl is None:
            bl_row = [''] * 7
            decls_b = []
            obi_b = None
        else:
            bl_row = [
                bl.get('bl_number'), bl.get('inbound_type'),
                to_date(bl.get('etd')), to_date(bl.get('eta')),
                to_date(bl.get('actual_arrival')),
                num(bl.get('cif_amount_krw')) or '', bl.get('status'),
            ]
            decls_b = decls_by_bl.get(bl['bl_id'], [])
            obi_b = obi_by_bl.get(bl['bl_id'])

        if decls_b:
            for d in decls_b:
                stage = 4
                if obi_b and int(num(obi_b['outbound_count'])) > 0:
                    stage = 5
                if obi_b and num(obi_b.get('supply_amount_sum')) > 0:
                    stage = 6
                row = po_row + lc_row + bl_row + [
                    d.get('declaration_number'),
                    d.get('paid_qty'), d.get('free_qty'),
                    d.get('landed_total_krw'),
                    int(num(obi_b['outbound_count'])) if obi_b else 0,
                    num(obi_b['allocated_qty']) if obi_b else 0,
                    num(obi_b.get('supply_amount_sum')) if obi_b else 0,
                    STAGE_LABEL[stage],
                ]
                ws.append(row)
                ws.cell(row=ws.max_row, column=len(headers)).fill = STAGE_FILL[stage]
        else:
            stage = 1
            if lc is not None: stage = 2
            if bl is not None: stage = max(stage, 3)
            row = po_row + lc_row + bl_row + ['', '', '', '', 0, 0, 0, STAGE_LABEL[stage]]
            ws.append(row)
            ws.cell(row=ws.max_row, column=len(headers)).fill = STAGE_FILL[stage]

    if not lcs_p and not bls_p:
        emit(po_row_base, None, None)
    else:
        for lc in lcs_p:
            bls_l = bls_under_lc.get(lc['lc_id'], [])
            if not bls_l:
                emit(po_row_base, lc, None)
            else:
                for bl in bls_l:
                    emit(po_row_base, lc, bl)
        for bl in bls_no_lc:
            emit(po_row_base, None, bl)

# 숫자/날짜 포맷
num_cols_long = [5, 6, 11, 19, 22, 23, 24, 25, 26, 27]
date_cols_long = [4, 10, 12, 16, 17, 18]
for col_idx in num_cols_long:
    for cell in ws[get_column_letter(col_idx)][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
for col_idx in date_cols_long:
    for cell in ws[get_column_letter(col_idx)][1:]:
        cell.number_format = 'yyyy-mm-dd'

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.font.name != '맑은 고딕':
            cell.font = BODY_FONT
        cell.border = BORDER

ws.freeze_panes = 'B2'
autosize(ws, max_w=30)


# ─── 안내 시트 (맨 앞으로) ─────────────────────────────────
ws = wb.create_sheet('읽는법', 0)
ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='PO 라이프사이클 — 발주 → LC → BL → 면장 → 출고 → 매출')
c.font = Font(name='맑은 고딕', bold=True, size=14)
c.alignment = Alignment(horizontal='center', vertical='center')

guide = [
    ('생성일', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('출처', 'gx10 운영 DB (Supabase) — 직접 쿼리'),
    ('PO 총 건수', f'{len(pos)} 건 (탑솔라/디원/화신 합산)'),
    ('LC 총 건수', f'{len(lcs)} 건  (PO 연결 {sum(1 for l in lcs if l.get("po_id"))}건)'),
    ('BL 총 건수', f'{len(bls)} 건  (PO 직접 연결 {sum(1 for b in bls if b.get("po_id"))} / LC 연결 {sum(1 for b in bls if b.get("lc_id"))})'),
    ('면장 총 건수', f'{len(decls)} 건'),
    ('출고가 있는 BL', f'{len(obi_agg)} 건'),
    ('', ''),
    ('시트 안내', ''),
    ('  1. 요약', 'PO 1행 = 라이프사이클 stage 별 요약. 진행단계 컬럼 색으로 어디까지 갔는지 표시.'),
    ('  2. PO 헤더', 'purchase_orders 원본. po_id 컬럼은 다른 시트 조인용.'),
    ('  3. PO 라인', 'po_line_items — PO 안 품목/수량/단가.'),
    ('  4. LC 마스터', 'lc_records — 은행/개설일/만기/상환 여부.'),
    ('  5. BL 마스터', 'bl_shipments — ETD/ETA/입항/면장번호/CIF.'),
    ('  6. 면장', 'import_declarations + cost_details — paid/free 수량, 부대비용, landed.'),
    ('  7. BL별 출고집계', 'outbound_bl_items 그룹 — BL 1개당 출고건수/수량/매출.'),
    ('  8. 전체 사슬', 'PO×LC×BL×면장 펼친 long format. 한 행 = 라이프사이클의 한 leaf.'),
    ('', ''),
    ('진행단계 색', ''),
    ('  1. PO 만', '빨강 — 발주만 등록, LC 미개설'),
    ('  2. + LC', '주황 — LC 개설됨'),
    ('  3. + BL', '노랑 — 선적/BL 발행됨'),
    ('  4. + 면장', '하늘 — 통관 면장 발급됨'),
    ('  5. + 출고', '파랑 — 창고 출고 시작됨'),
    ('  6. + 매출', '초록 — 매출 인식됨 (외부 판매)'),
    ('', ''),
    ('주의', 'BL ↔ PO 연결은 (a) 직접 po_id, (b) LC 경유 lc_id → po_id 두 경로. 요약/전체사슬 시트는 두 경로 합집합 사용.'),
]
for label, value in guide:
    ws.append([label, value])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 90
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
# 진행단계 색 표 영역에 배경색
stage_rows = {19: 1, 20: 2, 21: 3, 22: 4, 23: 5, 24: 6}
for r, st in stage_rows.items():
    ws.cell(row=r, column=2).fill = STAGE_FILL[st]


OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f'WROTE: {OUT}  ({OUT.stat().st_size:,} bytes)')
