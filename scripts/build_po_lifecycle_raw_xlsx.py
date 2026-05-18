"""PO 라이프사이클 통합 엑셀 (raw 자료 기반).

DB 가 아니라 운영자 정본 (수입진행상황 + 통합정리자료 + BL별 출고현황) 을 정본으로 사용.

입력 JSON: C:/Users/user/AppData/Local/Temp/sf_raw/*.json
출력:     C:/Users/user/Downloads/PO_라이프사이클_raw_YYYYMMDD.xlsx
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path('C:/Users/user/AppData/Local/Temp/sf_raw')
OUT = Path('C:/Users/user/Downloads') / f'PO_라이프사이클_raw_{datetime.now():%Y%m%d}.xlsx'


def load(name):
    return json.loads((SRC / f'{name}.json').read_text(encoding='utf-8'))


def to_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, date) else v.date()
    s = str(v)[:10]
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def first_date_str(*candidates):
    """여러 후보 중 첫 유효 날짜 반환 (date 객체)."""
    for c in candidates:
        d = to_date(c)
        if d:
            return d
    return None


def cell_str(v):
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def has(v):
    return cell_str(v) not in ('', 'None')


# ─── 스타일 ───────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
SUB_HEADER_FILL = PatternFill('solid', fgColor='374151')
BODY_FONT = Font(name='맑은 고딕', size=10)
NOTE_FONT = Font(name='맑은 고딕', size=9, italic=True, color='6B7280')
BORDER = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)
STAGE_FILL = {
    1: PatternFill('solid', fgColor='FEE2E2'),  # red
    2: PatternFill('solid', fgColor='FED7AA'),  # orange
    3: PatternFill('solid', fgColor='FEF3C7'),  # amber
    4: PatternFill('solid', fgColor='DBEAFE'),  # blue
    5: PatternFill('solid', fgColor='BFDBFE'),  # blue-200
    6: PatternFill('solid', fgColor='BBF7D0'),  # green
}
STAGE_LABEL = {
    1: '1. PO 발주',
    2: '2. + LC 개설',
    3: '3. + 선적 ETD',
    4: '4. + 입항 ETA',
    5: '5. + 통관/BL',
    6: '6. + 배송 완료',
}


def style_header(ws, row=1, fill=HEADER_FILL):
    for cell in ws[row]:
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_w=45, min_w=6, sample=300):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = min_w
        for cell in col[:sample]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            max_len = max(max_len, w)
        ws.column_dimensions[letter].width = min(max_w, max_len + 2)


def apply_body_style(ws, start_row=2, num_cols=(), date_cols=()):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
    for c in num_cols:
        for cell in ws[get_column_letter(c)][start_row - 1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
    for c in date_cols:
        for cell in ws[get_column_letter(c)][start_row - 1:]:
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = 'yyyy-mm-dd'


# ─── 데이터 로드 ─────────────────────────────────────────────
s_2024 = load('raw25_2024')
s_2025_full = load('raw25_2025')
s_25po = load('integrated_s11')      # 통합본 25년 PO (정리됨)
s_lc_master = load('integrated_s12')  # 통합본 LC 마스터
s_26po = load('integrated_s16')      # 통합본 26년 PO
s_bl_cost = load('integrated_s13' if (SRC / 'integrated_s13.json').exists() else 'integrated_s19')
s_bl_pum = load('raw25_품의서_서식1')
s_forex_ts = load('raw25_외환')         # 탑솔라 LC
s_forex_dw = load('raw25_디원_외환')      # 디원 LC

# 26 원본 제조사별
s_26jinko = load('raw26_징코')
s_26longi = load('raw26_론지솔라')
s_26trina = load('raw26_트리나')
s_26risen = load('raw26_라이젠')
s_26sum1 = load('raw26_Sheet1')
s_26sum2 = load('raw26_Sheet2')

# BL ↔ 출고
blo_jinko = load('blo_진코솔라')
blo_jinko2 = load('blo_진코솔라_2')
blo_ja = load('blo_JA솔라')
blo_trina = load('blo_트리나솔라')
blo_risen = load('blo_라이젠에너지')
blo_longi = load('blo_론지솔라')

print(f'24년 raw 2024 시트: {len(s_2024)}행')
print(f'25년 raw 2025 시트: {len(s_2025_full)}행')
print(f'25년 통합 s11:  {len(s_25po)}행')
print(f'26년 통합 s16:  {len(s_26po)}행')
print(f'BL 운송료 s19:  {len(s_bl_cost)}행')

# ─── 24년 PO 추출 (raw 2024 시트에서) ────────────────────────
# 컬럼: 0=No, 1=발주건명, 2=업체, 3=P/O No, 4=품명, 5=pcs, 6=F/M, 7=Wp,
#       8=Unit price, 9=Amount, 10=LC 은행, 11=LC 개설일, 12=L/C No,
#       13=LC 수량, 14=LC F/M, 15=LC W.P, 16=LC 개설금액, 17=LC 만기일, 18=유산스 금액,
#       19=ETD, 20=ETA, 21=수량 F/M包, 22=F/S, 23=PLT, 24=1PLT 수량, 25=포장40", 26=반출기한,
#       27=L/G 발행, 28=통관, 29=(?), 30=포워더, 31=B/L No,
#       32=현장배송 입고일, 33=현장, 34=Wp단가, 35=수량
COL_24 = dict(
    no=0, batch=1, vendor=2, po_no=3, product=4,
    pcs=5, fm=6, wp=7, unit_price=8, amount=9,
    lc_bank=10, lc_open=11, lc_no=12, lc_qty=13, lc_fm=14, lc_wp=15,
    lc_open_amount=16, lc_maturity=17, usance_amount=18,
    etd=19, eta=20, qty_fmpkg=21, fs=22, plt=23, plt_qty=24,
    container_40=25, lg_issue=26, customs=27,
    bl_issue_date=28, forwarder=29, bl_no=30,
    delivery_date=31, delivery_site=32, delivery_wp=33, delivery_qty=34,
)
COL_25 = COL_24.copy()  # 25년 시트는 거의 동일 구조 — 데이터 시트에서 직접 fill

# 25년 시트는 컬럼이 1개 다름: 인덱스 16 (LC 금액) + 17 (LC 개설금액) — 16과 17이 한 칸 밀림
# 실제 헤더: ...16=금액, 17=개설금액, 18=만기일, 19=유산스금액, 20=선적, 21=입항...
# 24년: 16=LC 개설금액, 17=LC 만기일, 18=유산스 금액, 19=선적, 20=입항...
# 즉 25년은 16에 '금액' 컬럼이 추가됐고 그 뒤로 한 칸씩 밀림
COL_25.update(dict(
    lc_amount_alt=16,     # 25년만의 '금액' 컬럼
    lc_open_amount=17,    # 25년의 개설금액
    lc_maturity=18,
    usance_amount=19,
    etd=20, eta=21,
    qty_fmpkg=22, fs=23, plt=25, plt_qty=26,
    container_40=27, release_due=28, lg_issue=29, customs=30,
    forwarder=31, bl_no=32,
    delivery_date=33, delivery_site=34, delivery_wp=None, delivery_qty=None,
))


def extract_pos_from_2024_style(sheet_rows, col_map, year_tag, sheet_label):
    """2024/2025 raw 시트에서 PO 추출.

    PO 헤더 행 = No 컬럼이 정수이고 P/O No 가 채워진 행.
    같은 PO 의 추가 LC 행 = No 컬럼 비어있으나 LC No / Amount 등이 채워진 행.
    같은 PO 의 분할 배송 행 = 위 두 가지 아닌 행 중 delivery_date 또는 site 가 있는 행.
    """
    pos = []
    current = None
    for row_idx, row in enumerate(sheet_rows):
        def g(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]
        no = g('no')
        po_no = cell_str(g('po_no'))
        # PO No 패턴 식별
        has_letter_or_hangul = bool(re.search(r'[A-Za-z]{3,}', po_no)) or bool(re.search(r'[가-힣]{3,}', po_no))
        looks_like_po = (
            has(po_no)
            and len(po_no.strip()) >= 4
            and has_letter_or_hangul
            and po_no not in ('P/O No.', 'P/O No', 'pcs', '품명', 'No.', 'No')
            and not po_no.startswith('=')
            and not re.match(r'^[=\d\s,.\-+()*/]+$', po_no)
            and po_no.strip() not in ('컨', '수량', '보험', '매매기준율')
            and not po_no.endswith('pcs')
            and '*' not in po_no  # 공식 거름
        )
        is_po_header = looks_like_po
        if is_po_header:
            if current:
                pos.append(current)
            current = {
                'year': year_tag, 'source': sheet_label, 'row_no': row_idx + 1,
                'no': no, 'batch': g('batch'), 'vendor': g('vendor'),
                'po_no': po_no, 'product': g('product'),
                'pcs': g('pcs'), 'fm': g('fm'), 'wp': g('wp'),
                'unit_price': g('unit_price'), 'amount': g('amount'),
                'lc_bank': g('lc_bank'), 'lc_open': g('lc_open'), 'lc_no': g('lc_no'),
                'lc_qty': g('lc_qty'), 'lc_open_amount': g('lc_open_amount'),
                'lc_maturity': g('lc_maturity'), 'usance_amount': g('usance_amount'),
                'etd': g('etd'), 'eta': g('eta'), 'qty_fmpkg': g('qty_fmpkg'),
                'plt': g('plt'), 'container_40': g('container_40'),
                'lg_issue': g('lg_issue'), 'customs': g('customs'),
                'bl_issue_date': g('bl_issue_date'), 'forwarder': g('forwarder'),
                'bl_no': g('bl_no'),
                'deliveries': [], 'lc_extras': [],
            }
            if has(g('delivery_date')) or has(g('delivery_site')):
                current['deliveries'].append({
                    'date': g('delivery_date'), 'site': g('delivery_site'),
                    'qty': g('delivery_qty'),
                })
        elif current:
            # 같은 PO 의 추가 LC 또는 분할 배송
            lc_no = g('lc_no')
            lc_open = g('lc_open')
            etd = g('etd')
            eta = g('eta')
            bl_no = g('bl_no')
            if has(lc_no) or has(lc_open) or has(etd) or has(eta) or has(bl_no):
                current['lc_extras'].append({
                    'lc_bank': g('lc_bank'), 'lc_open': lc_open, 'lc_no': lc_no,
                    'lc_open_amount': g('lc_open_amount'), 'lc_maturity': g('lc_maturity'),
                    'etd': etd, 'eta': eta, 'forwarder': g('forwarder'), 'bl_no': bl_no,
                    'pcs': g('pcs'), 'amount': g('amount'),
                    'customs': g('customs'),
                })
            # 배송 정보도 추출 (분할 배송)
            d = g('delivery_date')
            s_ = g('delivery_site')
            q = g('delivery_qty')
            if has(d) or has(s_) or has(q):
                current['deliveries'].append({'date': d, 'site': s_, 'qty': q})
    if current:
        pos.append(current)
    return pos


pos_24 = extract_pos_from_2024_style(s_2024, COL_24, 2024, '2024시트(raw)')
pos_25_raw = extract_pos_from_2024_style(s_2025_full, COL_25, 2025, '2025시트(raw)')
print(f'24년 PO 추출: {len(pos_24)}건')
print(f'25년 raw 추출: {len(pos_25_raw)}건')


# ─── 25년 통합본 s11 도 추출 (보강 정본) ─────────────────────
# s11 columns: 0=No, 1=발주처, 2=업체, 3=P/O No, 4=품명, 5=pcs, 6=Unit$, 7=Amount$,
#              8=LC 은행, 9=LC 개설일, 10=LC No, 11=LC 금액, 12=LC 만기,
#              13=ETD, 14=ETA, 15=FR 수량, 16=포워더, 17=B/L No
def extract_pos_from_integrated_s11(rows):
    pos = []
    current = None
    for row_idx, row in enumerate(rows):
        if not row or len(row) < 4:
            continue
        po_no = cell_str(row[3])
        no = row[0]
        if has(no) and has(po_no) and po_no != 'P/O No':
            if current:
                pos.append(current)
            current = {
                'year': 2025, 'source': 's11.E1.25년PO(통합본)', 'row_no': row_idx + 1,
                'no': no, 'batch': row[1], 'vendor': row[2], 'po_no': po_no,
                'product': row[4], 'pcs': row[5], 'fm': None, 'wp': None,
                'unit_price': row[6], 'amount': row[7],
                'lc_bank': row[8], 'lc_open': row[9], 'lc_no': row[10],
                'lc_qty': None, 'lc_open_amount': row[11], 'lc_maturity': row[12],
                'usance_amount': None,
                'etd': row[13], 'eta': row[14], 'qty_fmpkg': row[15],
                'plt': None, 'container_40': None, 'lg_issue': None,
                'customs': None, 'bl_issue_date': None, 'forwarder': row[16],
                'bl_no': row[17],
                'deliveries': [],
                'lc_extras': [],  # 추가 LC (같은 PO 의 행 2+)
            }
        elif current:
            # 같은 PO 의 추가 LC / BL / ETD 행
            extra = {
                'lc_open': row[9] if len(row) > 9 else None,
                'lc_no': row[10] if len(row) > 10 else None,
                'lc_open_amount': row[11] if len(row) > 11 else None,
                'lc_maturity': row[12] if len(row) > 12 else None,
                'etd': row[13] if len(row) > 13 else None,
                'eta': row[14] if len(row) > 14 else None,
                'bl_no': row[17] if len(row) > 17 else None,
                'forwarder': row[16] if len(row) > 16 else None,
            }
            if any(has(v) for v in extra.values()):
                current['lc_extras'].append(extra)
    if current:
        pos.append(current)
    return pos


pos_25_int = extract_pos_from_integrated_s11(s_25po)
print(f'25년 통합본 s11 PO 추출: {len(pos_25_int)}건')


# ─── 26년 통합본 s16 추출 ────────────────────────────────────
# s16 columns: 0=제조사, 1=No, 2=업체, 3=P/O No, 4=품명, 5=pcs, 6=F/M, 7=Wp,
#              8=Unit price, 9=Amount, 10=L/C No, 11=선적, 12=입항, 13=B/L No,
#              14=현장, 15=입고일
def extract_pos_from_integrated_s16(rows):
    pos = []
    current = None
    for row_idx, row in enumerate(rows):
        if not row or len(row) < 4:
            continue
        po_no = cell_str(row[3])
        if has(po_no) and po_no != 'P/O No.':
            if current:
                pos.append(current)
            current = {
                'year': 2026, 'source': 's16.E2.26년PO(통합본)', 'row_no': row_idx + 1,
                'manufacturer': row[0], 'no': None, 'batch': None, 'vendor': row[2],
                'po_no': po_no, 'product': row[4],
                'pcs': row[5], 'fm': row[6], 'wp': row[7],
                'unit_price': row[8], 'amount': row[9],
                'lc_bank': None, 'lc_open': None, 'lc_no': row[10],
                'lc_qty': None, 'lc_open_amount': None, 'lc_maturity': None,
                'usance_amount': None,
                'etd': row[11], 'eta': row[12], 'qty_fmpkg': None,
                'plt': None, 'container_40': None, 'lg_issue': None,
                'customs': None, 'bl_issue_date': None, 'forwarder': None,
                'bl_no': row[13],
                'deliveries': [],
            }
            if has(row[14]) or has(row[15]):
                current['deliveries'].append({'date': row[15], 'site': row[14], 'qty': None})
        elif current and len(row) >= 16:
            d, s, q = row[15], row[14], None
            extra_bl = row[13] if len(row) > 13 else None
            if has(d) or has(s):
                current['deliveries'].append({'date': d, 'site': s, 'qty': None})
            if has(extra_bl) and current.get('bl_no') and extra_bl != current['bl_no']:
                # 같은 PO 의 추가 BL
                current['bl_no'] = f"{current['bl_no']} / {extra_bl}"
    if current:
        pos.append(current)
    return pos


pos_26 = extract_pos_from_integrated_s16(s_26po)
print(f'26년 통합본 s16 PO 추출: {len(pos_26)}건')


# ─── BL ↔ 출고 매핑 (bl_outbound 6시트) ───────────────────────
def parse_bl_outbound(rows, header_idx, manufacturer):
    """B/L 컬럼 인덱스를 찾아서 BL → 출고 raw 행 반환."""
    if header_idx >= len(rows):
        return []
    header = rows[header_idx]
    # B/L 컬럼 찾기
    bl_col = None
    site_col = None
    qty_col = None
    date_col = None
    for i, h in enumerate(header):
        if h is None: continue
        hs = str(h).strip().lower()
        if 'b/l' in hs or 'bl' == hs:
            bl_col = i
        if '출고지' in str(h) or '지역' in str(h):
            if site_col is None:
                site_col = i
        if '출고수량' in str(h):
            qty_col = i
        if '출고일' in str(h):
            date_col = i
    if bl_col is None:
        return []
    out = []
    for r in rows[header_idx + 1:]:
        if not r or bl_col >= len(r):
            continue
        bl = r[bl_col]
        if not has(bl):
            continue
        out.append({
            'manufacturer': manufacturer,
            'bl': cell_str(bl),
            'site': cell_str(r[site_col]) if site_col is not None and site_col < len(r) else '',
            'qty': r[qty_col] if qty_col is not None and qty_col < len(r) else None,
            'date': r[date_col] if date_col is not None and date_col < len(r) else None,
        })
    return out


blo_all = []
blo_all += parse_bl_outbound(blo_jinko, 1, '징코')
blo_all += parse_bl_outbound(blo_jinko2, 2, '징코(25년)')
blo_all += parse_bl_outbound(blo_ja, 2, 'JA솔라')
blo_all += parse_bl_outbound(blo_trina, 2, '트리나')
blo_all += parse_bl_outbound(blo_risen, 2, '라이젠')
blo_all += parse_bl_outbound(blo_longi, 1, '론지')

# BL 별 그룹
blo_by_bl = defaultdict(list)
for x in blo_all:
    blo_by_bl[x['bl']].append(x)
print(f'BL 출고 매핑: {len(blo_all)}건, distinct BL {len(blo_by_bl)}개')


# ─── 라이프사이클 단계 매기기 ────────────────────────────────
def lifecycle_stage(po):
    stage = 1  # PO 만 있음
    if has(po.get('lc_no')):
        stage = 2
    if to_date(po.get('etd')):
        stage = max(stage, 3)
    if to_date(po.get('eta')):
        stage = max(stage, 4)
    if has(po.get('bl_no')) or to_date(po.get('customs')):
        stage = max(stage, 5)
    # 배송 완료 여부 — deliveries 가 있으면 6
    if po.get('deliveries'):
        stage = max(stage, 6)
    return stage


# 모든 PO 통합 — 사용자 요청: raw 우선
# 25년은 raw 2025 시트 (배송 분할까지 다 있음) 우선
all_pos = pos_24 + pos_25_raw + pos_26
print(f'전체 통합 PO: {len(all_pos)}건')


# ─── workbook 작성 ───────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)


# ─── Sheet 1: 읽는법 ────────────────────────────────────────
ws = wb.create_sheet('읽는법')
ws.merge_cells('A1:B1')
c = ws.cell(row=1, column=1, value='PO 라이프사이클 — 발주 → LC → 선적 → 입항 → 통관/BL → 배송')
c.font = Font(name='맑은 고딕', bold=True, size=14)
c.alignment = Alignment(horizontal='center', vertical='center')

guide = [
    ('생성일', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('정본', '운영자 raw 자료 (수입진행상황 25/26 + 통합정리자료 + BL별 출고현황)'),
    ('출처 파일',
     '• 수입진행상황(module)-2025년도.xlsx (191MB) — 24/25년 PO 마스터 + 외환/디원외환 LC + 품의서\n'
     '• 수입진행상황(module)-2026년도.xlsx — 26년 제조사별 트래커 (징코/론지/트리나/라이젠)\n'
     '• 솔라플로우_통합정리자료_2026-05-15.xlsx — raw 정리본 (시트 11/12/13/16/19)\n'
     '• (수입) BL별 출고현황리스트.xlsx — BL ↔ 현장 출고 매핑'),
    ('', ''),
    ('PO 통계', f'24년 raw {len(pos_24)}건 + 25년 raw {len(pos_25_raw)}건 + 26년 통합 {len(pos_26)}건 = 총 {len(all_pos)}건 (25년 통합본 s11 12건은 별도 정리본 시트로 보존)'),
    ('BL 출고 매핑', f'{len(blo_all)}건 (distinct BL {len(blo_by_bl)}개)'),
    ('', ''),
    ('시트 안내', ''),
    ('  1. 읽는법', '본 페이지'),
    ('  2. PO 라이프사이클', '한 PO = 한 행. 진행단계 색으로 표시'),
    ('  3. 24년 PO (raw)', 'import_2025.xlsx::2024 시트 그대로'),
    ('  4. 25년 PO (raw)', 'import_2025.xlsx::2025 시트 그대로 — PO 라이프사이클 정본'),
    ('  5. 25년 PO 정리본', '통합정리자료 s11 — 운영자가 raw 를 정리한 깔끔한 버전 (참고용)'),
    ('  6. 26년 PO 정리본', '통합정리자료 s16 — 26년 PO 트래커 통합 정리'),
    ('  7. 26년 징코 (raw)', 'import_2026.xlsx::징코 시트 — 광범한 컬럼 (46열)'),
    ('  8. 26년 론지 (raw)', '동일 (48열, T/T 10%/90% 분할)'),
    ('  9. 26년 트리나 (raw)', '동일'),
    ('  10. 26년 라이젠 (raw)', '동일'),
    ('  11. LC 마스터 (탑솔라)', 'import_2025.xlsx::외환 시트 — 탑솔라(주) LC 현황'),
    ('  12. LC 마스터 (디원)', 'import_2025.xlsx::디원 외환 시트'),
    ('  13. BL별 CIF·운송료', 'import_2025.xlsx::품의서 서식1 — BL 정산 양식 (한 BL 한 블록)'),
    ('  14. BL↔출고 통합', '(수입) BL별 출고현황리스트 6개 제조사 시트 통합'),
    ('  15. BL별 운송료청구', '통합본 s19 — BL별 운송료 청구서 50+ 건'),
    ('', ''),
    ('진행단계 색', ''),
    ('  1. PO 발주만', '빨강 — LC 미개설'),
    ('  2. + LC 개설', '주황 — LC No 채워짐'),
    ('  3. + 선적 ETD', '노랑 — 출항 일자 있음'),
    ('  4. + 입항 ETA', '하늘 — 도착 일자 있음'),
    ('  5. + 통관/BL', '파랑 — B/L No 또는 통관 일자 있음'),
    ('  6. + 배송 완료', '초록 — 현장 배송 분할 행 1+ 있음'),
    ('', ''),
    ('주의 — raw 자료의 모호함', ''),
    ('  • PO 정의', '운영자가 "발주건" 단위로 1행. 한 PO 가 여러 LC 로 분할 또는 여러 BL 로 분할 가능 — '
                  '본 자료는 PO 헤더 행의 LC/BL 만 추출하고, 나머지는 raw 시트에서 확인.'),
    ('  • 25년 데이터 출처', '통합본 s11(정리됨) 우선. raw 시트와 중복되면 통합본 사용, raw 에만 있으면 raw 채택.'),
    ('  • 배송 완료 = stage 6', '배송 행 1+ 있으면 6으로 간주. 실제 PO 수량 vs 배송 수량 합계 검증은 미수행.'),
    ('  • LC 매핑 한계', '한 PO 가 여러 LC 로 분할되는 경우, "라이프사이클" 시트는 첫 LC 만 표시. '
                       '나머지 LC 는 "LC 마스터 (탑솔라/디원)" 시트 또는 raw 시트에서 확인.'),
]
for label, value in guide:
    ws.append([label, value])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 110
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
# 단계 색칠
stage_rows = {22: 1, 23: 2, 24: 3, 25: 4, 26: 5, 27: 6}
for r, st in stage_rows.items():
    ws.cell(row=r, column=2).fill = STAGE_FILL[st]


# ─── Sheet 2: PO 라이프사이클 ────────────────────────────────
ws = wb.create_sheet('PO 라이프사이클')
headers = [
    '연도', 'No', '발주처', '업체(제조사)', 'P/O No', '품명',
    '수량(pcs)', 'F/M', 'Wp', 'Unit$', 'Amount$',
    'LC 은행', 'LC 개설일', 'L/C No', 'LC 개설금액', 'LC 만기일', '유산스 금액',
    'ETD(선적)', 'ETA(입항)', 'F/M包 수량', 'PLT', '40\"',
    '통관 일자', '포워더', 'B/L No',
    '배송 횟수', '첫 배송일', '마지막 배송일', '배송 수량 합', '배송지 목록',
    '진행단계', '원본 시트', '원본 행',
]
ws.append(headers)
style_header(ws)

stage_col_idx = headers.index('진행단계') + 1

def aggregate(po):
    """헤더 행 + lc_extras 의 LC/BL/ETD/ETA 합치기."""
    lcs_no = [po.get('lc_no')] + [e.get('lc_no') for e in po.get('lc_extras', [])]
    lcs_no = [cell_str(x) for x in lcs_no if has(x)]
    bls = [po.get('bl_no')] + [e.get('bl_no') for e in po.get('lc_extras', [])]
    bls = [cell_str(x) for x in bls if has(x)]
    open_dates = [to_date(po.get('lc_open'))] + [to_date(e.get('lc_open')) for e in po.get('lc_extras', [])]
    open_dates = [d for d in open_dates if d]
    mat_dates = [to_date(po.get('lc_maturity'))] + [to_date(e.get('lc_maturity')) for e in po.get('lc_extras', [])]
    mat_dates = [d for d in mat_dates if d]
    etds = [to_date(po.get('etd'))] + [to_date(e.get('etd')) for e in po.get('lc_extras', [])]
    etds = [d for d in etds if d]
    etas = [to_date(po.get('eta'))] + [to_date(e.get('eta')) for e in po.get('lc_extras', [])]
    etas = [d for d in etas if d]
    customs = [to_date(po.get('customs'))] + [to_date(e.get('customs')) for e in po.get('lc_extras', [])]
    customs = [d for d in customs if d]
    open_amounts = [po.get('lc_open_amount')] + [e.get('lc_open_amount') for e in po.get('lc_extras', [])]
    open_amount_sum = 0.0
    for v in open_amounts:
        if isinstance(v, (int, float)):
            open_amount_sum += float(v)
    return {
        'lc_no_join': ' | '.join(dict.fromkeys(lcs_no)),
        'bl_no_join': ' | '.join(dict.fromkeys(bls)),
        'open_date_range': fmt_range(open_dates),
        'maturity_range': fmt_range(mat_dates),
        'etd_range': fmt_range(etds),
        'eta_range': fmt_range(etas),
        'customs_range': fmt_range(customs),
        'open_amount_sum': open_amount_sum if open_amount_sum else None,
        'lc_count': len(set(lcs_no)),
        'bl_count': len(set(bls)),
    }

def fmt_range(dates):
    if not dates:
        return ''
    dates = sorted(set(dates))
    if dates[0] == dates[-1]:
        return dates[0].isoformat()
    return f'{dates[0].isoformat()} ~ {dates[-1].isoformat()}'

# stage 함수도 lc_extras 고려해서 강화
def lifecycle_stage2(po):
    agg = aggregate(po)
    stage = 1
    if agg['lc_no_join']:
        stage = 2
    if agg['etd_range']:
        stage = max(stage, 3)
    if agg['eta_range']:
        stage = max(stage, 4)
    if agg['bl_no_join'] or agg['customs_range']:
        stage = max(stage, 5)
    if po.get('deliveries'):
        stage = max(stage, 6)
    return stage

for po in all_pos:
    deliveries = po.get('deliveries', [])
    dd = [to_date(d.get('date')) for d in deliveries]
    dd = [d for d in dd if d]
    delivery_qty_sum = sum(
        (float(d['qty']) if isinstance(d.get('qty'), (int, float)) else 0)
        for d in deliveries
    )
    sites = []
    seen = set()
    for d in deliveries:
        s = cell_str(d.get('site'))
        if s and s not in seen:
            seen.add(s)
            sites.append(s)

    agg = aggregate(po)
    stage = lifecycle_stage2(po)
    row = [
        po.get('year'),
        po.get('no'),
        po.get('batch') or po.get('manufacturer'),
        po.get('vendor') or po.get('manufacturer'),
        po.get('po_no'),
        po.get('product'),
        po.get('pcs'),
        po.get('fm'),
        po.get('wp'),
        po.get('unit_price'),
        po.get('amount'),
        po.get('lc_bank'),
        agg['open_date_range'],
        f"{agg['lc_count']}건: {agg['lc_no_join']}" if agg['lc_count'] > 1 else agg['lc_no_join'],
        agg['open_amount_sum'],
        agg['maturity_range'],
        po.get('usance_amount'),
        agg['etd_range'],
        agg['eta_range'],
        po.get('qty_fmpkg'),
        po.get('plt'),
        po.get('container_40'),
        agg['customs_range'],
        po.get('forwarder'),
        f"{agg['bl_count']}건: {agg['bl_no_join']}" if agg['bl_count'] > 1 else agg['bl_no_join'],
        len(deliveries) if deliveries else None,
        min(dd) if dd else None,
        max(dd) if dd else None,
        delivery_qty_sum if delivery_qty_sum else None,
        ' | '.join(sites) if sites else '',
        STAGE_LABEL[stage],
        po.get('source'),
        po.get('row_no'),
    ]
    ws.append(row)
    ws.cell(row=ws.max_row, column=stage_col_idx).fill = STAGE_FILL[stage]
    ws.cell(row=ws.max_row, column=stage_col_idx).font = Font(name='맑은 고딕', size=10, bold=True)

apply_body_style(ws, num_cols=[2, 7, 8, 9, 10, 11, 15, 20, 21, 22, 26, 29],
                 date_cols=[13, 16, 18, 19, 23, 27, 28])
ws.freeze_panes = 'F2'
autosize(ws, max_w=35)


# ─── raw 시트들 복사 헬퍼 ────────────────────────────────────
def copy_raw_sheet(rows, sheet_name, max_rows=None, has_header_at_row=1):
    ws = wb.create_sheet(sheet_name)
    if max_rows:
        rows = rows[:max_rows]
    for r in rows:
        # date string → date object 로 복원
        nr = []
        for v in r:
            if isinstance(v, str) and re.match(r'^\d{4}-\d{2}-\d{2}', v):
                d = to_date(v)
                nr.append(d if d else v)
            else:
                nr.append(v)
        ws.append(nr)
    if ws.max_row >= has_header_at_row:
        style_header(ws, row=has_header_at_row)
    apply_body_style(ws, start_row=has_header_at_row + 1)
    ws.freeze_panes = f'A{has_header_at_row + 1}'
    autosize(ws, max_w=30, sample=80)


# ─── Sheet 3~8: raw 시트 복사 ────────────────────────────────
copy_raw_sheet(s_2024, '24년 PO (raw)', has_header_at_row=3)
copy_raw_sheet(s_2025_full, '25년 PO (raw)', has_header_at_row=3)
copy_raw_sheet(s_26jinko, '26년 징코 (raw)', has_header_at_row=3)
copy_raw_sheet(s_26longi, '26년 론지 (raw)', has_header_at_row=3)
copy_raw_sheet(s_26trina, '26년 트리나 (raw)', has_header_at_row=3)
copy_raw_sheet(s_26risen, '26년 라이젠 (raw)', has_header_at_row=3)
copy_raw_sheet(s_forex_ts, 'LC 마스터 (탑솔라)', has_header_at_row=6)
copy_raw_sheet(s_forex_dw, 'LC 마스터 (디원)', has_header_at_row=6)
copy_raw_sheet(s_bl_pum, 'BL별 CIF·운송료 (품의서)', has_header_at_row=4)


# ─── Sheet 12: BL ↔ 출고 통합 ────────────────────────────────
ws = wb.create_sheet('BL↔출고 통합')
headers = ['제조사', 'B/L No', '출고일', '출고지/지역', '출고수량']
ws.append(headers)
style_header(ws)
# 정렬: 제조사, BL, 출고일
def sortkey(x):
    d = to_date(x['date'])
    return (x['manufacturer'], x['bl'], d or date(1900, 1, 1))
for x in sorted(blo_all, key=sortkey):
    ws.append([
        x['manufacturer'], x['bl'], to_date(x['date']),
        x['site'], x['qty'] if isinstance(x.get('qty'), (int, float)) else x.get('qty'),
    ])
apply_body_style(ws, num_cols=[5], date_cols=[3])
ws.freeze_panes = 'A2'
autosize(ws, max_w=40)


# ─── 통합본 정리 시트들 (s11/s16) ────────────────────────────
copy_raw_sheet(s_25po, '25년 PO 정리본 (통합 s11)', has_header_at_row=2)
copy_raw_sheet(s_26po, '26년 PO 정리본 (통합 s16)', has_header_at_row=4)

# ─── Sheet 13: BL 운송료 청구서 (통합본 s19 그대로) ──────────
copy_raw_sheet(s_bl_cost, 'BL별 운송료청구', has_header_at_row=1)


# ─── 시트 순서 정리 ──────────────────────────────────────────
order = [
    '읽는법',
    'PO 라이프사이클',
    '24년 PO (raw)',
    '25년 PO (raw)',
    '25년 PO 정리본 (통합 s11)',
    '26년 PO 정리본 (통합 s16)',
    '26년 징코 (raw)',
    '26년 론지 (raw)',
    '26년 트리나 (raw)',
    '26년 라이젠 (raw)',
    'LC 마스터 (탑솔라)',
    'LC 마스터 (디원)',
    'BL별 CIF·운송료 (품의서)',
    'BL↔출고 통합',
    'BL별 운송료청구',
]
wb._sheets = [wb[s] for s in order if s in wb.sheetnames]

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f'WROTE: {OUT}  ({OUT.stat().st_size:,} bytes)')
