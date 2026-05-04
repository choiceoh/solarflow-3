"""
재고 시트(94행) → inventory_snapshots + products.safety_stock/available_stock backfill (D-064, PR 20).

정책 (D-064): 안전장치보다 데이터 살림. 충돌 시 ERP 가 더 신뢰도 높음.
- 재고 시트의 'M-XX-NN' erp_code 로 우리 products 매칭 (PR 19 결과 활용)
- inventory_snapshots: 같은 (snapshot_date, product_id, source) 면 UPSERT (덮어쓰기, ERP 최신값 우선)
- products.safety_stock / available_stock: 항상 ERP 값으로 동기화 (NULL 도 OK)
- 시트의 erp_code 가 products 에 없으면 → 신규 product 등록 + 재고도 등록
"""
import os, re, json, sys, datetime, openpyxl, psycopg2
from openpyxl import load_workbook

FILE = "/tmp/erp_data.xlsx"
SHEET = "재고"
SNAPSHOT_DATE = datetime.date(2025, 12, 31)  # ERP 자료 export 시점 추정 — 사용자 확인 가능

# 컬럼 인덱스 (헤더 행 1)
COLS = {
    "erp_code": 0, "model": 1, "spec": 2, "unit": 3,
    "beginning": 4, "inbound": 5, "outbound": 6, "ending": 7,
    "safety": 8, "available": 9, "factor": 10, "mgmt_unit": 11, "ending_mgmt": 12,
}

def normalize_code(s):
    if not s: return ""
    chars = []
    for ch in str(s):
        c = ord(ch)
        if (ord('A') <= c <= ord('Z')) or (ord('0') <= c <= ord('9')):
            chars.append(ch)
        elif ord('a') <= c <= ord('z'):
            chars.append(chr(c - 32))
        elif 0xAC00 <= c <= 0xD7AF or 0x4E00 <= c <= 0x9FFF:
            chars.append(ch)
    return "".join(chars)

def to_int(v):
    if v in (None, ""): return None
    try: return int(float(str(v).replace(",", "")))
    except: return None

def to_float(v):
    if v in (None, ""): return None
    try: return float(str(v).replace(",", ""))
    except: return None

def parse_wp(s):
    if not s: return None
    m = re.search(r'(\d+)', str(s))
    return int(m.group(1)) if m else None

c = psycopg2.connect(os.environ["SUPABASE_DB_URL"])
c.autocommit = False
cur = c.cursor()

# 마스터 캐시 — erp_code → product_id (PR 19 backfill 결과)
cur.execute("SELECT product_id, product_code, erp_code FROM products")
prod_by_erp = {}
prod_by_code = {}
for pid, code, erp in cur.fetchall():
    if erp:
        prod_by_erp[erp] = pid
    if code:
        prod_by_code[normalize_code(code)] = pid

cur.execute("SELECT manufacturer_id, name_kr, short_name FROM manufacturers")
mfg_kr = {}
for mid, kr, short in cur.fetchall():
    if kr: mfg_kr[normalize_code(kr)] = mid
    if short: mfg_kr[normalize_code(short)] = mid

PREFIX_RULES = [
    (re.compile(r"^TSM", re.I), "트리나"),
    (re.compile(r"^LR[5789]", re.I), "론지"),
    (re.compile(r"^JKM", re.I), "징코"),
    (re.compile(r"^JAM", re.I), "ja"),
    (re.compile(r"^RSM", re.I), "라이젠"),
    (re.compile(r"^Q\.?TRON|^QPEAK|^Q\.?PEAK", re.I), "한화"),
    (re.compile(r"^HA", re.I), "한솔"),
    (re.compile(r"^HS", re.I), "한솔"),
    (re.compile(r"^CS", re.I), "캐솔"),
    (re.compile(r"^CWP|^TWMHF", re.I), "통웨이"),
]

def infer_mfg(model):
    for pat, kw in PREFIX_RULES:
        if pat.match(model):
            return mfg_kr.get(normalize_code(kw))
    return None

# 재고 시트 처리
wb = load_workbook(FILE, data_only=True)
ws = wb[SHEET]
rows = list(ws.iter_rows(values_only=True))
header_idx = -1
for i, row in enumerate(rows[:5]):
    if row and any(str(c).strip() == "품번" for c in row if c):
        header_idx = i
        break
if header_idx < 0:
    print("헤더 행 미발견"); sys.exit(1)

snapshot_inserted = 0
snapshot_updated = 0
products_synced = 0
new_products = 0
errors = []

for ridx, row in enumerate(rows[header_idx+1:], start=header_idx+2):
    erp = row[COLS["erp_code"]]
    if not erp or not str(erp).strip().startswith("M-"):
        continue
    erp_code = str(erp).strip()
    model = str(row[COLS["model"]]).strip() if row[COLS["model"]] else ""
    spec = parse_wp(row[COLS["spec"]])
    beginning = to_int(row[COLS["beginning"]])
    inbound = to_int(row[COLS["inbound"]])
    outbound = to_int(row[COLS["outbound"]])
    ending = to_int(row[COLS["ending"]])
    safety = to_int(row[COLS["safety"]])
    available = to_int(row[COLS["available"]])
    factor = to_float(row[COLS["factor"]])

    # product_id 결정
    pid = prod_by_erp.get(erp_code)
    if not pid:
        # PR 19 후에도 매칭 안 됨 — 모델명으로 한 번 더 시도
        norm = normalize_code(model)
        pid = prod_by_code.get(norm)
        if pid:
            # erp_code 미등록인 product 발견 — UPDATE
            try:
                cur.execute("UPDATE products SET erp_code = %s WHERE product_id = %s AND erp_code IS NULL",
                            (erp_code, pid))
                if cur.rowcount > 0:
                    prod_by_erp[erp_code] = pid
            except Exception as e:
                errors.append(f"row={ridx} erp_code 매칭 update 실패: {e}")
                c.rollback()
                continue
        else:
            # 신규 product 등록
            mfg_id = infer_mfg(model)
            try:
                cur.execute("""
                INSERT INTO products (product_code, product_name, manufacturer_id, spec_wp, wattage_kw, is_active, erp_code)
                VALUES (%s, %s, %s, %s, %s, true, %s) RETURNING product_id
                """, (model[:30], model[:100], mfg_id, spec, (spec/1000.0 if spec else None), erp_code))
                pid = cur.fetchone()[0]
                new_products += 1
                prod_by_erp[erp_code] = pid
                prod_by_code[normalize_code(model)] = pid
            except Exception as e:
                errors.append(f"row={ridx} 신규 product 실패 ({erp_code}, {model}): {e}")
                c.rollback()
                continue

    # products 의 safety/available 동기화 (ERP 신뢰)
    try:
        cur.execute("UPDATE products SET safety_stock = %s, available_stock = %s WHERE product_id = %s",
                    (safety, available, pid))
        products_synced += cur.rowcount
    except Exception as e:
        errors.append(f"row={ridx} products UPDATE 실패: {e}")
        c.rollback()
        continue

    # inventory_snapshots upsert
    payload = json.dumps({
        "model": model, "spec": spec, "unit": str(row[COLS["unit"]]) if row[COLS["unit"]] else None,
        "ending_mgmt": to_int(row[COLS["ending_mgmt"]]),
    }, ensure_ascii=False)
    try:
        cur.execute("""
        INSERT INTO inventory_snapshots
          (snapshot_date, product_id, beginning_qty, inbound_qty, outbound_qty, ending_qty,
           safety_qty, available_qty, unit_factor, source, source_payload)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'erp_export', %s::jsonb)
        ON CONFLICT (snapshot_date, product_id, source) DO UPDATE SET
          beginning_qty = EXCLUDED.beginning_qty,
          inbound_qty = EXCLUDED.inbound_qty,
          outbound_qty = EXCLUDED.outbound_qty,
          ending_qty = EXCLUDED.ending_qty,
          safety_qty = EXCLUDED.safety_qty,
          available_qty = EXCLUDED.available_qty,
          unit_factor = EXCLUDED.unit_factor,
          source_payload = EXCLUDED.source_payload
        RETURNING (xmax = 0) AS inserted
        """, (SNAPSHOT_DATE, pid, beginning, inbound, outbound, ending,
              safety, available, factor, payload))
        is_new = cur.fetchone()[0]
        if is_new: snapshot_inserted += 1
        else: snapshot_updated += 1
    except Exception as e:
        errors.append(f"row={ridx} snapshot 실패: {e}")
        c.rollback()
        continue

c.commit()

# 검증
cur.execute("SELECT count(*) FROM inventory_snapshots WHERE source = 'erp_export'")
total_snapshots = cur.fetchone()[0]
cur.execute("SELECT count(*) FROM products WHERE safety_stock IS NOT NULL OR available_stock IS NOT NULL")
synced_products = cur.fetchone()[0]

print(f"\n=== 결과 ===")
print(f"snapshot 신규: {snapshot_inserted}")
print(f"snapshot UPDATE: {snapshot_updated}")
print(f"products safety/available 동기화: {products_synced}")
print(f"신규 product 등록 (재고 시트만 있던 모델): {new_products}")
print(f"errors: {len(errors)}")
for e in errors[:5]: print(f"  {e}")
print(f"\n=== 누적 ===")
print(f"전체 inventory_snapshots: {total_snapshots}")
print(f"products 중 재고 정보 있는 행: {synced_products}")
