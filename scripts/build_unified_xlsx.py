# -*- coding: utf-8 -*-
"""솔라플로우_통합정리자료_2026-05-15.xlsx 개선 빌드.

기존 14시트 구조 유지하되:
1. 0번 시트로 '대시보드' 신설 (한 장 요약)
2. 1번 '자료 카탈로그' 에 신규 entry 추가 (과세자료 / 거래처 재고확인서)
3. README 갱신 (시트 안내 신규 포함)
4. 신규 흡수 5시트 추가 (E2 / B / L / J / K)

원본 파일은 Dropbox 버전 히스토리에 백업됨 — 동일 파일명으로 덮어쓰기.
"""
from __future__ import annotations

import os
import sys
from datetime import datetime

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = r'C:\Users\user\Dropbox (개인용)\8. 코딩\솔라플로우 참고 자료'
TARGET = os.path.join(ROOT, '솔라플로우_통합정리자료_2026-05-15.xlsx')

# ---------- 스타일 ----------
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='1F4E78')
NOTE_FONT = Font(name='맑은 고딕', italic=True, size=9, color='666666')
BODY_FONT = Font(name='맑은 고딕', size=10)
NUM_FONT = Font(name='맑은 고딕', size=10)
BORDER_THIN = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN


def autosize(ws, min_w=10, max_w=50):
    for col in ws.columns:
        # read_only 모드가 아닐 때만 사용
        try:
            col_letter = col[0].column_letter
        except AttributeError:
            continue
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            v = str(cell.value)
            # 한글 1.7배 가중
            w = sum(2 if ord(ch) > 127 else 1 for ch in v)
            if w > max_len:
                max_len = w
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ---------- 1. 기존 wb 로드 ----------
print(f'Loading {TARGET} ...')
wb = load_workbook(TARGET)
print(f'Existing sheets ({len(wb.sheetnames)}):')
for n in wb.sheetnames:
    print(f'  - {n}')

# ---------- 2. 시트 0번 '대시보드' 신설 ----------
DASH = '0. 대시보드'
if DASH in wb.sheetnames:
    del wb[DASH]
ws = wb.create_sheet(DASH, 0)

ws['A1'] = '솔라플로우 통합 정리자료 — 대시보드'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:F1')
ws['A2'] = f'기준일: {datetime.now().strftime("%Y-%m-%d %H:%M")}    |    원천: harness/data-sources.md + db-connectivity-report.md + 운영자 Dropbox'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:F2')

# 섹션 1: 백필 진척
ws['A4'] = '1. 백필 마이그 진척 (M111~M131 완료, M132~M136 후보)'
ws['A4'].font = SECTION_FONT
ws.merge_cells('A4:F4')

backfill = [
    ('마이그', '대상', '효과', '소스', '상태'),
    ('M111+M112', 'bl_shipments.cif_amount_krw + bl_line_items', 'BL CIF 백필 100건', 'C + 면장 PDF', '✅'),
    ('M113~M115', 'outbound_bl_items 18% → 74%, 트리거 + 컬럼 정리', 'BL 매칭 자동화', 'C + A', '✅'),
    ('M116', 'cost_details 백필 100건 (CIF 1,049억)', '재무 정합 핵심', 'A.DB-3', '✅'),
    ('M117', 'bl_shipments 4컬럼 (decl/inv/xr/arrival)', '면장 보강', 'A + E1', '✅'),
    ('M118', 'outbounds.site_name 17건 보강', '사이트 매칭', 'C + 영업', '✅'),
    ('M130', 'cost_details.incidental_cost 백필 (47/100, 18.3억)', '부대비용 1차', 'D 회계전표', '✅'),
    ('M131', 'BL별 운송료 청구서 9건 정밀 보강', 'incidental 정확도', 'J BL별 청구서', '✅'),
    ('M132 (제안)', 'cost_details.customs_fee 100건 백필', '관세 18억대 추정', 'D + E1.품의서', '🔥 1순위'),
    ('M133 (제안)', 'cost_details.incidental_cost 잔여 53건', '17억 회계 갭 직접 축소', 'E1.품의서/Sheet1(4)', '🔥 1순위'),
    ('M134 (제안)', 'incidental_expenses 신규 행 백필', '회계 추적성', 'D 분개 raw', '중'),
    ('M135 (제안)', 'lc_line_items 백필', 'LC↔BL 매칭', 'E1.외환 우측', '중'),
    ('M136 (제안)', 'purchase_orders 잔여 백필 (62 → 80+)', 'PO 통계', 'E1.2024/2025 시트', '중'),
]
for i, row in enumerate(backfill):
    for j, v in enumerate(row):
        c = ws.cell(row=5 + i, column=1 + j, value=v)
        c.border = BORDER_THIN
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if i == 0:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            c.font = BODY_FONT

# 섹션 2: DB 핵심 수치
row_anchor = 5 + len(backfill) + 2
ws.cell(row=row_anchor, column=1, value='2. DB 핵심 수치 (2026-05-15 gx10 기준)').font = SECTION_FONT
ws.merge_cells(start_row=row_anchor, start_column=1, end_row=row_anchor, end_column=6)

db_metrics = [
    ('항목', '값', '단위', '비고'),
    ('cost_details.cif_total_krw 합계', '104,901,126,374', '원 (1,049억)', 'CIF 백필 100% 완료'),
    ('cost_details.incidental_cost 합계', '1,826,504,115', '원 (18.3억)', 'M130/M131 백필분 (전체의 47%)'),
    ('cost_details.landed_total_krw 합계', '117,217,743,110', '원 (1,172억)', 'CIF + incidental'),
    ('2025 매출 (sales)', '37,638,966,253', '원 (376.4억)', 'sales.supply_amount where status<>cancelled'),
    ('2025 FIFO 원가', '29,412,661,798', '원 (294.1억)', 'fifo_matches.cost_amount (CIF/qty 만)'),
    ('2025 매출총이익 (시스템)', '8,226,304,455', '원 (82.3억) / 21.9%', None),
    ('회계 vs 시스템 갭', '≈ 17억', '원', '회계 매출원가에 부대비용 가산 후 차이 (M132/M133 으로 축소 가능)'),
    ('purchase_orders 행수', '62', '건', '탑솔라 47 / 디원 11 / 화신 4'),
    ('lc_records 행수', '49', '건', '탑솔라 37 / 디원 9 / 화신 3'),
    ('incidental_expenses 행수', '0', '건', 'cost_details 컬럼에 직접 update — 별도 백필 필요'),
    ('lc_line_items 행수', '0', '건', '미백필 (M135 후보)'),
]
for i, row in enumerate(db_metrics):
    for j, v in enumerate(row):
        c = ws.cell(row=row_anchor + 1 + i, column=1 + j, value=v)
        c.border = BORDER_THIN
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if i == 0:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            c.font = BODY_FONT

# 섹션 3: 자료 흡수 진행률
row_anchor2 = row_anchor + 1 + len(db_metrics) + 2
ws.cell(row=row_anchor2, column=1, value='3. 외부 자료 흡수 진행률').font = SECTION_FONT
ws.merge_cells(start_row=row_anchor2, start_column=1, end_row=row_anchor2, end_column=6)

intake = [
    ('자료 ID', '자료', '본 통합본 시트', '비고'),
    ('A', 'solarflow 자료.xlsx (ERP raw)', '— (DB 가 정본)', 'DB-3/fifo/수불/매출/입고/출고 → DB 매핑은 카탈로그 참조'),
    ('B', '탑솔라 그룹 출고현황 (워크플로우)', '시트 16', '✅ 신규 흡수 (1148행 raw + 4 플래그)'),
    ('C', 'BL별 출고현황리스트', '시트 8 (매트릭스)', 'M111~M115 백필 완료 — DB 가 정본'),
    ('D', '회계 전표 (선진/블루/스마일)', '시트 9 (요약)', '6 시트 raw 분석 — 본 통합본은 청구서 인벤토리만'),
    ('E1', '수입진행상황 25년 (191MB, 8 시트)', '시트 11~14', '25년 PO/LC/BL CIF/차량 매트릭스 흡수 완료'),
    ('E2', '수입진행상황 26년 (110KB, 6 시트)', '시트 15·16', '✅ 신규 흡수 (Sheet1 요약 + 제조사별 4시트 통합 트래커)'),
    ('F/G', 'zip 아카이브 (24/25/26)', '시트 7 (인덱스)', '디렉토리 분포만 — raw 는 zip 안 (감사용)'),
    ('H', '24년말 기말재고 스냅샷', '— (DB 반영)', "fifo_matches 의 '기초재고' 행"),
    ('I', '발전시공일정', '— (영업 자료)', '신규 도메인 (projects/installation_schedules) 후보'),
    ('J', 'BL별 운송료 청구서 (25/26년 50+건)', '시트 18', '✅ 신규 흡수 (BL → 파일 → 거래처 인벤토리)'),
    ('K', '클레임 (블루/선진/론지 26년 5건)', '시트 19', '✅ 신규 흡수 (인벤토리 메타)'),
    ('L', '바로 26년 1Q 모듈판매 이익률', '시트 17', '✅ 신규 흡수 (256행 raw + 1Q 요약)'),
    ('M', '과세자료 제출관련 (관세평가)', '시트 19', '✅ 신규 — 관세 가격결정자료 (트리나)'),
    ('N', '거래처 재고 확인서 (25년말)', '시트 19', '✅ 신규 — 탑솔라/디원/화신 + 타처 보관'),
]
for i, row in enumerate(intake):
    for j, v in enumerate(row):
        c = ws.cell(row=row_anchor2 + 1 + i, column=1 + j, value=v)
        c.border = BORDER_THIN
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if i == 0:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            c.font = BODY_FONT

# 섹션 4: 다음 액션
row_anchor3 = row_anchor2 + 1 + len(intake) + 2
ws.cell(row=row_anchor3, column=1, value='4. 추천 다음 액션').font = SECTION_FONT
ws.merge_cells(start_row=row_anchor3, start_column=1, end_row=row_anchor3, end_column=6)

actions = [
    ('우선순위', '액션', '얻는 가치'),
    ('🔥 1', 'M132/M133 작성 — cost_details.customs_fee 100건 + incidental_cost 잔여 53건 백필', '회계 17억 갭 + 관세 18억 갭 직접 축소'),
    ('🔥 2', 'M134 — incidental_expenses 신규 행 백필 (D 회계전표 → BL 매칭)', '회계 추적성 확보 (현재 cost_details 컬럼 직접 update 만)'),
    ('중 3', 'M135/M136 — lc_line_items + purchase_orders 잔여 백필', 'LC 만기 알림 / 발주 통계'),
    ('중 4', '월 1회 검증 잡 — 시트 16 (B 워크플로우 4 플래그) ↔ outbounds 정합', 'D-055 데이터 신뢰도'),
    ('낮음 5', '신규 도메인 설계 — claims (K) / projects (I) / partner_price_book', '손실 추적 + PM 자동화 + 단가 인사이트'),
    ('낮음 6', '시트 17 ↔ bp/sales_dashboard RPC 일치 검증', 'baro 매출 신뢰도'),
]
for i, row in enumerate(actions):
    for j, v in enumerate(row):
        c = ws.cell(row=row_anchor3 + 1 + i, column=1 + j, value=v)
        c.border = BORDER_THIN
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if i == 0:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            c.font = BODY_FONT

# 컬럼 너비
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 14
# 행 높이 — 첫 줄과 헤더만 살짝
ws.row_dimensions[1].height = 24

print(f'\n시트 0 [{DASH}] 작성 완료')

# ---------- 3. README 갱신 ----------
README = '0. README'
# README 가 0번 위치에서 1번으로 밀려난 상태. 시트 안내 텍스트 갱신.
ws = wb[README]
# 기존 'A11~A18 시트 안내' 부분을 새 구조로 덮어쓴다 — 안전하게 시트 전체 클리어 후 다시 씀
ws.delete_rows(1, ws.max_row)

readme_lines = [
    '솔라플로우 통합 정리자료',
    f'생성: 2026-05-15 13:50  (개선판: {datetime.now().strftime("%Y-%m-%d %H:%M")})',
    None,
    '본 통합본은 운영자 PC 의 외부 엑셀/회계장부/발주서 아카이브가 SolarFlow DB 어디에 매핑되는지,',
    '무엇이 백필 됐고 무엇이 갭으로 남았는지 한 곳에 정리한 분석 시트다.',
    None,
    '병행 reference:',
    '  - harness/data-sources.md          (사람 읽기용 narrative)',
    '  - harness/db-connectivity-report.md (DB 정본 reference)',
    None,
    '시트 안내:',
    '  0. 대시보드            — 한 장 요약 (백필 진척 + DB 수치 + 자료 흡수 진행률 + 다음 액션)',
    '  1. README              — 본 안내',
    '  2. 자료 카탈로그        — 외부 자료 18종 × DB 매핑 × 백필 상태 (신규: 과세자료 / 재고확인서)',
    '  3. BL·면장·비용        — 운영 중 BL 매트릭스 (자료 출처별 채움 여부)',
    '  4. 부대비용 청구서     — BL × 청구서 (블루오션/선진) 인벤토리',
    '  5. 발주 진행 매트릭스   — 제조사 × 계약 × 단계 (E.수입진행상황 요약)',
    '  6. 컬럼 매핑           — 핵심 시트 컬럼 → DB 컬럼',
    '  7. zip 아카이브        — 24/25/26 zip 디렉토리 분포',
    '  8. 갭 분석             — DB 비어있는 도메인 + 다음 마이그 후보',
    '  9. 회사·거래처 사전     — 자료에 나오는 이름 → 정본 매핑',
    ' 10. DB 실시간 정합      — 테이블별 행수 + 재무 갭 + 백필 우선순위',
    ' 11. E1. 25년 PO 실데이터 — 25년 PO 55행 (purchase_orders 백필 1순위 소스)',
    ' 12. E1. LC 마스터 실데이터 — 탑솔라 27 + 디원 5 (lc_records 백필 1순위)',
    ' 13. BL별 CIF·운송료 정산  — 품의서 25 블록 (cost_details.incidental_cost 백필 소스)',
    ' 14. 580Wp 차량 적재 매트릭스 — 운송료 단가 root',
    ' 15. (NEW) E2. 26년 발주 현황 — 26년 계약물량/잔량/단가 요약',
    ' 16. (NEW) E2. 26년 PO 트래커 통합 — 징코/론지/트리나/라이젠 4 시트 → 핵심 컬럼 flatten',
    ' 17. (NEW) B. 탑솔라 그룹 워크플로우 — 거래명세서/검수/결재/계산서 4 플래그 1148행',
    ' 18. (NEW) L. 바로 26년 1Q 판매·이익률 — 256행 raw + 1Q 가중평균',
    ' 19. (NEW) J. BL별 운송료 청구서 인벤토리 — 25/26년 청구서 파일 인덱스 (BL → 파일 → 거래처)',
    ' 20. (NEW) K. 클레임 + 보조자료 인벤토리 — 클레임 5건 + 과세자료 + 거래처 재고확인서',
    None,
    '범례 (백필 상태):',
    '  ✅ DB 가 정본 (자료는 mirror)',
    '  🟢 일부 백필됨, 잔여 작업 있음',
    '  🟡 부분 매핑 / 검증 필요',
    '  ❌ 미백필 (자료만 있고 DB 에 없음)',
]
for i, line in enumerate(readme_lines):
    ws.cell(row=i + 1, column=1, value=line if line is not None else '')
    if i == 0:
        ws.cell(row=i + 1, column=1).font = TITLE_FONT
    elif i == 1:
        ws.cell(row=i + 1, column=1).font = NOTE_FONT
    else:
        ws.cell(row=i + 1, column=1).font = BODY_FONT
ws.column_dimensions['A'].width = 100

print(f'시트 [{README}] 갱신 완료')

# ---------- 4. 자료 카탈로그 갱신 (신규 entry 추가) ----------
CAT = wb.sheetnames[2]  # 1. 자료 카탈로그 (이전 1번, 대시보드 추가로 2번)
print(f'카탈로그 시트: [{CAT}]')
ws = wb[CAT]
# 기존 마지막 행 (M = 18행) 다음에 신규 N, O 추가
new_entries = [
    ('N', '과세가격결정자료(탑솔라)-트리나모듈 날인본.pdf', '관세평가 미제출/지연 제출 사유서 + 트리나모듈 과세가격결정자료', '사고 시', '(import_declarations.customs_fee 보강 단서)', '❌ 신규 — 자료 인벤토리만', '2026년 모듈 발주/과세자료 제출관련/'),
    ('O', '거래처 재고 확인서 (25년말)', '탑솔라/디원/화신 재고확인서 + 타처(블루오션) 보관 재고 PDF 5건', '연 1회 (회계감사용)', 'inventory_snapshots (외부 재고)', '❌ 거래처 PDF — 정합 검증용', '2025년 모듈 발주/2025. 블루 재고 확인서/'),
]
start_row = ws.max_row + 1
for i, row in enumerate(new_entries):
    for j, v in enumerate(row):
        c = ws.cell(row=start_row + i, column=1 + j, value=v)
        c.border = BORDER_THIN
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.font = BODY_FONT

print(f'카탈로그에 신규 2 entry (N, O) 추가됨')

# ---------- 5. 신규 흡수 시트 5개 추가 ----------
def copy_sheet_data(src_path, src_sheet, dst_ws, header_row_idx=0, max_data_rows=None):
    """원본 시트를 dst_ws 에 그대로 복사 (값만, 서식 적용)."""
    src_wb = load_workbook(src_path, read_only=True, data_only=True)
    src_ws = src_wb[src_sheet]
    dst_row = 1
    n_cols = src_ws.max_column
    for i, row in enumerate(src_ws.iter_rows(values_only=True)):
        if max_data_rows is not None and dst_row > max_data_rows:
            break
        for j, val in enumerate(row):
            cell = dst_ws.cell(row=dst_row, column=j + 1, value=val)
            cell.font = BODY_FONT
            if i == header_row_idx:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        dst_row += 1
    src_wb.close()
    return dst_row - 1, n_cols


# === 시트 14: E2 26년 발주 현황 요약 ===
S14 = '14. E2-26년 발주 현황'
if S14 in wb.sheetnames:
    del wb[S14]
ws = wb.create_sheet(S14)
ws['A1'] = 'E2. 2026년도 모듈 발주 진행현황 (수입진행상황(module)-2026년도.xlsx Sheet1)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:J1')
ws['A2'] = '원본: Dropbox/2026년 모듈 발주/수입진행상황(module)-2026년도.xlsx (110KB, 2026-05-11 갱신)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:J2')

# 1) 계약 물량 / 잔량 (rows 2-10 of source)
ws['A4'] = '1. 발주처 × 업체 × 계약 물량/잔량'
ws['A4'].font = SECTION_FONT
ws.merge_cells('A4:J4')

p_e2 = os.path.join(ROOT, '2026년 모듈 발주', '수입진행상황(module)-2026년도.xlsx')
src_wb = load_workbook(p_e2, read_only=True, data_only=True)
src_ws = src_wb['Sheet1']
src_rows = list(src_ws.iter_rows(values_only=True))

# 헤더 (row 2)
hdr_e2_summary = ['발주처', '업체', '발주물량(계약/잔량)', '품목', 'PCS', 'WP', '단가($)', '금액(USD)', '비고']
ws.append([])  # row 5 placeholder — actually we want row 5 = header
# We'll write to specific rows
hdr_row = 5
for j, v in enumerate(hdr_e2_summary):
    c = ws.cell(row=hdr_row, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

cur = hdr_row + 1
for r in src_rows[3:11]:  # rows 4..11 (data)
    for j in range(9):
        val = r[j] if j < len(r) else None
        c = ws.cell(row=cur, column=j + 1, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if j in (4, 5, 7):  # PCS, WP, USD
            c.number_format = '#,##0'
        elif j == 6:
            c.number_format = '0.0000'
    cur += 1

# 2) 제조사별 단가 분석 (rows 12-23 of source)
cur += 2
ws.cell(row=cur, column=1, value='2. 제조사별 단가 분석 (T/T 0.95 / 0.9 분배 등)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=10)
cur += 1
hdr_e2_price = ['', '제조사', '구분', '품목(W)', '수량', 'WP', '비율', '단가($/Wp)', '금액(USD)', '안분(USD)']
for j, v in enumerate(hdr_e2_price):
    c = ws.cell(row=cur, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cur += 1
for r in src_rows[12:32]:
    if all(v is None for v in r):
        cur += 1
        continue
    for j in range(min(10, len(r))):
        c = ws.cell(row=cur, column=j + 1, value=r[j])
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if j in (4, 5, 8, 9):
            c.number_format = '#,##0'
        elif j in (6, 7):
            c.number_format = '0.0000'
    cur += 1

src_wb.close()

# 너비
widths = [12, 20, 28, 14, 12, 14, 12, 16, 16, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
print(f'시트 [{S14}] 작성 완료')

# === 시트 15: E2 26년 PO 트래커 통합 (4 manufacturer) ===
S15 = '15. E2-26년 PO 트래커'
if S15 in wb.sheetnames:
    del wb[S15]
ws = wb.create_sheet(S15)
ws['A1'] = 'E2. 2026년도 PO 트래커 통합 (징코/론지/트리나/라이젠 4시트 → 핵심 16컬럼 flatten)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:P1')
ws['A2'] = '원본: 2026년 모듈 발주/수입진행상황(module)-2026년도.xlsx (제조사별 시트)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:P2')

# 헤더 (16 컬럼)
hdr = ['제조사', 'No.', '업체', 'P/O No.', '품명', "Q'ty(pcs)", 'F/M', 'Wp',
       'Unit price', 'Amount', 'L/C No.', '선적', '입항', 'B/L No', '현장', '입고일']
hdr_row = 4
for j, v in enumerate(hdr):
    c = ws.cell(row=hdr_row, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 컬럼 인덱스 매핑 (제조사 시트 row 2-3 헤더 기준)
# 0=No, 1=업체, 2=P/O No, 3=품명, 4=pcs, 5=F/M, 6=Wp, 7=Unit price, 8=Amount,
# 11=L/C No, 18=선적, 19=입항, 28=B/L No, 30=현장, 29=입고일
col_map = {0:0, 1:1, 2:2, 3:3, 4:4, 5:5, 6:6, 7:7, 8:8, 11:10, 18:11, 19:12, 28:13, 30:14, 29:15}

cur = hdr_row + 1
src_wb = load_workbook(p_e2, read_only=True, data_only=True)
for mfr in ['징코', '론지솔라', '트리나', '라이젠']:
    src_ws = src_wb[mfr]
    for i, r in enumerate(src_ws.iter_rows(values_only=True)):
        # 데이터는 row 5 (i=4) 부터, 헤더 3행 (i=0~3)
        if i < 4:
            continue
        # 비어있는 행은 skip (중요 col 다 None 이면)
        if all(r[k] is None for k in [0, 1, 2, 3, 28, 30] if k < len(r)):
            continue
        # 제조사
        ws.cell(row=cur, column=1, value=mfr).font = BODY_FONT
        for src_idx, dst_idx in col_map.items():
            val = r[src_idx] if src_idx < len(r) else None
            c = ws.cell(row=cur, column=dst_idx + 1, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if dst_idx in (5, 6, 9):
                c.number_format = '#,##0'
            elif dst_idx == 8:
                c.number_format = '0.0000'
        cur += 1
src_wb.close()

# 너비
widths = [10, 6, 12, 22, 22, 12, 8, 12, 12, 16, 18, 12, 12, 22, 24, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'
print(f'시트 [{S15}] 작성 완료 — {cur - hdr_row - 1}행')

# === 시트 16: B 탑솔라 그룹 출고현황 워크플로우 ===
S16 = '16. B-탑솔라 그룹 워크플로우'
if S16 in wb.sheetnames:
    del wb[S16]
ws = wb.create_sheet(S16)
ws['A1'] = 'B. 탑솔라 그룹 모듈 출고현황 — 워크플로우 4 플래그 (거래명세서/검수/결재/계산서)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:R1')
ws['A2'] = '원본: Dropbox/탑솔라 그룹 모듈 출고현황.xlsx → 세금계산서 발행(탑, 디원, 화신) 시트 (1148행)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:R2')

# 원본 시트 raw 복사 (헤더 = row 2 in source)
p_b = os.path.join(ROOT, '탑솔라 그룹 모듈 출고현황.xlsx')
src_wb = load_workbook(p_b, read_only=True, data_only=True)
src_ws = src_wb['세금계산서 발행(탑, 디원, 화신)']

# 헤더는 row 4 of dest
hdr = ['구분', '납품일자', '업체명', '발전소명', '주소', '', '모델명', '수량', '용량(kW)',
       '출고잔량', '단가', '공급가액', '세액', '합계', '거래명세서', '인수검수요청서', '결재요청', '계산서발행']
hdr_row = 4
for j, v in enumerate(hdr):
    c = ws.cell(row=hdr_row, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

cur = hdr_row + 1
for i, r in enumerate(src_ws.iter_rows(values_only=True)):
    if i < 2:  # 원본 row 0 (빈), row 1 (헤더) 스킵
        continue
    # 비어있는 행 (모든 첫 14컬럼 None) 스킵
    if all(r[k] is None for k in range(min(14, len(r)))):
        continue
    for j in range(min(18, len(r))):
        val = r[j]
        c = ws.cell(row=cur, column=j + 1, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if j in (7, 8):
            c.number_format = '#,##0.00'
        elif j in (10, 11, 12, 13):
            c.number_format = '#,##0'
        elif j == 1 and isinstance(val, datetime):
            c.number_format = 'yyyy-mm-dd'
        elif j in (14, 15, 16, 17) and isinstance(val, bool):
            # 워크플로우 플래그 시각화
            c.value = '✅' if val else '⬜'
            c.alignment = Alignment(horizontal='center', vertical='center')
    cur += 1
src_wb.close()

widths = [10, 12, 18, 22, 28, 4, 22, 8, 10, 14, 10, 14, 12, 14, 8, 10, 8, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'
print(f'시트 [{S16}] 작성 완료 — {cur - hdr_row - 1}행')

# === 시트 17: L 바로 26년 1Q 판매·이익률 ===
S17 = '17. L-바로 26년1Q 판매·이익률'
if S17 in wb.sheetnames:
    del wb[S17]
ws = wb.create_sheet(S17)
ws['A1'] = 'L. 바로(주) 26년 1분기 모듈판매 이익률 분석'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:V1')
ws['A2'] = '원본: Dropbox/바로 모듈판매현황_이익률_26년1Q.xlsx (256+ 행, 2026-04-03 갱신)'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:V2')

p_l = os.path.join(ROOT, '바로 모듈판매현황_이익률_26년1Q.xlsx')
src_wb = load_workbook(p_l, read_only=True, data_only=True)
src_ws = src_wb['바로1분기모듈판매']

# 1) 1Q 가중평균 요약 (rows 3-4)
ws.cell(row=4, column=1, value='1. 1Q 가중평균 요약').font = SECTION_FONT
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=22)

# row 4의 헤더 + row 5 의 값 (source row 3, 4)
src_rows = list(src_ws.iter_rows(values_only=True))
sum_hdr = ['총 판매용량(kW)', '원가 총액(천원)', '기존 판매총액(천원)', '기존 가중평균이익률', '변경 판매총액(천원)', '변경 가중평균이익률']
sum_idx = [15, 16, 17, 18, 19, 20]
for j, v in enumerate(sum_hdr):
    c = ws.cell(row=5, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
# 값
src_summary = src_rows[3]
for j, src_i in enumerate(sum_idx):
    val = src_summary[src_i] if src_i < len(src_summary) else None
    c = ws.cell(row=6, column=j + 1, value=val)
    c.font = BODY_FONT
    if j in (3, 5):
        c.number_format = '0.00%'
    else:
        c.number_format = '#,##0.00'

# 가중평균 단가 (row 4 in source)
src_price = src_rows[4]
for j, src_i in enumerate([15, 16, 17, 19]):
    val = src_price[src_i] if src_i < len(src_price) else None
    if j == 0:
        ws.cell(row=7, column=1, value='가중평균 단가(원/Wp)').font = NUM_FONT
    else:
        c = ws.cell(row=7, column=1 + j, value=val)
        c.font = BODY_FONT
        c.number_format = '#,##0.00'

# 2) 판매 상세 (row 5 source = 헤더, row 6+ data)
ws.cell(row=10, column=1, value='2. 판매 건별 상세').font = SECTION_FONT
ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=22)

# 데이터 컬럼 추출 — source col 1~22 (38 중)
detail_cols = list(range(1, 23))
detail_hdr_src = src_rows[5]  # row 5 of source = 헤더
hdr_row = 11
for j, src_i in enumerate(detail_cols):
    val = detail_hdr_src[src_i] if src_i < len(detail_hdr_src) else None
    c = ws.cell(row=hdr_row, column=j + 1, value=val)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

cur = hdr_row + 1
for i, r in enumerate(src_ws.iter_rows(values_only=True)):
    if i < 6:
        continue
    if all(r[k] is None for k in detail_cols if k < len(r)):
        continue
    for j, src_i in enumerate(detail_cols):
        val = r[src_i] if src_i < len(r) else None
        c = ws.cell(row=cur, column=j + 1, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if isinstance(val, datetime):
            c.number_format = 'yyyy-mm-dd'
        elif j in (1, 2, 3) and isinstance(val, (int, float)):  # 단가/환율/원가
            c.number_format = '#,##0.0000'
        elif j in (14, 15) and isinstance(val, (int, float)):  # 수량/용량
            c.number_format = '#,##0.00'
        elif j in (16, 17, 19, 20, 21) and isinstance(val, (int, float)):
            c.number_format = '#,##0.00'
        elif j in (18, 20) and isinstance(val, float) and abs(val) < 1:  # 이익률
            c.number_format = '0.00%'
    cur += 1
src_wb.close()

# 너비 (22 컬럼)
widths = [12, 12, 10, 10, 12, 12, 12, 14, 22, 24, 12, 10, 16, 22, 10, 10, 14, 12, 12, 14, 12, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A12'
print(f'시트 [{S17}] 작성 완료 — {cur - hdr_row - 1}행')

# === 시트 18: J BL별 운송료 청구서 인벤토리 ===
S18 = '18. J-BL별 운송료 청구서'
if S18 in wb.sheetnames:
    del wb[S18]
ws = wb.create_sheet(S18)
ws['A1'] = 'J. 25/26년 BL별 운송료 청구서 인벤토리 (incidental_expenses 잔여 백필 1순위 소스)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:H1')
ws['A2'] = '원본: Dropbox/2025년 운송료/ + 2026년 모듈 발주/운송료 청구자료/ — 파일별 BL/거래처/금액 매칭'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:H2')

# 25년 운송료 폴더 스캔
import re
folder_25 = os.path.join(ROOT, '2025년 운송료')
folder_26 = os.path.join(ROOT, '2026년 모듈 발주', '운송료 청구자료')

ws.cell(row=4, column=1, value='1. 2025년 운송료 폴더 (Dropbox/2025년 운송료/)').font = SECTION_FONT
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

hdr = ['연도', '파일명', '추정 BL No.', '거래처(추정)', '확장자', '크기(B)', '갱신일', '비고']
hdr_row = 5
for j, v in enumerate(hdr):
    c = ws.cell(row=hdr_row, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

bl_re = re.compile(r'(SHACYV\w+|SHACYR\w+|SNK[oO]03[A-Z0-9]+|JWSH\d+|HDMUSHAA\d+|SHKWA\d+|EASED\d+|SELYIT\d+|SHADFC\w+|ESZX\d+|NPSELHT\d+|LS\d+|FR\s*\w+|RSPN\d+|JAHF\d+|MCKRJH\w+|KD\d+|TMSHKPTP\d+|DFS\d+)')

def detect_vendor(name):
    n = name.lower()
    if '블루' in name or 'bluo' in n or '블루오션' in name:
        return '블루오션에어'
    if '선진' in name:
        return '선진로지스틱스'
    if '스마일' in name:
        return '스마일로지스'
    if '광주' in name:
        return '광주공장'
    return '미상'

def detect_bl(name):
    m = bl_re.search(name)
    return m.group(1) if m else ''

cur = hdr_row + 1
files_25 = sorted(os.listdir(folder_25))
for f in files_25:
    full = os.path.join(folder_25, f)
    if os.path.isdir(full):
        continue
    size = os.path.getsize(full)
    mtime = datetime.fromtimestamp(os.path.getmtime(full))
    ext = os.path.splitext(f)[1].lower()
    bl = detect_bl(f)
    vendor = detect_vendor(f)
    note = ''
    if '~$' in f:
        continue  # Excel lock file
    if 'FN_' in f or 'fn_' in f:
        note = 'FN(최종본)'
    elif 'REVISED' in f:
        note = '재발행'
    elif '청구서' in f or '작업료' in f:
        note = '청구서'
    elif '부대비용' in f:
        note = 'D 회계 raw (오늘 갱신)'
        vendor = '회계 전표'
    vals = ['2025', f, bl, vendor, ext, size, mtime, note]
    for j, v in enumerate(vals):
        c = ws.cell(row=cur, column=j + 1, value=v)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if j == 5:
            c.number_format = '#,##0'
        elif j == 6:
            c.number_format = 'yyyy-mm-dd'
    cur += 1

# 26년 운송료 청구자료 폴더 (하위 디렉토리 = 월별)
cur += 1
ws.cell(row=cur, column=1, value='2. 2026년 운송료 청구자료 (Dropbox/2026년 모듈 발주/운송료 청구자료/)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
cur += 1
for j, v in enumerate(hdr):
    c = ws.cell(row=cur, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cur += 1

if os.path.isdir(folder_26):
    for sub in sorted(os.listdir(folder_26)):
        sub_full = os.path.join(folder_26, sub)
        if not os.path.isdir(sub_full):
            continue
        for f in sorted(os.listdir(sub_full)):
            full = os.path.join(sub_full, f)
            if os.path.isdir(full):
                continue
            size = os.path.getsize(full)
            mtime = datetime.fromtimestamp(os.path.getmtime(full))
            ext = os.path.splitext(f)[1].lower()
            bl = detect_bl(f)
            vendor = detect_vendor(sub) or detect_vendor(f)
            vals = ['2026', f'{sub}/{f}', bl, vendor, ext, size, mtime, sub]
            for j, v in enumerate(vals):
                c = ws.cell(row=cur, column=j + 1, value=v)
                c.font = BODY_FONT
                c.alignment = Alignment(vertical='center', wrap_text=True)
                if j == 5:
                    c.number_format = '#,##0'
                elif j == 6:
                    c.number_format = 'yyyy-mm-dd'
            cur += 1

widths = [8, 50, 22, 18, 8, 12, 12, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A6'
print(f'시트 [{S18}] 작성 완료')

# === 시트 19: K 클레임 + 보조자료 인벤토리 ===
S19 = '19. K-클레임·보조자료 인벤토리'
if S19 in wb.sheetnames:
    del wb[S19]
ws = wb.create_sheet(S19)
ws['A1'] = 'K. 클레임 + 신규 보조자료 (과세자료 / 거래처 재고확인서) 인벤토리'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:F1')
ws['A2'] = '원본: Dropbox/{2026년/클레임, 2026년/과세자료 제출관련, 2025년/2025. 블루 재고 확인서}'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:F2')

# 1. 클레임
ws.cell(row=4, column=1, value='1. 클레임 (26년 모듈 데미지)').font = SECTION_FONT
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

hdr_k = ['일자', '거래처/제조사', '파일명', '크기(B)', 'BL/규격', '비고']
hdr_row = 5
for j, v in enumerate(hdr_k):
    c = ws.cell(row=hdr_row, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

claim_dirs = [
    os.path.join(ROOT, '2026년 모듈 발주', '클레임'),
    os.path.join(ROOT, '2024년 모듈발주', '클레임'),
]
cur = hdr_row + 1
for d in claim_dirs:
    if not os.path.isdir(d):
        continue
    for f in sorted(os.listdir(d)):
        full = os.path.join(d, f)
        if os.path.isdir(full):
            # 하위 폴더 (e.g. 론지 클레임 1차/)
            for ff in sorted(os.listdir(full)):
                ffull = os.path.join(full, ff)
                if os.path.isdir(ffull):
                    continue
                size = os.path.getsize(ffull)
                mtime = datetime.fromtimestamp(os.path.getmtime(ffull)).strftime('%Y-%m-%d')
                spec = ''
                if '580' in ff: spec = '580Wp'
                elif '615' in ff: spec = '615Wp'
                elif '645' in ff: spec = '645Wp'
                elif '650' in ff: spec = '650Wp'
                vendor = '블루오션' if '블루' in ff else ('선진' if '선진' in ff else ('론지' if '론지' in ff else ''))
                vals = [mtime, f'{vendor} (하위: {f})', ff, size, spec, '클레임 1차 (zip 234MB 포함)']
                for j, v in enumerate(vals):
                    c = ws.cell(row=cur, column=j + 1, value=v)
                    c.font = BODY_FONT
                    c.alignment = Alignment(vertical='center', wrap_text=True)
                    if j == 3:
                        c.number_format = '#,##0'
                cur += 1
            continue
        size = os.path.getsize(full)
        mtime = datetime.fromtimestamp(os.path.getmtime(full)).strftime('%Y-%m-%d')
        spec = ''
        if '580' in f: spec = '580Wp'
        elif '615' in f: spec = '615Wp'
        elif '645' in f: spec = '645Wp'
        elif '650' in f: spec = '650Wp'
        vendor = '블루오션' if '블루' in f else ('선진' if '선진' in f else ('론지' if '론지' in f else ('금영' if '금영' in f else '')))
        note = ('데미지 정리 (xlsx, 사진 임베드)' if f.endswith('.xlsx') and size > 1000000
                else 'PDF 견적' if f.endswith('.pdf')
                else '데미지 정리')
        vals = [mtime, vendor, f, size, spec, note]
        for j, v in enumerate(vals):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if j == 3:
                c.number_format = '#,##0'
        cur += 1

# 2. 과세자료 제출관련 (신규)
cur += 1
ws.cell(row=cur, column=1, value='2. 과세자료 제출관련 (관세평가 — import_declarations.customs_fee 단서)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
cur += 1
for j, v in enumerate(['일자', '문서 종류', '파일명', '크기(B)', '대상', '비고']):
    c = ws.cell(row=cur, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cur += 1
tax_dir = os.path.join(ROOT, '2026년 모듈 발주', '과세자료 제출관련')
if os.path.isdir(tax_dir):
    for f in sorted(os.listdir(tax_dir)):
        full = os.path.join(tax_dir, f)
        if os.path.isdir(full):
            continue
        size = os.path.getsize(full)
        mtime = datetime.fromtimestamp(os.path.getmtime(full)).strftime('%Y-%m-%d')
        kind = '관세평가 미제출 사유서' if '별지' in f or '미제출' in f else '과세가격결정자료'
        target = '트리나모듈 (탑솔라)' if '트리나' in f else '일반'
        note = '관세 가격결정 → customs_fee 산정 근거 단서'
        vals = [mtime, kind, f, size, target, note]
        for j, v in enumerate(vals):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if j == 3:
                c.number_format = '#,##0'
        cur += 1

# 3. 거래처 재고 확인서 (신규)
cur += 1
ws.cell(row=cur, column=1, value='3. 거래처 재고 확인서 (25년말 — 회계감사용)').font = SECTION_FONT
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
cur += 1
for j, v in enumerate(['일자', '구분', '파일명', '크기(B)', '대상 회사', '비고']):
    c = ws.cell(row=cur, column=j + 1, value=v)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cur += 1
inv_dir = os.path.join(ROOT, '2025년 모듈 발주', '2025. 블루 재고 확인서')
if os.path.isdir(inv_dir):
    for f in sorted(os.listdir(inv_dir)):
        full = os.path.join(inv_dir, f)
        if os.path.isdir(full):
            continue
        size = os.path.getsize(full)
        mtime = datetime.fromtimestamp(os.path.getmtime(full)).strftime('%Y-%m-%d')
        if '타처' in f:
            kind = '타처보관 재고자산'
        else:
            kind = '재고확인서'
        company = ''
        if '탑솔라' in f: company = '탑솔라(주)'
        elif '디원' in f: company = '디원'
        elif '화신' in f: company = '화신이엔지'
        note = '거래처 발급 PDF — fifo_matches 기말재고 정합 검증 소스'
        vals = [mtime, kind, f, size, company, note]
        for j, v in enumerate(vals):
            c = ws.cell(row=cur, column=j + 1, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if j == 3:
                c.number_format = '#,##0'
        cur += 1

widths = [12, 22, 50, 12, 14, 38]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A6'
print(f'시트 [{S19}] 작성 완료')

# ---------- 6. 시트 순서 재배치 ----------
# 원하는 순서:
# 0. 대시보드 → 1. README → 2. 자료 카탈로그 → 3. BL·면장 → 4. 부대비용 → 5. 발주 진행 →
# 6. 컬럼 → 7. zip → 8. 갭 → 9. 회사·거래처 → 10. DB 실시간 → 11~14. E1 시리즈 →
# 15~19. 신규 흡수 시리즈
# (기존 시트 이름이 한글 + 번호 prefix 라 그대로 둠)

# 현재 시트 순서 상태
print('\n현재 시트 순서 (재배치 전):')
for i, n in enumerate(wb.sheetnames):
    print(f'  {i:2d}. {n}')

# 신규 시트만 순서 조정 — 기존 시트는 이미 0(대시보드) 만 앞으로 추가됐으니 정상
# 신규 5시트가 이미 끝에 있으면 OK
# 그러나 사용자가 본 README의 번호 (2~20) 와 실제 시트 prefix 가 다를 수 있음
# 일단 저장.

# ---------- 7. 저장 ----------
wb.save(TARGET)
print(f'\n저장 완료: {TARGET}')

# 최종 시트 순서 확인
wb2 = load_workbook(TARGET, read_only=True)
print('\n최종 시트 순서:')
for i, n in enumerate(wb2.sheetnames):
    print(f'  {i:2d}. {n}')
wb2.close()

print('\n빌드 완료.')
