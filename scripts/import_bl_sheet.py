"""
1회성 import: (수입) BL별 출고현황 리스트 — multi-sheet (제조사별).
- 각 시트: 발주처/B/L/항구/포워더/ETD/ETA/모델명/모듈수량/용량 + 출고 분할 행 (forward-fill)
- BL 마스터 자동 등록 (dedup by bl_number)
- 출고 분할 행 → 자연키 매칭 (출고일+모델+출고지+출고수량) → 기존 outbound 매핑
- 매칭 안 되면 신규 outbound 자동 등록 (manufacturer/wattage 자동 추론은 PR 11 룰 활용)
- outbound_bl_items INSERT 로 BL→출고 N:1 연결
"""
import os, sys, re, json, datetime, openpyxl, psycopg2
from openpyxl import load_workbook

FILE = "/tmp/bl_inbound.xlsx"

# 시트명 → 제조사 매핑 (마스터의 short_name/name_kr 기준)
SHEET_MFG = {
    "진코솔라": "징코",
    "진코솔라 (2)": "징코",
    "JA솔라": "JA",
    "트리나솔라": "트리나",
    "라이젠에너지": "라이젠",
    "론지솔라": "론지",
}

# 발주처 → company short_name 또는 그대로 매칭
SELLER_MAP = {
    "탑솔라": "탑솔라(주)",
    "디원": "디원",
    "화신": "화신이엔지",
    "화신이엔지": "화신이엔지",
}

WP_TO_PRODUCT_PREFIX = {}  # heuristic 으로 product 자동 등록 시 사용

def normalize_corp(s):
    if not s:
        return ""
    out = str(s).replace("(주)", "").replace("㈜", "").replace("주식회사", "")
    chars = []
    for ch in out:
        c = ord(ch)
        if (ord('A') <= c <= ord('Z')) or (ord('0') <= c <= ord('9')):
            chars.append(ch)
        elif ord('a') <= c <= ord('z'):
            chars.append(chr(c - 32))
        elif 0xAC00 <= c <= 0xD7AF or 0x4E00 <= c <= 0x9FFF:
            chars.append(ch)
    return "".join(chars)

def normalize_code(s):
    if not s:
        return ""
    chars = []
    for ch in str(s):
        c = ord(ch)
        if (ord('A') <= c <= ord('Z')) or (ord('0') <= c <= ord('9')):
            chars.append(ch)
        elif ord('a') <= c <= ord('z'):
            chars.append(chr(c - 32))
        elif 0xAC00 <= c <= 0xD7AF or 0x4E00 <= c <= 0x9FFF:
            chars.append(ch)
    return "".join(chars)

def to_iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None

def to_int(v):
    if v in (None, ""):
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except:
        return None

def to_float(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", ""))
    except:
        return None

c = psycopg2.connect(os.environ["SUPABASE_DB_URL"])
c.autocommit = False
cur = c.cursor()

# 마스터 캐시
cur.execute("SELECT manufacturer_id, name_kr, short_name FROM manufacturers")
mfg_map = {}
for mid, kr, short in cur.fetchall():
    if kr: mfg_map[normalize_code(kr)] = mid
    if short: mfg_map[normalize_code(short)] = mid

cur.execute("SELECT company_id, company_code, company_name FROM companies")
comp_map = {}
for cid, code, name in cur.fetchall():
    if code: comp_map[normalize_corp(code)] = cid
    if name: comp_map[normalize_corp(name)] = cid

cur.execute("SELECT bl_id, bl_number FROM bl_shipments")
bl_idx = {bn: bid for bid, bn in cur.fetchall()}

cur.execute("SELECT product_id, product_code, wattage_kw FROM products")
prod_map = {}
for pid, code, w in cur.fetchall():
    prod_map[normalize_code(code)] = (pid, float(w) if w else 0)

cur.execute("SELECT warehouse_id, warehouse_name FROM warehouses LIMIT 1")
default_wh = cur.fetchone()[0]  # backup default

# 통계
new_bls = 0
matched_outbounds = 0
new_outbounds = 0
new_products = 0
errors = []

wb = load_workbook(FILE, data_only=True)
print(f"sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    mfg_kw = SHEET_MFG.get(sheet_name)
    if not mfg_kw:
        print(f"[skip] {sheet_name} — manufacturer 매핑 없음")
        continue
    mfg_id = mfg_map.get(normalize_code(mfg_kw))
    if not mfg_id:
        print(f"[skip] {sheet_name} — manufacturer {mfg_kw} 마스터 미존재")
        continue
    print(f"\n=== {sheet_name} (manufacturer={mfg_kw}) ===")

    rows = list(ws.iter_rows(values_only=True))
    # 헤더 행 자동 탐지
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        if row and any(cell == "B/L" for cell in row if cell):
            header_idx = i
            break
    if header_idx < 0:
        print(f"[skip] {sheet_name} — 헤더 행 미발견")
        continue
    header = list(rows[header_idx])
    H = {}
    for j, lab in enumerate(header):
        if lab is None: continue
        key = str(lab).strip()
        H[key] = j
    # alias 보정
    def col(name):
        return H.get(name, -1)

    company_col = col("발주처")
    bl_col = col("B/L")
    port_col = col("항구")
    fwd_col = col("포워더")
    etd_col = col("ETD")
    eta_col = col("ETA")
    model_col = col("모델명")
    mod_qty_col = col("모듈수량")
    cap_col = col("용량")
    out_date_col = col("출고일")
    out_site_col = col("출고지")
    region_col = col("지역")
    wp_col = col("WP")
    out_qty_col = col("출고수량")

    if -1 in (bl_col, model_col):
        print(f"[skip] {sheet_name} — 필수 컬럼 누락")
        continue

    # forward-fill: BL 정보 (회사·BL·항구·포워더·ETD·ETA·모델·모듈수량·용량) 가 첫 행에만 있음
    last_bl = {}
    sheet_bls = 0
    sheet_outbounds_matched = 0
    sheet_outbounds_new = 0

    for ridx, row in enumerate(rows[header_idx+1:], start=header_idx+2):
        # B/L 정보 갱신 — bl_col 값 있으면 새 BL row 시작
        if bl_col != -1 and row[bl_col]:
            last_bl = {
                "company": row[company_col] if company_col != -1 else None,
                "bl": str(row[bl_col]).strip(),
                "port": row[port_col] if port_col != -1 else None,
                "forwarder": row[fwd_col] if fwd_col != -1 else None,
                "etd": to_iso(row[etd_col]) if etd_col != -1 else None,
                "eta": to_iso(row[eta_col]) if eta_col != -1 else None,
                "model": str(row[model_col]).strip() if model_col != -1 and row[model_col] else "",
                "module_qty": to_int(row[mod_qty_col]) if mod_qty_col != -1 else None,
                "capacity": to_float(row[cap_col]) if cap_col != -1 else None,
            }

            # bls 테이블 INSERT (dedup by bl_number)
            if last_bl["bl"] not in bl_idx:
                comp_id = comp_map.get(normalize_corp(last_bl["company"] or ""))
                if not comp_id:
                    errors.append(f"{sheet_name} row={ridx} BL={last_bl['bl']} — company 매칭 실패: {last_bl['company']}")
                    continue
                try:
                    cur.execute("""
                    INSERT INTO bl_shipments (bl_number, company_id, manufacturer_id, inbound_type, currency, status, etd, eta, port, forwarder)
                    VALUES (%s, %s, %s, 'import', 'USD', 'arrived', %s, %s, %s, %s)
                    RETURNING bl_id
                    """, (last_bl["bl"], comp_id, mfg_id, last_bl["etd"], last_bl["eta"],
                          str(last_bl["port"]) if last_bl["port"] else None,
                          str(last_bl["forwarder"]) if last_bl["forwarder"] else None))
                    new_bl_id = cur.fetchone()[0]
                    bl_idx[last_bl["bl"]] = new_bl_id
                    new_bls += 1
                    sheet_bls += 1
                    c.commit()
                except Exception as e:
                    c.rollback()
                    errors.append(f"BL INSERT 실패 {last_bl['bl']}: {e}")
                    continue

        # 출고 행 — 출고일 + 출고수량 둘 다 있으면 처리
        if out_date_col == -1 or out_qty_col == -1:
            continue
        out_date = to_iso(row[out_date_col])
        out_qty = to_int(row[out_qty_col])
        if not out_date or not out_qty or out_qty <= 0:
            continue
        out_site = (str(row[out_site_col]).strip() if out_site_col != -1 and row[out_site_col] else None)
        out_region = (str(row[region_col]).strip() if region_col != -1 and row[region_col] else None)

        if not last_bl.get("model"):
            continue

        # product 매칭
        prod_meta = prod_map.get(normalize_code(last_bl["model"]))
        if not prod_meta:
            # 자동 등록
            try:
                wp = to_int(row[wp_col]) if wp_col != -1 else None
                wattage_kw = wp / 1000.0 if wp else None
                body = {
                    "product_code": last_bl["model"][:30],
                    "product_name": last_bl["model"][:100],
                    "manufacturer_id": mfg_id,
                    "is_active": True,
                }
                if wp:
                    body["spec_wp"] = wp
                if wattage_kw:
                    body["wattage_kw"] = wattage_kw
                cur.execute("""
                INSERT INTO products (product_code, product_name, manufacturer_id, spec_wp, wattage_kw, is_active)
                VALUES (%s, %s, %s, %s, %s, true) RETURNING product_id, wattage_kw
                """, (body["product_code"], body["product_name"], mfg_id, body.get("spec_wp"), body.get("wattage_kw")))
                pid, w = cur.fetchone()
                prod_meta = (pid, float(w) if w else 0)
                prod_map[normalize_code(last_bl["model"])] = prod_meta
                new_products += 1
                c.commit()
            except Exception as e:
                c.rollback()
                errors.append(f"product 등록 실패 {last_bl['model']}: {e}")
                continue

        prod_id, wattage = prod_meta

        # 자연키 매칭 — 출고일 + product_id + 출고수량
        cur.execute("""
        SELECT outbound_id FROM outbounds
        WHERE outbound_date = %s AND product_id = %s AND quantity = %s
        """, (out_date, prod_id, out_qty))
        candidates = cur.fetchall()

        if candidates:
            # 첫 매치에 BL 연결 (bl_id 비어있을 때만)
            ob_id = candidates[0][0]
            # outbound_bl_items 도 dedup 후 INSERT
            cur.execute("""
            SELECT 1 FROM outbound_bl_items WHERE outbound_id = %s AND bl_id = %s
            """, (ob_id, bl_idx[last_bl["bl"]]))
            if not cur.fetchone():
                cur.execute("""
                INSERT INTO outbound_bl_items (outbound_id, bl_id, quantity)
                VALUES (%s, %s, %s)
                """, (ob_id, bl_idx[last_bl["bl"]], out_qty))
            matched_outbounds += 1
            sheet_outbounds_matched += 1
            c.commit()
        else:
            # 신규 출고 등록 — 회사 = last_bl["company"], 창고 = default
            comp_id = comp_map.get(normalize_corp(last_bl["company"] or ""))
            if not comp_id:
                errors.append(f"{sheet_name} row={ridx} 출고 신규 등록 실패 — company 매칭 안됨")
                continue
            try:
                cap_kw = (out_qty * wattage) if wattage > 0 else None
                payload = json.dumps({
                    "source": "bl_sheet",
                    "sheet": sheet_name,
                    "sheet_row": ridx,
                    "wp": to_int(row[wp_col]) if wp_col != -1 else None,
                    "region": out_region,
                    "bl_id": str(bl_idx[last_bl["bl"]]),
                }, ensure_ascii=False)
                cur.execute("""
                INSERT INTO outbounds (
                  outbound_date, company_id, product_id, quantity, capacity_kw,
                  warehouse_id, usage_category, status, site_name, site_address,
                  source_payload
                ) VALUES (%s, %s, %s, %s, %s, %s, 'sale', 'active', %s, %s, %s::jsonb)
                RETURNING outbound_id
                """, (out_date, comp_id, prod_id, out_qty, cap_kw, default_wh,
                      out_site, out_region, payload))
                new_ob_id = cur.fetchone()[0]
                cur.execute("""
                INSERT INTO outbound_bl_items (outbound_id, bl_id, quantity)
                VALUES (%s, %s, %s)
                """, (new_ob_id, bl_idx[last_bl["bl"]], out_qty))
                new_outbounds += 1
                sheet_outbounds_new += 1
                c.commit()
            except Exception as e:
                c.rollback()
                errors.append(f"outbound 신규 INSERT 실패 row={ridx}: {e}")
                continue

    print(f"  BL 신규 {sheet_bls}, 기존 outbound 매칭 {sheet_outbounds_matched}, 신규 outbound {sheet_outbounds_new}")

print(f"\n========== 전체 결과 ==========")
print(f"BL 신규 등록: {new_bls}")
print(f"product 신규 등록: {new_products}")
print(f"기존 outbound BL 연결: {matched_outbounds}")
print(f"신규 outbound 등록: {new_outbounds}")
print(f"errors: {len(errors)}")
if errors[:5]:
    print("샘플 에러:")
    for e in errors[:5]:
        print(f"  {e}")
