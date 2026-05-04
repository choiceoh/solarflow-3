"""
탑솔라 그룹 모듈 출고현황 → SolarFlow 출고 업로드 양식 변환기

입력: 탑솔라 그룹의 카카오톡 공유 엑셀 (월별 누적, 섹션 구분, 자유 양식)
출력: SolarFlow OUTBOUND_FIELDS 형식 xlsx + 매출 보조 시트

사용:
    python convert_topsolar_outbound.py <입력.xlsx> [출력.xlsx]
"""

import re
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# 구분 → seller company_code 매핑 (탑솔라 그룹 3사)
# NOTE: 실제 SolarFlow DB의 company_code 와 맞아야 함. 1차 cut: 한글 그대로.
SELLER_MAP = {
    "탑": "탑솔라",
    "탑솔라": "탑솔라",
    "디원": "디원",
    "화신": "화신이엔지",
    "화신이엔지": "화신이엔지",
}

# 구분에 섞이는 태그(부가 메모)
GUBUN_TAGS = ("외판", "단가확인", "확정", "차감", "9시착", "오후착", "오전착")

OUTBOUND_HEADERS = [
    "outbound_date", "company_code", "product_code", "quantity",
    "warehouse_code", "usage_category", "order_number", "site_name",
    "site_address", "spare_qty", "group_trade", "target_company_code",
    "erp_outbound_no", "memo",
]
OUTBOUND_LABELS = [
    "출고일*", "법인코드*", "품번코드*", "수량*",
    "창고코드*", "용도*", "수주번호", "현장명",
    "현장주소", "스페어수량", "그룹거래(Y/N)", "상대법인코드",
    "ERP출고번호", "메모",
]

SALE_HEADERS = [
    "row_link", "customer_name", "unit_price_wp",
    "supply_amount", "vat", "total", "memo",
]
SALE_LABELS = [
    "출고행번호", "거래처명*", "Wp단가(원)*",
    "공급가액", "세액", "합계", "메모",
]


def normalize_seller(gubun: str, current_section: str | None) -> tuple[str | None, str]:
    """구분 셀과 섹션 컨텍스트로 판매 법인코드 결정. (code, note)"""
    raw = (gubun or "").strip()
    note_parts: list[str] = []

    if any(tag in raw for tag in GUBUN_TAGS):
        note_parts.append(raw)

    # 슬래시 분리: '탑 / 확정', '외판/탑'
    tokens = [t.strip() for t in re.split(r"[/]", raw) if t.strip()]
    seller = None
    for t in tokens:
        if t in SELLER_MAP:
            seller = SELLER_MAP[t]
            break
        for k, v in SELLER_MAP.items():
            if t.startswith(k):
                seller = v
                break
        if seller:
            break

    if not seller and current_section:
        seller = SELLER_MAP.get(current_section)

    return seller, " ".join(note_parts)


def parse_section_label(s: str) -> str | None:
    """'탑솔라 (1월)' → '탑솔라', '디원 (2월)' → '디원' 형태에서 회사명 추출."""
    if not isinstance(s, str):
        return None
    m = re.match(r"^\s*(탑솔라|디원|화신이엔지)\s*\(", s)
    return m.group(1) if m else None


def parse_date(v) -> tuple[str, str]:
    """(date_iso, note). 파싱 실패 시 date_iso=''."""
    if pd.isna(v):
        return "", ""
    if isinstance(v, (datetime, pd.Timestamp)):
        return v.strftime("%Y-%m-%d"), ""
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d"), ""
    s = str(v).strip()
    # '1/12 오후착', '4/30 오후', '1/19 오전' — 월/일만 추출하고 연도는 채울 수 없음
    m = re.match(r"^(\d{1,2})/(\d{1,2})", s)
    if m:
        return "", f"원본 날짜: {s}"
    return "", f"원본 날짜: {s}"


SPARE_PAT = re.compile(r"SP\s*(\d+)\s*EA", re.IGNORECASE)


def parse_spare(remarks) -> int | None:
    if pd.isna(remarks):
        return None
    m = SPARE_PAT.search(str(remarks))
    return int(m.group(1)) if m else None


def to_int(v) -> int | None:
    if pd.isna(v):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def to_float(v) -> float | None:
    if pd.isna(v):
        return None
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


def is_skip_row(row: pd.Series) -> bool:
    """헤더 반복/합계/빈 행."""
    g = str(row[0]).strip() if pd.notna(row[0]) else ""
    if g in ("구분", "합 계", "합계"):
        return True
    # 모든 핵심 컬럼이 비어있으면 빈 행
    if all(pd.isna(row[i]) for i in (1, 2, 3, 6, 7)):
        return True
    return False


def convert(src: Path, dst: Path) -> dict:
    raw = pd.read_excel(src, header=None)

    out_rows: list[dict] = []
    sale_rows: list[dict] = []
    warnings: list[str] = []
    current_section: str | None = None

    for idx, row in raw.iterrows():
        # 섹션 마커 탐지
        section = parse_section_label(row[0]) if pd.notna(row[0]) else None
        if section:
            current_section = section
            continue

        if is_skip_row(row):
            continue

        out_no = len(out_rows) + 1
        seller, gubun_note = normalize_seller(
            str(row[0]) if pd.notna(row[0]) else "", current_section
        )

        date_iso, date_note = parse_date(row[1])

        product_code = str(row[6]).strip() if pd.notna(row[6]) else ""
        qty = to_int(row[7])
        order_number = str(row[5]).strip() if pd.notna(row[5]) else ""
        # 'BRB-260277-1 차감' 같이 부가 텍스트가 붙은 경우 코드만 분리
        m = re.match(r"^([A-Z]+-\d+(?:-\d+)?)", order_number)
        if m:
            order_code = m.group(1)
            extra = order_number[len(order_code):].strip()
        else:
            order_code = order_number
            extra = ""

        site_name = str(row[3]).strip() if pd.notna(row[3]) else ""
        site_addr = str(row[4]).strip() if pd.notna(row[4]) else ""
        remarks = str(row[9]).strip() if pd.notna(row[9]) else ""
        spare = parse_spare(remarks)

        memo_bits = [b for b in (gubun_note, date_note, extra, remarks) if b]
        memo = " | ".join(memo_bits)

        out_rows.append({
            "outbound_date": date_iso,
            "company_code": seller or "",
            "product_code": product_code,
            "quantity": qty if qty is not None else "",
            "warehouse_code": "",
            "usage_category": "sale",
            "order_number": order_code,
            "site_name": site_name,
            "site_address": site_addr,
            "spare_qty": spare if spare is not None else "",
            "group_trade": "",
            "target_company_code": "",
            "erp_outbound_no": "",
            "memo": memo,
        })

        # 행 누락 사유
        missing = []
        if not seller:
            missing.append("법인코드")
        if not date_iso:
            missing.append("출고일")
        if not product_code:
            missing.append("품번")
        if qty is None:
            missing.append("수량")
        if missing:
            warnings.append(
                f"엑셀 {idx + 1}행 → 변환 {out_no}행: 필수 누락 [{', '.join(missing)}]"
            )

        # 매출 보조 시트
        unit_price = to_float(row[10])
        supply = to_float(row[11])
        vat = to_float(row[12])
        total = to_float(row[13])
        customer = str(row[2]).strip() if pd.notna(row[2]) else ""
        if customer or unit_price or supply:
            sale_rows.append({
                "row_link": out_no,
                "customer_name": customer,
                "unit_price_wp": unit_price if unit_price is not None else "",
                "supply_amount": supply if supply is not None else "",
                "vat": vat if vat is not None else "",
                "total": total if total is not None else "",
                "memo": "",
            })

    write_workbook(dst, out_rows, sale_rows, warnings)
    return {
        "outbound_count": len(out_rows),
        "sale_count": len(sale_rows),
        "warning_count": len(warnings),
    }


def write_workbook(dst: Path, out_rows, sale_rows, warnings):
    wb = Workbook()

    ws = wb.active
    ws.title = "출고 업로드"
    write_sheet(ws, OUTBOUND_LABELS, OUTBOUND_HEADERS, out_rows)

    ws_sale = wb.create_sheet("매출 보조 (출고 후 사용)")
    write_sheet(ws_sale, SALE_LABELS, SALE_HEADERS, sale_rows)

    ws_warn = wb.create_sheet("변환 검토")
    ws_warn.append(["항목"])
    ws_warn["A1"].font = Font(bold=True)
    if warnings:
        for w in warnings:
            ws_warn.append([w])
    else:
        ws_warn.append(["경고 없음"])
    ws_warn.column_dimensions["A"].width = 80

    ws_help = wb.create_sheet("사용법")
    help_lines = [
        "[변환 결과 안내]",
        "",
        "1. '출고 업로드' 시트 → SolarFlow '출고 일괄 등록'에 그대로 업로드 가능한 양식.",
        "   * 표시 컬럼은 필수. 변환기는 데이터를 채울 수 있는 만큼 채우고, 비어있으면",
        "   '변환 검토' 시트에 사유를 남깁니다.",
        "",
        "2. 업로드 전 반드시 확인할 것:",
        "   - 법인코드 (탑솔라/디원/화신이엔지) 가 SolarFlow 마스터의 코드와 일치하는지",
        "   - 품번코드 (모델명) 가 제품 마스터의 품번과 일치하는지",
        "   - 창고코드 — 변환기는 비워둠. 수동으로 기본 창고 코드 입력 필요",
        "   - 용도 — 'sale' 로 기본 채움. 외판/공사/AS 등 다른 케이스면 수정",
        "",
        "3. '매출 보조 (출고 후 사용)' 시트:",
        "   - 출고 등록이 끝난 뒤 매출 일괄 등록에서 사용",
        "   - '출고행번호' 는 위 출고 시트의 몇 번째 행인지 가리킴",
        "   - 출고 등록 후 outbound_id 를 받아서 그 컬럼만 추가하면 매출 업로드 양식이 됨",
        "",
        "4. '변환 검토' 시트: 자동 매핑 실패한 행 목록. 빈 셀을 직접 채워서 업로드.",
    ]
    for i, line in enumerate(help_lines, 1):
        ws_help.cell(row=i, column=1, value=line)
    ws_help.column_dimensions["A"].width = 100

    wb.save(dst)


def write_sheet(ws, labels: list[str], keys: list[str], rows: list[dict]):
    header_fill = PatternFill("solid", start_color="DCE6F1")
    for i, (label, key) in enumerate(zip(labels, keys), 1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=i, value=key).font = Font(italic=True, color="888888")
    for r_i, row in enumerate(rows, 3):
        for c_i, key in enumerate(keys, 1):
            ws.cell(row=r_i, column=c_i, value=row.get(key, ""))
    for c_i in range(1, len(keys) + 1):
        ws.column_dimensions[get_column_letter(c_i)].width = 18
    ws.freeze_panes = "A3"


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_name(
        src.stem + "_변환됨.xlsx"
    )
    if not src.exists():
        print(f"입력 파일 없음: {src}")
        sys.exit(1)
    stats = convert(src, dst)
    print(f"출고 행: {stats['outbound_count']}")
    print(f"매출 보조 행: {stats['sale_count']}")
    print(f"검토 필요: {stats['warning_count']}")
    print(f"저장: {dst}")


if __name__ == "__main__":
    main()
